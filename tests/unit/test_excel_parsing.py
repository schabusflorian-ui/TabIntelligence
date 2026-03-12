"""
Unit tests for _excel_to_structured_repr() and _structured_to_markdown().

Tests run against the real sample_model.xlsx fixture to verify that
structured extraction preserves formulas, formatting, hierarchy, merged
cells, cell references, and sheet metadata.
"""

import io

import openpyxl
import pytest

from src.core.exceptions import InvalidFileError
from src.extraction.stages.parsing import ParsingStage

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_minimal_xlsx(
    data: dict | None = None,
    bold_cells: set | None = None,
    formulas: dict | None = None,
    merged_ranges: list | None = None,
    hidden_sheets: set | None = None,
    number_formats: dict | None = None,
    font_colors: dict | None = None,
    fill_colors: dict | None = None,
    borders: dict | None = None,
    comments: dict | None = None,
    very_hidden_sheets: set | None = None,
    underline_cells: set | None = None,
) -> bytes:
    """
    Create a tiny in-memory .xlsx from a dict of {sheet_name: [[cell, ...], ...]}.

    ``bold_cells``       – set of (sheet_name, row, col) tuples where font should be bold.
    ``formulas``         – dict of (sheet_name, row, col) -> formula string.
    ``merged_ranges``    – list of (sheet_name, range_string) to merge.
    ``hidden_sheets``    – set of sheet names to mark hidden.
    ``number_formats``   – dict of (sheet_name, row, col) -> format string.
    ``font_colors``      – dict of (sheet_name, row, col) -> "RRGGBB" hex color.
    ``fill_colors``      – dict of (sheet_name, row, col) -> "RRGGBB" hex color.
    ``borders``          – dict of (sheet_name, row, col) -> {"bottom": bool, "right": bool}.
    ``comments``         – dict of (sheet_name, row, col) -> comment text.
    ``very_hidden_sheets`` – set of sheet names to mark veryHidden.
    ``underline_cells``  – set of (sheet_name, row, col) tuples where font should be underlined.
    """
    from openpyxl.comments import Comment
    from openpyxl.styles import Border, Font, PatternFill, Side

    if data is None:
        data = {"Sheet1": [["A", "B"], [1, 2]]}
    bold_cells = bold_cells or set()
    formulas = formulas or {}
    merged_ranges = merged_ranges or []
    hidden_sheets = hidden_sheets or set()
    number_formats = number_formats or {}
    font_colors = font_colors or {}
    fill_colors = fill_colors or {}
    borders = borders or {}
    comments = comments or {}
    very_hidden_sheets = very_hidden_sheets or set()
    underline_cells = underline_cells or set()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name)
        if sheet_name in very_hidden_sheets:
            ws.sheet_state = "veryHidden"
        elif sheet_name in hidden_sheets:
            ws.sheet_state = "hidden"
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                key = (sheet_name, r_idx, c_idx)
                cell_obj = ws.cell(
                    row=r_idx,
                    column=c_idx,
                    value=formulas[key] if key in formulas else val,
                )
                # Consolidate font (bold + color + underline)
                is_bold = key in bold_cells
                is_underline = key in underline_cells
                color = font_colors.get(key)
                if is_bold or color or is_underline:
                    cell_obj.font = Font(
                        bold=is_bold,
                        color=color,
                        underline="single" if is_underline else None,
                    )
                if key in number_formats:
                    cell_obj.number_format = number_formats[key]
                if key in fill_colors:
                    hex_color = fill_colors[key]
                    cell_obj.fill = PatternFill(
                        start_color=hex_color,
                        end_color=hex_color,
                        fill_type="solid",
                    )
                if key in borders:
                    border_spec = borders[key]
                    cell_obj.border = Border(
                        bottom=Side(style="thin") if border_spec.get("bottom") else Side(),
                        right=Side(style="thin") if border_spec.get("right") else Side(),
                    )
                if key in comments:
                    cell_obj.comment = Comment(comments[key], "Test")
        for merge_sheet, merge_range in merged_ranges:
            if merge_sheet == sheet_name:
                ws.merge_cells(merge_range)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ===========================================================================
# Tests against sample_model.xlsx fixture
# ===========================================================================


class TestStructuredReprWithFixture:
    """Test _excel_to_structured_repr on the real sample_model.xlsx."""

    def test_top_level_keys(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        assert "sheets" in result
        assert "named_ranges" in result
        assert "sheet_count" in result
        assert "total_rows" in result

    def test_sheet_count(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        assert result["sheet_count"] == 4
        names = [s["sheet_name"] for s in result["sheets"]]
        assert "Income Statement" in names
        assert "Balance Sheet" in names
        assert "Cash Flow" in names
        assert "Scratch - Working" in names

    def test_total_rows_positive(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        assert result["total_rows"] > 0

    def test_sheet_has_required_keys(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        for sheet in result["sheets"]:
            assert "sheet_name" in sheet
            assert "is_hidden" in sheet
            assert "merged_regions" in sheet
            assert "rows" in sheet
            assert isinstance(sheet["is_hidden"], bool)
            assert isinstance(sheet["merged_regions"], list)

    def test_cell_has_required_keys(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        income = next(s for s in result["sheets"] if s["sheet_name"] == "Income Statement")
        assert len(income["rows"]) > 0
        first_row = income["rows"][0]
        assert "row_index" in first_row
        assert "cells" in first_row
        first_cell = first_row["cells"][0]
        for key in ("ref", "value", "formula", "is_bold", "indent_level", "number_format"):
            assert key in first_cell, f"Missing key '{key}' in cell dict"

    def test_cell_refs_are_valid(self, sample_xlsx):
        """Every cell ref must look like A1, B12, AA3 etc."""
        import re

        ref_pat = re.compile(r"^[A-Z]+\d+$")
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        for sheet in result["sheets"]:
            for row in sheet["rows"]:
                for cell in row["cells"]:
                    assert ref_pat.match(cell["ref"]), f"Invalid ref: {cell['ref']}"

    def test_revenue_row_present(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        income = next(s for s in result["sheets"] if s["sheet_name"] == "Income Statement")
        # Row 2 should contain "Revenue" in cell A2
        row2 = next((r for r in income["rows"] if r["row_index"] == 2), None)
        assert row2 is not None
        a2 = next((c for c in row2["cells"] if c["ref"] == "A2"), None)
        assert a2 is not None
        assert a2["value"] == "Revenue"

    def test_bold_header_detection(self, sample_xlsx):
        """Header row (FY2022, FY2023, FY2024E) should have bold=True."""
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        income = next(s for s in result["sheets"] if s["sheet_name"] == "Income Statement")
        row1 = next((r for r in income["rows"] if r["row_index"] == 1), None)
        assert row1 is not None
        bold_cells = [c for c in row1["cells"] if c["is_bold"]]
        assert len(bold_cells) >= 3, "FY2022/FY2023/FY2024E should be bold"

    def test_subtotal_detection(self, sample_xlsx):
        """'Total Operating Expenses' and 'Total Assets' should be marked as subtotal."""
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        income = next(s for s in result["sheets"] if s["sheet_name"] == "Income Statement")
        # Row 12: Total Operating Expenses
        row12 = next((r for r in income["rows"] if r["row_index"] == 12), None)
        assert row12 is not None
        a12 = next((c for c in row12["cells"] if c["ref"] == "A12"), None)
        assert a12 is not None
        assert a12.get("is_subtotal") is True

    def test_gross_profit_subtotal(self, sample_xlsx):
        """'Gross Profit' contains the word 'Gross' which matches the subtotal pattern."""
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        income = next(s for s in result["sheets"] if s["sheet_name"] == "Income Statement")
        row6 = next((r for r in income["rows"] if r["row_index"] == 6), None)
        assert row6 is not None
        a6 = next((c for c in row6["cells"] if c["ref"] == "A6"), None)
        assert a6 is not None
        assert a6.get("is_subtotal") is True

    def test_empty_rows_skipped(self, sample_xlsx):
        """Row 7 in Income Statement is empty and should not appear."""
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        income = next(s for s in result["sheets"] if s["sheet_name"] == "Income Statement")
        row_indices = {r["row_index"] for r in income["rows"]}
        assert 7 not in row_indices

    def test_named_ranges_is_dict(self, sample_xlsx):
        result = ParsingStage._excel_to_structured_repr(sample_xlsx)
        assert isinstance(result["named_ranges"], dict)


# ===========================================================================
# Tests with synthetic workbooks (edge cases)
# ===========================================================================


class TestStructuredReprSynthetic:
    """Test edge cases using in-memory workbooks."""

    def test_hidden_sheet_marked(self):
        xlsx = _make_minimal_xlsx(
            data={"Visible": [[1]], "Secret": [[2]]},
            hidden_sheets={"Secret"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        visible = next(s for s in result["sheets"] if s["sheet_name"] == "Visible")
        secret = next(s for s in result["sheets"] if s["sheet_name"] == "Secret")
        assert visible["is_hidden"] is False
        assert secret["is_hidden"] is True

    def test_merged_cells_detected(self):
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header", None, None], [1, 2, 3]]},
            merged_ranges=[("S1", "A1:C1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        s1 = result["sheets"][0]
        assert "A1:C1" in s1["merged_regions"]

    def test_formula_captured(self):
        xlsx = _make_minimal_xlsx(
            data={"S1": [[10, 20, None]]},
            formulas={("S1", 1, 3): "=A1+B1"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        s1 = result["sheets"][0]
        row1 = s1["rows"][0]
        c3 = next((c for c in row1["cells"] if c["ref"] == "C1"), None)
        assert c3 is not None
        assert c3["formula"] == "=A1+B1"

    def test_bold_formatting_preserved(self):
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header", "Normal"]]},
            bold_cells={("S1", 1, 1)},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row = result["sheets"][0]["rows"][0]
        a1 = next(c for c in row["cells"] if c["ref"] == "A1")
        b1 = next(c for c in row["cells"] if c["ref"] == "B1")
        assert a1["is_bold"] is True
        assert b1["is_bold"] is False

    def test_corrupted_file_raises_invalid_file_error(self):
        with pytest.raises(InvalidFileError):
            ParsingStage._excel_to_structured_repr(b"not an xlsx file at all")

    def test_empty_workbook(self):
        """A workbook with one sheet but no data should produce 0 rows."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = ParsingStage._excel_to_structured_repr(buf.read())
        assert result["sheet_count"] == 1
        assert result["total_rows"] == 0


# ===========================================================================
# Markdown conversion tests
# ===========================================================================


class TestStructuredToMarkdown:
    """Test _structured_to_markdown output."""

    def test_returns_string(self, sample_xlsx):
        structured = ParsingStage._excel_to_structured_repr(sample_xlsx)
        md = ParsingStage._structured_to_markdown(structured)
        assert isinstance(md, str)
        assert len(md) > 0

    def test_sheet_headers_present(self, sample_xlsx):
        structured = ParsingStage._excel_to_structured_repr(sample_xlsx)
        md = ParsingStage._structured_to_markdown(structured)
        assert "## Sheet: Income Statement" in md
        assert "## Sheet: Balance Sheet" in md
        assert "## Sheet: Cash Flow" in md

    def test_hidden_sheet_annotated(self):
        structured = {
            "sheets": [
                {
                    "sheet_name": "Hidden",
                    "is_hidden": True,
                    "merged_regions": [],
                    "rows": [],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 0,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "[HIDDEN]" in md

    def test_merged_regions_in_output(self):
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": ["A1:C1"],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "x",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "Merged: A1:C1" in md

    def test_markdown_table_pipe_delimited(self, sample_xlsx):
        structured = ParsingStage._excel_to_structured_repr(sample_xlsx)
        md = ParsingStage._structured_to_markdown(structured)
        # Should have pipe-delimited table rows
        lines = md.strip().split("\n")
        table_lines = [l for l in lines if l.startswith("|")]
        assert len(table_lines) > 0

    def test_empty_structured_produces_output(self):
        md = ParsingStage._structured_to_markdown(
            {"sheets": [], "named_ranges": {}, "sheet_count": 0, "total_rows": 0}
        )
        assert isinstance(md, str)

    def test_named_ranges_section(self):
        structured = {
            "sheets": [],
            "named_ranges": {"EBITDA": "Sheet1!B15"},
            "sheet_count": 0,
            "total_rows": 0,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "## Named Ranges" in md
        assert "EBITDA" in md
        assert "Sheet1!B15" in md


# ===========================================================================
# Merged cell propagation tests
# ===========================================================================


class TestMergedCellPropagation:
    """Test that merged cell values are propagated to all cells in the region."""

    def test_merged_header_propagated_horizontally(self):
        """A1:C1 merged with 'Header' -- B1 and C1 should get the value."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header", None, None], [1, 2, 3]]},
            merged_ranges=[("S1", "A1:C1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row1 = next(r for r in result["sheets"][0]["rows"] if r["row_index"] == 1)
        refs = {c["ref"]: c for c in row1["cells"]}
        assert "A1" in refs
        assert "B1" in refs
        assert "C1" in refs
        assert refs["A1"]["value"] == "Header"
        assert refs["B1"]["value"] == "Header"
        assert refs["C1"]["value"] == "Header"

    def test_merged_cells_have_is_merged_flag(self):
        """All cells in a merged region should have is_merged=True."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Title", None, None], [1, 2, 3]]},
            merged_ranges=[("S1", "A1:C1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row1 = next(r for r in result["sheets"][0]["rows"] if r["row_index"] == 1)
        for cell in row1["cells"]:
            if cell["ref"] in ("A1", "B1", "C1"):
                assert cell.get("is_merged") is True, f"{cell['ref']} missing is_merged"

    def test_merged_vertical_propagation(self):
        """A1:A3 merged vertically -- rows 2 and 3 should get the value."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Section", 10], [None, 20], [None, 30]]},
            merged_ranges=[("S1", "A1:A3")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        rows = result["sheets"][0]["rows"]
        for r in rows:
            a_cell = next((c for c in r["cells"] if c["ref"].startswith("A")), None)
            assert a_cell is not None, f"Row {r['row_index']} missing column A cell"
            assert a_cell["value"] == "Section"

    def test_merged_bold_formatting_propagated(self):
        """Bold formatting on top-left should propagate to all merged cells."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Bold Header", None], [1, 2]]},
            bold_cells={("S1", 1, 1)},
            merged_ranges=[("S1", "A1:B1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row1 = next(r for r in result["sheets"][0]["rows"] if r["row_index"] == 1)
        for cell in row1["cells"]:
            if cell["ref"] in ("A1", "B1"):
                assert cell["is_bold"] is True, f"{cell['ref']} not bold"

    def test_merged_rect_block(self):
        """A 2x2 merged block propagates to all 4 cells."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Block", None], [None, None], ["data", "data2"]]},
            merged_ranges=[("S1", "A1:B2")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        rows = result["sheets"][0]["rows"]
        merged_refs = set()
        for r in rows:
            for c in r["cells"]:
                if c.get("is_merged"):
                    merged_refs.add(c["ref"])
        assert {"A1", "B1", "A2", "B2"} <= merged_refs

    def test_merged_regions_still_listed(self):
        """Merged regions in metadata should still be listed."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header", None, None], [1, 2, 3]]},
            merged_ranges=[("S1", "A1:C1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        assert "A1:C1" in result["sheets"][0]["merged_regions"]


# ===========================================================================
# Chunking tests
# ===========================================================================


class TestChunking:
    """Test chunking threshold detection and merge logic."""

    def test_estimate_token_count_returns_int(self):
        structured = {
            "sheets": [{"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": []}],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 0,
        }
        count = ParsingStage._estimate_token_count(structured)
        assert isinstance(count, int)
        assert count >= 0

    def test_should_chunk_false_for_small_file(self):
        structured = {
            "sheets": [{"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": []}],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 0,
        }
        assert ParsingStage._should_chunk(structured) is False

    def test_should_chunk_true_for_large_file(self):
        """Generate enough data to exceed 50K token threshold."""
        # 50K tokens ≈ 200K chars of markdown. Generate many rows with
        # multiple columns to produce enough output.
        big_rows = []
        for i in range(2000):
            big_rows.append(
                {
                    "row_index": i,
                    "cells": [
                        {
                            "ref": f"A{i}",
                            "value": f"Financial line item label for row number {i}",
                            "formula": None,
                            "is_bold": False,
                            "indent_level": 0,
                            "number_format": "General",
                        },
                        {
                            "ref": f"B{i}",
                            "value": 12345.67,
                            "formula": None,
                            "is_bold": False,
                            "indent_level": 0,
                            "number_format": "#,##0",
                        },
                        {
                            "ref": f"C{i}",
                            "value": 23456.78,
                            "formula": None,
                            "is_bold": False,
                            "indent_level": 0,
                            "number_format": "#,##0",
                        },
                        {
                            "ref": f"D{i}",
                            "value": 34567.89,
                            "formula": None,
                            "is_bold": False,
                            "indent_level": 0,
                            "number_format": "#,##0",
                        },
                        {
                            "ref": f"E{i}",
                            "value": 45678.90,
                            "formula": None,
                            "is_bold": False,
                            "indent_level": 0,
                            "number_format": "#,##0",
                        },
                    ],
                }
            )
        structured = {
            "sheets": [
                {"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": big_rows}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 2000,
        }
        assert ParsingStage._should_chunk(structured) is True

    def test_merge_parsed_sheets(self):
        r1 = {"sheets": [{"sheet_name": "S1", "rows": []}]}
        r2 = {"sheets": [{"sheet_name": "S2", "rows": []}]}
        merged = ParsingStage._merge_parsed_sheets([r1, r2])
        assert len(merged["sheets"]) == 2
        names = [s["sheet_name"] for s in merged["sheets"]]
        assert "S1" in names
        assert "S2" in names


# ===========================================================================
# Formula reference extraction tests
# ===========================================================================


class TestFormulaReferenceExtraction:
    """Test _extract_cell_references() with various formula patterns."""

    def test_simple_addition(self):
        refs = ParsingStage._extract_cell_references("=A1+B1")
        assert "A1" in refs
        assert "B1" in refs

    def test_sum_range(self):
        refs = ParsingStage._extract_cell_references("=SUM(A1:A10)")
        assert "A1" in refs
        assert "A10" in refs

    def test_cross_sheet_ref(self):
        refs = ParsingStage._extract_cell_references("=Sheet2!B5")
        assert "Sheet2!B5" in refs

    def test_quoted_sheet_ref(self):
        refs = ParsingStage._extract_cell_references("='Income Statement'!A1")
        assert "'Income Statement'!A1" in refs

    def test_mixed_refs(self):
        refs = ParsingStage._extract_cell_references("=A1+Sheet2!B5+SUM(C1:C10)")
        assert "A1" in refs
        assert "Sheet2!B5" in refs
        assert "C1" in refs
        assert "C10" in refs

    def test_absolute_refs(self):
        refs = ParsingStage._extract_cell_references("=$A$1+$B2+C$3")
        assert "A1" in refs
        assert "B2" in refs
        assert "C3" in refs

    def test_no_formula_returns_empty(self):
        assert ParsingStage._extract_cell_references("hello") == []
        assert ParsingStage._extract_cell_references("") == []
        assert ParsingStage._extract_cell_references(None) == []

    def test_references_in_structured_repr(self):
        """Verify references appear in structured repr for formula cells."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[10, 20, None]]},
            formulas={("S1", 1, 3): "=A1+B1"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row = result["sheets"][0]["rows"][0]
        c3 = next(c for c in row["cells"] if c["ref"] == "C1")
        assert "references" in c3
        assert "A1" in c3["references"]
        assert "B1" in c3["references"]


# ===========================================================================
# Hidden sheet markdown tests
# ===========================================================================


class TestHiddenSheetMarkdown:
    """Test hidden sheet markdown rendering with [HIDDEN] prefix."""

    def test_hidden_sheet_uses_prefix(self):
        structured = {
            "sheets": [
                {"sheet_name": "Secret", "is_hidden": True, "merged_regions": [], "rows": []}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 0,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "## Sheet: [HIDDEN] Secret" in md

    def test_visible_sheet_no_prefix(self):
        structured = {
            "sheets": [
                {"sheet_name": "Visible", "is_hidden": False, "merged_regions": [], "rows": []}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 0,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "[HIDDEN]" not in md
        assert "## Sheet: Visible" in md


# ===========================================================================
# Number format annotation tests
# ===========================================================================


class TestFormatAnnotations:
    """Test number format intelligence in markdown output."""

    def test_percentage_column_annotated(self):
        """Column with >50% percentage-formatted cells should be annotated."""
        rows = [
            {
                "row_index": 1,
                "cells": [
                    {
                        "ref": "A1",
                        "value": "Label",
                        "formula": None,
                        "is_bold": True,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B1",
                        "value": 0.05,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0%",
                    },
                ],
            },
            {
                "row_index": 2,
                "cells": [
                    {
                        "ref": "A2",
                        "value": "Rate2",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B2",
                        "value": 0.08,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0%",
                    },
                ],
            },
        ]
        structured = {
            "sheets": [
                {"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": rows}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 2,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "| Fmt |" in md
        assert "%" in md

    def test_currency_column_annotated(self):
        """Column with >50% currency-formatted cells gets $ annotation."""
        rows = [
            {
                "row_index": 1,
                "cells": [
                    {
                        "ref": "A1",
                        "value": "Revenue",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B1",
                        "value": 100000,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "#,##0",
                    },
                ],
            },
            {
                "row_index": 2,
                "cells": [
                    {
                        "ref": "A2",
                        "value": "Costs",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B2",
                        "value": 50000,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "#,##0",
                    },
                ],
            },
        ]
        structured = {
            "sheets": [
                {"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": rows}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 2,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "| Fmt |" in md
        assert "$" in md

    def test_no_annotation_when_mixed_formats(self):
        """No format annotation when column has mixed formats (<50% any)."""
        rows = [
            {
                "row_index": 1,
                "cells": [
                    {
                        "ref": "B1",
                        "value": 100,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "#,##0",
                    },
                ],
            },
            {
                "row_index": 2,
                "cells": [
                    {
                        "ref": "B2",
                        "value": 0.05,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0%",
                    },
                ],
            },
            {
                "row_index": 3,
                "cells": [
                    {
                        "ref": "B3",
                        "value": 42,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                ],
            },
        ]
        structured = {
            "sheets": [
                {"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": rows}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 3,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "| Fmt |" not in md

    def test_format_row_has_empty_indent_formula(self):
        """Fmt annotation row should have empty entries for Indent and Formula columns."""
        rows = [
            {
                "row_index": 1,
                "cells": [
                    {
                        "ref": "A1",
                        "value": "Label",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B1",
                        "value": 0.05,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0%",
                    },
                ],
            },
            {
                "row_index": 2,
                "cells": [
                    {
                        "ref": "A2",
                        "value": "Rate2",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B2",
                        "value": 0.08,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0%",
                    },
                ],
            },
        ]
        structured = {
            "sheets": [
                {"sheet_name": "S1", "is_hidden": False, "merged_regions": [], "rows": rows}
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 2,
        }
        md = ParsingStage._structured_to_markdown(structured)
        fmt_line = [l for l in md.split("\n") if l.startswith("| Fmt")]
        assert len(fmt_line) == 1
        # Fmt row should end with 3 empty trailing columns (Bold, Indent, Formula)
        parts = fmt_line[0].split("|")
        # Last 3 non-empty-ish parts before trailing empty should be empty
        stripped = [p.strip() for p in parts]
        # Indent and Formula columns should be empty
        assert stripped[-2] == ""  # Formula
        assert stripped[-3] == ""  # Indent


# ===========================================================================
# Markdown Formula column tests
# ===========================================================================


class TestMarkdownFormulaColumn:
    """Test that formula text appears in the markdown Formula column."""

    def test_formula_column_in_header(self):
        """Header row should contain 'Formula' column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "X",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        header_line = [l for l in md.split("\n") if l.startswith("| Row")][0]
        assert "Formula" in header_line

    def test_formula_shown_for_formula_cell(self):
        """Row with a formula cell should show the formula in the Formula column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Total",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                                {
                                    "ref": "B1",
                                    "value": 300,
                                    "formula": "=SUM(B2:B5)",
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        assert len(data_lines) == 1
        assert "=SUM(B2:B5)" in data_lines[0]

    def test_no_formula_for_plain_value(self):
        """Row without formula should have empty Formula column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Revenue",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                                {
                                    "ref": "B1",
                                    "value": 100,
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        assert len(data_lines) == 1
        # Last pipe-delimited field (before trailing |) should be empty
        parts = [p.strip() for p in data_lines[0].split("|")]
        # parts: ['', '1', 'Revenue', '100', 'N', '0', '', '']
        formula_field = parts[-2]  # second-to-last (last is empty after trailing |)
        assert formula_field == ""

    def test_formula_picks_first_formula_cell(self):
        """When multiple cells have formulas, use the first one encountered."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Item",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                                {
                                    "ref": "B1",
                                    "value": 100,
                                    "formula": "=C1+D1",
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                                {
                                    "ref": "C1",
                                    "value": 50,
                                    "formula": "=D1*2",
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        assert "=C1+D1" in data_lines[0]


# ===========================================================================
# Markdown Indent column tests
# ===========================================================================


class TestMarkdownIndentColumn:
    """Test that indent level appears in the markdown Indent column."""

    def test_indent_column_in_header(self):
        """Header row should contain 'Indent' column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "X",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        header_line = [l for l in md.split("\n") if l.startswith("| Row")][0]
        assert "Indent" in header_line

    def test_indent_level_rendered(self):
        """Cell with indent_level=2 should show '2' in the Indent column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Sub-item",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 2,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        parts = [p.strip() for p in data_lines[0].split("|")]
        # parts: ['', '1', 'Sub-item', 'N', '2', '', '']
        # Indent is after Bold, before Formula
        indent_field = parts[-3]
        assert indent_field == "2"

    def test_zero_indent_default(self):
        """Cell with no indent should show '0' in the Indent column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Top-level",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        parts = [p.strip() for p in data_lines[0].split("|")]
        indent_field = parts[-3]
        assert indent_field == "0"


# ===========================================================================
# WS-1: Color extraction tests
# ===========================================================================


class TestColorExtraction:
    """Test font_color, fill_color, and border extraction."""

    def test_font_color_blue_detected(self):
        """Blue font color should be extracted as lowercase hex."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            font_colors={("S1", 1, 1): "0000FF"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["font_color"] is not None
        assert "00" in cell["font_color"] and "ff" in cell["font_color"]

    def test_font_color_default_resolves_to_white(self):
        """Default font (theme=1/white) resolves after theme color support."""
        xlsx = _make_minimal_xlsx(data={"S1": [[100]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        # openpyxl default font uses theme=1 (lt1/white), now correctly resolved
        assert cell["font_color"] == "ffffff"

    def test_fill_color_yellow_detected(self):
        """Solid yellow fill should be extracted."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            fill_colors={("S1", 1, 1): "FFFF00"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["fill_color"] is not None

    def test_fill_color_none_for_no_fill(self):
        """Cell without fill should yield None."""
        xlsx = _make_minimal_xlsx(data={"S1": [[100]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["fill_color"] is None

    def test_border_bottom_detected(self):
        """Cell with bottom border should set has_border_bottom=True."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            borders={("S1", 1, 1): {"bottom": True, "right": False}},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["has_border_bottom"] is True
        assert cell["has_border_right"] is False

    def test_border_right_detected(self):
        """Cell with right border should set has_border_right=True."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            borders={("S1", 1, 1): {"bottom": False, "right": True}},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["has_border_right"] is True
        assert cell["has_border_bottom"] is False

    def test_no_border_defaults_false(self):
        """Cell without borders should default to False."""
        xlsx = _make_minimal_xlsx(data={"S1": [[100]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["has_border_bottom"] is False
        assert cell["has_border_right"] is False


# ===========================================================================
# WS-1: Cell type derivation tests
# ===========================================================================


class TestCellType:
    """Test cell_type classification."""

    def test_formula_cell_type(self):
        """Cell with a formula should be classified as 'formula'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[10, 20, None]]},
            formulas={("S1", 1, 3): "=A1+B1"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cells = result["sheets"][0]["rows"][0]["cells"]
        formula_cell = next(c for c in cells if c.get("formula"))
        assert formula_cell["cell_type"] == "formula"

    def test_input_cell_blue_font(self):
        """Blue font, no formula should be classified as 'input'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            font_colors={("S1", 1, 1): "0000FF"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["cell_type"] == "input"

    def test_input_cell_yellow_fill(self):
        """Light yellow fill, no formula should be classified as 'input'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            # Light yellow: FFFFCC -> high R, high G, low B
            fill_colors={("S1", 1, 1): "FFFFCC"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["cell_type"] == "input"

    def test_label_cell_in_column_a(self):
        """String in column A should be classified as 'label'."""
        xlsx = _make_minimal_xlsx(data={"S1": [["Revenue"]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["cell_type"] == "label"

    def test_label_cell_in_column_c(self):
        """String in column C (3rd column, index 2) should be 'label'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[None, None, "Note"]]},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["ref"] == "C1"
        assert cell["cell_type"] == "label"

    def test_value_cell_numeric_column_d(self):
        """Numeric in column D (index 3, outside label range) should be 'value'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[None, None, None, 42]]},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell["ref"] == "D1"
        assert cell["cell_type"] == "value"

    def test_formula_takes_priority_over_blue_font(self):
        """Formula should override blue font → 'formula' not 'input'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[10, None]]},
            formulas={("S1", 1, 2): "=A1*2"},
            font_colors={("S1", 1, 2): "0000FF"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cells = result["sheets"][0]["rows"][0]["cells"]
        b1 = next(c for c in cells if c["ref"] == "B1")
        assert b1["cell_type"] == "formula"


# ===========================================================================
# WS-1: Merged cell enhancement tests
# ===========================================================================


class TestMergedCellEnhancements:
    """Test merge_origin and field propagation for merged cells."""

    def test_merge_origin_horizontal(self):
        """Horizontal merge A1:C1 should set merge_origin='A1' on all cells."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header", None, None], [1, 2, 3]]},
            merged_ranges=[("S1", "A1:C1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row1 = next(r for r in result["sheets"][0]["rows"] if r["row_index"] == 1)
        for cell in row1["cells"]:
            assert cell.get("merge_origin") == "A1"
            assert cell.get("is_merged") is True

    def test_merge_origin_vertical(self):
        """Vertical merge A1:A3 should set merge_origin='A1' on A2 and A3."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header"], [None], [None]]},
            merged_ranges=[("S1", "A1:A3")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        for r in result["sheets"][0]["rows"]:
            for cell in r["cells"]:
                if cell["ref"] in ("A1", "A2", "A3"):
                    assert cell.get("merge_origin") == "A1"

    def test_merged_cells_propagate_new_fields(self):
        """New formatting fields should propagate from top-left to merged cells."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Header", None, None]]},
            bold_cells={("S1", 1, 1)},
            font_colors={("S1", 1, 1): "FF0000"},
            merged_ranges=[("S1", "A1:C1")],
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row1 = next(r for r in result["sheets"][0]["rows"] if r["row_index"] == 1)
        for cell in row1["cells"]:
            assert cell.get("is_bold") is True
            assert cell.get("font_color") is not None


# ===========================================================================
# WS-1: Section boundary detection tests
# ===========================================================================


class TestSectionBoundaryDetection:
    """Test section boundary detection from formatting cues."""

    def test_bold_after_gap_is_boundary(self):
        """Bold row preceded by gap in row_index should be a section boundary."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    [1, 2],  # row 1
                    [None, None],  # row 2 (will be skipped as empty)
                    ["Revenue", 100],  # row 3
                ]
            },
            bold_cells={("S1", 3, 1)},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        boundaries = result["sheets"][0]["section_boundaries"]
        assert any(b["label"] == "Revenue" for b in boundaries)

    def test_first_bold_row_is_boundary(self):
        """First row that is bold should be a section boundary."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [["Income Statement", None]]},
            bold_cells={("S1", 1, 1)},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        boundaries = result["sheets"][0]["section_boundaries"]
        assert len(boundaries) >= 1
        assert boundaries[0]["label"] == "Income Statement"

    def test_bold_with_border_is_boundary(self):
        """Bold row with border_bottom should be a section boundary."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Data", 100],
                    ["Revenue", 200],
                ]
            },
            bold_cells={("S1", 2, 1)},
            borders={("S1", 2, 1): {"bottom": True, "right": False}},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        boundaries = result["sheets"][0]["section_boundaries"]
        assert any(b["label"] == "Revenue" for b in boundaries)

    def test_non_bold_row_not_boundary(self):
        """Non-bold row should not be a section boundary even after gap."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    [1, 2],
                    [None, None],
                    ["Revenue", 100],
                ]
            },
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        boundaries = result["sheets"][0]["section_boundaries"]
        assert not any(b.get("label") == "Revenue" for b in boundaries)

    def test_empty_sheet_no_boundaries(self):
        """Empty sheet should have no section boundaries."""
        xlsx = _make_minimal_xlsx(data={"S1": [[None]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        boundaries = result["sheets"][0]["section_boundaries"]
        assert boundaries == []


# ===========================================================================
# WS-1: Formula graph summary tests
# ===========================================================================


class TestFormulaGraphSummary:
    """Test formula dependency graph summary."""

    def test_formula_count(self):
        """Should count all formula cells."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[1, 2, None], [3, 4, None]]},
            formulas={
                ("S1", 1, 3): "=A1+B1",
                ("S1", 2, 3): "=A2+B2",
            },
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        summary = result["sheets"][0]["formula_graph_summary"]
        assert summary["formula_count"] == 2

    def test_subtotal_detected(self):
        """SUM of contiguous same-column range should be detected as subtotal."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[10], [20], [None]]},
            formulas={("S1", 3, 1): "=SUM(A1:A2)"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        summary = result["sheets"][0]["formula_graph_summary"]
        assert summary["subtotal_count"] >= 1

    def test_cross_sheet_ref_detected(self):
        """Formula referencing another sheet should be detected."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[None]], "S2": [[100]]},
            formulas={("S1", 1, 1): "=S2!A1"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        summary = result["sheets"][0]["formula_graph_summary"]
        assert summary["cross_sheet_ref_count"] >= 1

    def test_sample_formulas_limited(self):
        """Sample formulas should be capped at 5."""
        data = {"S1": [[i] for i in range(10)]}
        formulas = {("S1", i, 1): f"=A{i}+1" for i in range(1, 11)}
        xlsx = _make_minimal_xlsx(data=data, formulas=formulas)
        result = ParsingStage._excel_to_structured_repr(xlsx)
        summary = result["sheets"][0]["formula_graph_summary"]
        assert len(summary["sample_formulas"]) <= 5

    def test_no_formulas_returns_zeros(self):
        """Sheet without formulas should return all-zero summary."""
        xlsx = _make_minimal_xlsx(data={"S1": [[1, 2], [3, 4]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        summary = result["sheets"][0]["formula_graph_summary"]
        assert summary["formula_count"] == 0
        assert summary["subtotal_count"] == 0
        assert summary["cross_sheet_ref_count"] == 0


# ===========================================================================
# WS-1: Comment extraction tests
# ===========================================================================


class TestCommentExtraction:
    """Test cell comment extraction and markdown footnotes."""

    def test_comment_extracted(self):
        """Cell comment should appear in structured output."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[100]]},
            comments={("S1", 1, 1): "Auditor note"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert cell.get("comment") == "Auditor note"

    def test_no_comment_field_when_absent(self):
        """Cell without comment should not have 'comment' key (sparse)."""
        xlsx = _make_minimal_xlsx(data={"S1": [[100]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        cell = result["sheets"][0]["rows"][0]["cells"][0]
        assert "comment" not in cell

    def test_comment_in_markdown_footnotes(self):
        """Comments should appear as footnotes in markdown output."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "section_boundaries": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Revenue",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                    "comment": "Source: mgmt",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "**Notes:**" in md
        assert "[A1]: Source: mgmt" in md


# ===========================================================================
# WS-1: Enhanced markdown tests
# ===========================================================================


class TestEnhancedMarkdown:
    """Test Type column, visibility labels, format annotations, section separators."""

    def test_type_column_in_header(self):
        """Header row should contain 'Type' column."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "section_boundaries": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "X",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        header_line = [l for l in md.split("\n") if l.startswith("| Row")][0]
        assert "Type" in header_line

    def test_type_column_shows_formula(self):
        """Row with formula cell should show Type=F."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "section_boundaries": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": 100,
                                    "formula": "=SUM(A2:A5)",
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                    "cell_type": "formula",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        parts = [p.strip() for p in data_lines[0].split("|")]
        # Type is between last data column and Bold
        assert "F" in parts

    def test_type_column_shows_input(self):
        """Row with input cell should show Type=I."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "section_boundaries": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": 100,
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                    "cell_type": "input",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        data_lines = [l for l in md.split("\n") if l.startswith("| 1")]
        parts = [p.strip() for p in data_lines[0].split("|")]
        assert "I" in parts

    def test_very_hidden_sheet_annotated(self):
        """veryHidden sheet should show [VERY HIDDEN] in markdown."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "Secret",
                    "is_hidden": True,
                    "visibility": "veryHidden",
                    "merged_regions": [],
                    "section_boundaries": [],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": 1,
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        }
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 1,
        }
        md = ParsingStage._structured_to_markdown(structured)
        assert "[VERY HIDDEN]" in md

    def test_section_separator_in_markdown(self):
        """Section boundary rows should be preceded by a separator."""
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "section_boundaries": [{"row_index": 3, "label": "Revenue"}],
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {
                                    "ref": "A1",
                                    "value": "Data",
                                    "formula": None,
                                    "is_bold": False,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        },
                        {
                            "row_index": 3,
                            "cells": [
                                {
                                    "ref": "A3",
                                    "value": "Revenue",
                                    "formula": None,
                                    "is_bold": True,
                                    "indent_level": 0,
                                    "number_format": "General",
                                },
                            ],
                        },
                    ],
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 2,
        }
        md = ParsingStage._structured_to_markdown(structured)
        lines = md.split("\n")
        # Find the Revenue row and check its predecessor is a separator
        for i, line in enumerate(lines):
            if line.startswith("| 3"):
                assert "---" in lines[i - 1]
                break
        else:
            pytest.fail("Row 3 not found in markdown output")

    def test_multiplier_format_annotation(self):
        """Column formatted as '0.0x' should show 'x' in Fmt row."""
        rows = [
            {
                "row_index": 1,
                "cells": [
                    {
                        "ref": "A1",
                        "value": "Turns",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B1",
                        "value": 3.5,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0",
                    },
                ],
            },
            {
                "row_index": 2,
                "cells": [
                    {
                        "ref": "A2",
                        "value": "Coverage",
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "General",
                    },
                    {
                        "ref": "B2",
                        "value": 2.1,
                        "formula": None,
                        "is_bold": False,
                        "indent_level": 0,
                        "number_format": "0.0",
                    },
                ],
            },
        ]
        structured = {
            "sheets": [
                {
                    "sheet_name": "S1",
                    "is_hidden": False,
                    "merged_regions": [],
                    "section_boundaries": [],
                    "rows": rows,
                }
            ],
            "named_ranges": {},
            "sheet_count": 1,
            "total_rows": 2,
        }
        md = ParsingStage._structured_to_markdown(structured)
        fmt_line = [l for l in md.split("\n") if l.startswith("| Fmt")]
        assert len(fmt_line) == 1
        assert "x" in fmt_line[0]


# ===========================================================================
# WS-1: Hidden sheet visibility field tests
# ===========================================================================


class TestVisibilityField:
    """Test visibility field in structured output."""

    def test_very_hidden_sheet_detected(self):
        """veryHidden sheet should have is_hidden=True and visibility='veryHidden'."""
        xlsx = _make_minimal_xlsx(
            data={"S1": [[1]], "Secret": [[2]]},
            very_hidden_sheets={"Secret"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        secret = next(s for s in result["sheets"] if s["sheet_name"] == "Secret")
        assert secret["is_hidden"] is True
        assert secret["visibility"] == "veryHidden"

    def test_visible_sheet_has_visibility(self):
        """Visible sheet should have visibility='visible'."""
        xlsx = _make_minimal_xlsx(data={"S1": [[1]]})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        assert result["sheets"][0]["visibility"] == "visible"

    def test_hidden_sheet_visibility(self):
        """Hidden sheet should have visibility='hidden'."""
        xlsx = _make_minimal_xlsx(
            data={"Visible": [[1]], "S1": [[2]]},
            hidden_sheets={"S1"},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        s1 = next(s for s in result["sheets"] if s["sheet_name"] == "S1")
        assert s1["visibility"] == "hidden"


# ===========================================================================
# WS-3: Messy-sheet detection heuristic tests
# ===========================================================================


class TestDetectLabelColumn:
    """Tests for _detect_label_column()."""

    def test_standard_layout_returns_A(self):
        """Column A has string labels, B-D have numbers -> 'A'."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100, 200, 300],
                    ["COGS", 50, 60, 70],
                    ["Gross Profit", 50, 140, 230],
                    ["OpEx", 20, 30, 40],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        col = ParsingStage._detect_label_column(result["sheets"][0]["rows"])
        assert col == "A"

    def test_offset_labels_in_column_B(self):
        """Column A has row numbers, column B has strings -> 'B'."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    [1, "Revenue", 100, 200],
                    [2, "COGS", 50, 60],
                    [3, "Gross Profit", 50, 140],
                    [4, "OpEx", 20, 30],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        col = ParsingStage._detect_label_column(result["sheets"][0]["rows"])
        assert col == "B"

    def test_all_numeric_returns_none(self):
        """All columns are numbers -> None."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    [1, 2, 3],
                    [4, 5, 6],
                    [7, 8, 9],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        col = ParsingStage._detect_label_column(result["sheets"][0]["rows"])
        assert col is None

    def test_empty_rows_returns_none(self):
        """Empty rows list -> None."""
        col = ParsingStage._detect_label_column([])
        assert col is None


class TestDetectHeaderRow:
    """Tests for _detect_header_row()."""

    def test_fiscal_year_header(self):
        """Row 1 has 'FY2022', 'FY2023', 'FY2024E' -> row 1."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["", "FY2022", "FY2023", "FY2024E"],
                    ["Revenue", 100, 200, 300],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row_idx = ParsingStage._detect_header_row(result["sheets"][0]["rows"])
        assert row_idx == 1

    def test_header_in_row_3(self):
        """Title rows 1-2, period values in row 3 -> row 3."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Company Name", None, None],
                    ["Financial Statements", None, None],
                    ["", "2022", "2023", "2024"],
                    ["Revenue", 100, 200, 300],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row_idx = ParsingStage._detect_header_row(result["sheets"][0]["rows"])
        assert row_idx == 3

    def test_no_period_values_returns_none(self):
        """No row has period-like values -> None."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100, 200],
                    ["COGS", 50, 60],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row_idx = ParsingStage._detect_header_row(result["sheets"][0]["rows"])
        assert row_idx is None

    def test_header_in_row_15(self):
        """Period headers pushed to row 15 by disclaimers -> still detected."""
        data = [[f"Disclaimer line {i}", None, None] for i in range(1, 15)]
        data.append(["", "2022", "2023", "2024"])
        data.append(["Revenue", 100, 200, 300])
        xlsx = _make_minimal_xlsx(data={"S1": data})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        row_idx = ParsingStage._detect_header_row(result["sheets"][0]["rows"])
        assert row_idx == 15


class TestDetectTableRegions:
    """Tests for _detect_table_regions()."""

    def test_single_region(self):
        """Contiguous rows -> 1 region."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["A", 1],
                    ["B", 2],
                    ["C", 3],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        regions = ParsingStage._detect_table_regions(result["sheets"][0]["rows"])
        assert len(regions) == 1
        assert regions[0]["start_row"] == 1
        assert regions[0]["end_row"] == 3

    def test_multiple_regions_with_gap(self):
        """Rows 1-3, blank rows 4-6, rows 7-9 -> 2 regions."""
        data = {
            "S1": [
                ["Revenue", 100],  # row 1
                ["COGS", 50],  # row 2
                ["Profit", 50],  # row 3
                [None, None],  # row 4 (blank)
                [None, None],  # row 5 (blank)
                [None, None],  # row 6 (blank)
                ["Assets", 1000],  # row 7
                ["Liabilities", 500],  # row 8
                ["Equity", 500],  # row 9
            ]
        }
        xlsx = _make_minimal_xlsx(data=data)
        result = ParsingStage._excel_to_structured_repr(xlsx)
        regions = ParsingStage._detect_table_regions(result["sheets"][0]["rows"])
        assert len(regions) == 2
        assert regions[0]["start_row"] == 1
        assert regions[0]["end_row"] == 3
        assert regions[1]["start_row"] == 7
        assert regions[1]["end_row"] == 9

    def test_empty_rows_returns_empty(self):
        """Empty rows list -> empty list."""
        regions = ParsingStage._detect_table_regions([])
        assert regions == []


class TestDetectTransposed:
    """Tests for _detect_transposed()."""

    def test_transposed_detected(self):
        """Periods in column A -> True."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["2020", 100, 200],
                    ["2021", 110, 210],
                    ["2022", 120, 220],
                    ["2023", 130, 230],
                    ["2024", 140, 240],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        is_t = ParsingStage._detect_transposed(result["sheets"][0]["rows"], "A")
        assert is_t is True

    def test_not_transposed(self):
        """Standard layout -> False."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100, 200],
                    ["COGS", 50, 60],
                    ["Profit", 50, 140],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        is_t = ParsingStage._detect_transposed(result["sheets"][0]["rows"], "A")
        assert is_t is False


class TestDetectNonFinancialRows:
    """Tests for _detect_non_financial_rows()."""

    def test_source_and_separator_detected(self):
        """'Source:' and '---' rows are marked as non-financial."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100],
                    ["COGS", 50],
                    ["Source: Company filings", None],
                    ["---", None],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        nf = ParsingStage._detect_non_financial_rows(result["sheets"][0]["rows"])
        assert 3 in nf
        assert 4 in nf
        assert 1 not in nf
        assert 2 not in nf

    def test_note_detected(self):
        """'Note: ...' rows are non-financial."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100],
                    ["Note: audited figures", None],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        nf = ParsingStage._detect_non_financial_rows(result["sheets"][0]["rows"])
        assert 2 in nf

    def test_management_estimates_detected(self):
        """'Management estimates' is a non-financial annotation."""
        rows = [
            {"row_index": 1, "cells": [{"ref": "A1", "value": "Revenue"}]},
            {"row_index": 2, "cells": [{"ref": "A2", "value": "Management estimates only"}]},
        ]
        nf = ParsingStage._detect_non_financial_rows(rows)
        assert 2 in nf

    def test_unaudited_detected(self):
        """'Unaudited' is a non-financial annotation."""
        rows = [
            {"row_index": 1, "cells": [{"ref": "A1", "value": "Revenue"}]},
            {"row_index": 2, "cells": [{"ref": "A2", "value": "Unaudited interim results"}]},
        ]
        nf = ParsingStage._detect_non_financial_rows(rows)
        assert 2 in nf

    def test_as_of_detected(self):
        """'As of December 31, 2023' is a non-financial annotation."""
        rows = [
            {"row_index": 1, "cells": [{"ref": "A1", "value": "As of December 31, 2023"}]},
        ]
        nf = ParsingStage._detect_non_financial_rows(rows)
        assert 1 in nf

    def test_for_the_period_detected(self):
        """'For the period ended...' is a non-financial annotation."""
        rows = [
            {"row_index": 1, "cells": [{"ref": "A1", "value": "For the period ended Dec 2023"}]},
        ]
        nf = ParsingStage._detect_non_financial_rows(rows)
        assert 1 in nf


class TestDetectUnitHint:
    """Tests for _detect_unit_hint()."""

    def test_thousands_in_cell(self):
        """'(in thousands)' in first rows -> ('thousands', 1000.0)."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["(in thousands)", None, None],
                    ["Revenue", 100, 200],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        hint, mult = ParsingStage._detect_unit_hint(result["sheets"][0]["rows"], "S1")
        assert hint == "thousands"
        assert mult == 1_000.0

    def test_millions_in_sheet_name(self):
        """Sheet name '($millions)' -> ('millions', 1000000.0)."""
        xlsx = _make_minimal_xlsx(
            data={
                "P&L ($millions)": [
                    ["Revenue", 100],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        hint, mult = ParsingStage._detect_unit_hint(result["sheets"][0]["rows"], "P&L ($millions)")
        assert hint == "millions"
        assert mult == 1_000_000.0

    def test_no_unit_returns_none(self):
        """No unit text -> (None, None)."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        hint, mult = ParsingStage._detect_unit_hint(result["sheets"][0]["rows"], "S1")
        assert hint is None
        assert mult is None

    def test_unit_in_row_8(self):
        """Unit annotation in row 8 (past old limit of 5) -> detected."""
        data = [
            ["Company Name", None],
            ["Report Title", None],
            ["Prepared By: Finance", None],
            ["Date: 2024-01-01", None],
            ["Confidential", None],
            [None, None],
            [None, None],
            ["(in millions)", None],
            ["Revenue", 100],
        ]
        xlsx = _make_minimal_xlsx(data={"S1": data})
        result = ParsingStage._excel_to_structured_repr(xlsx)
        hint, mult = ParsingStage._detect_unit_hint(result["sheets"][0]["rows"], "S1")
        assert hint == "millions"
        assert mult == 1_000_000.0


class TestSheetMetadata:
    """Tests for _detect_sheet_metadata() and markdown integration."""

    def test_metadata_dict_has_all_keys(self):
        """_detect_sheet_metadata returns dict with expected keys."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["Revenue", 100, 200],
                    ["COGS", 50, 60],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        meta = ParsingStage._detect_sheet_metadata(result["sheets"][0])
        assert "label_column" in meta
        assert "header_row_index" in meta
        assert "table_regions" in meta
        assert "is_transposed" in meta
        assert "non_financial_rows" in meta
        assert "unit_hint" in meta
        assert "unit_multiplier" in meta

    def test_max_row_captured(self):
        """max_row key is present in sheet dict after extraction."""
        xlsx = _make_minimal_xlsx(
            data={
                "S1": [
                    ["A", 1],
                    ["B", 2],
                ]
            }
        )
        result = ParsingStage._excel_to_structured_repr(xlsx)
        assert "max_row" in result["sheets"][0]
        assert result["sheets"][0]["max_row"] >= 2

    def test_metadata_in_markdown_label_column(self):
        """Markdown shows 'Labels: column B' when labels are in column B."""
        sheet = {
            "sheet_name": "S1",
            "is_hidden": False,
            "merged_regions": [],
            "label_column": "B",
            "header_row_index": None,
            "unit_hint": None,
            "is_transposed": False,
            "table_regions": [],
            "non_financial_rows": set(),
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "B1", "value": "Revenue", "number_format": "General"},
                    ],
                },
            ],
        }
        md = ParsingStage._structured_to_markdown({"sheets": [sheet]})
        assert "Labels: column B" in md

    def test_metadata_in_markdown_units(self):
        """Markdown shows 'Units: thousands' when unit_hint is set."""
        sheet = {
            "sheet_name": "S1",
            "is_hidden": False,
            "merged_regions": [],
            "label_column": "A",
            "header_row_index": None,
            "unit_hint": "thousands",
            "is_transposed": False,
            "table_regions": [],
            "non_financial_rows": set(),
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": "Revenue", "number_format": "General"},
                    ],
                },
            ],
        }
        md = ParsingStage._structured_to_markdown({"sheets": [sheet]})
        assert "Units: thousands" in md

    def test_metadata_in_markdown_table_regions(self):
        """Markdown shows table regions when >1 region detected."""
        sheet = {
            "sheet_name": "S1",
            "is_hidden": False,
            "merged_regions": [],
            "label_column": "A",
            "header_row_index": None,
            "unit_hint": None,
            "is_transposed": False,
            "table_regions": [
                {"start_row": 1, "end_row": 25},
                {"start_row": 28, "end_row": 50},
            ],
            "non_financial_rows": set(),
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": "Revenue", "number_format": "General"},
                    ],
                },
            ],
        }
        md = ParsingStage._structured_to_markdown({"sheets": [sheet]})
        assert "Table regions: 1-25, 28-50" in md

    def test_nf_marker_in_markdown(self):
        """Non-financial rows get 'NF' in the Bold column."""
        sheet = {
            "sheet_name": "S1",
            "is_hidden": False,
            "merged_regions": [],
            "non_financial_rows": {2},
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": "Revenue", "number_format": "General"},
                    ],
                },
                {
                    "row_index": 2,
                    "cells": [
                        {"ref": "A2", "value": "Source: filings", "number_format": "General"},
                    ],
                },
            ],
        }
        md = ParsingStage._structured_to_markdown({"sheets": [sheet]})
        lines = md.strip().split("\n")
        # Find the row for row_index=2
        row2_line = [l for l in lines if l.startswith("| 2")]
        assert len(row2_line) == 1
        assert "NF" in row2_line[0]

    def test_float_inf_nan_no_crash(self):
        """float('inf') and float('nan') should not crash markdown rendering."""
        sheet = {
            "sheet_name": "S1",
            "is_hidden": False,
            "merged_regions": [],
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": float("inf"), "number_format": "General"},
                    ],
                },
                {
                    "row_index": 2,
                    "cells": [
                        {"ref": "A2", "value": float("nan"), "number_format": "General"},
                    ],
                },
                {
                    "row_index": 3,
                    "cells": [
                        {"ref": "A3", "value": float("-inf"), "number_format": "General"},
                    ],
                },
            ],
        }
        md = ParsingStage._structured_to_markdown({"sheets": [sheet]})
        assert "inf" in md
        assert "nan" in md

    def test_detect_non_financial_rows_empty_input(self):
        """Empty rows -> empty set."""
        result = ParsingStage._detect_non_financial_rows([])
        assert result == set()

    def test_detect_non_financial_rows_all_financial(self):
        """Sheet with only financial data -> empty set."""
        rows = [
            {"row_index": 1, "cells": [{"ref": "A1", "value": "Revenue"}]},
            {"row_index": 2, "cells": [{"ref": "A2", "value": "COGS"}]},
        ]
        result = ParsingStage._detect_non_financial_rows(rows)
        assert result == set()


# ===========================================================================
# WS-1 Hardening: Theme/Indexed Color Resolution
# ===========================================================================


class TestThemeAndIndexedColors:
    """Tests for _normalize_color handling theme and indexed colors."""

    def test_theme_color_accent1_blue(self):
        """Theme=4 (accent1) resolves to the standard blue."""
        from unittest.mock import MagicMock

        from src.extraction.stages.parsing import _normalize_color

        color = MagicMock()
        color.type = "theme"
        color.theme = 4
        color.tint = 0.0
        result = _normalize_color(color)
        assert result == "4472c4"

    def test_theme_color_dk1_black(self):
        """Theme=0 (dk1) resolves to black."""
        from unittest.mock import MagicMock

        from src.extraction.stages.parsing import _normalize_color

        color = MagicMock()
        color.type = "theme"
        color.theme = 0
        color.tint = 0.0
        result = _normalize_color(color)
        assert result == "000000"

    def test_theme_color_with_positive_tint(self):
        """Theme=4 with positive tint lightens the color."""
        from unittest.mock import MagicMock

        from src.extraction.stages.parsing import _normalize_color

        color = MagicMock()
        color.type = "theme"
        color.theme = 4  # 4472C4
        color.tint = 0.4
        result = _normalize_color(color)
        assert result is not None
        # Lightened: each channel moves toward 255
        # R: 0x44 + (255-0x44)*0.4 = 68 + 74.8 = 142 → ~8e
        # G: 0x72 + (255-0x72)*0.4 = 114 + 56.4 = 170 → ~aa
        # B: 0xC4 + (255-0xC4)*0.4 = 196 + 23.6 = 219 → ~db
        r = int(result[0:2], 16)
        g = int(result[2:4], 16)
        b = int(result[4:6], 16)
        assert r > 0x44  # lighter than base red
        assert g > 0x72  # lighter than base green
        assert b > 0xC4  # lighter than base blue

    def test_theme_color_with_negative_tint(self):
        """Theme=1 (white) with negative tint darkens."""
        from unittest.mock import MagicMock

        from src.extraction.stages.parsing import _normalize_color

        color = MagicMock()
        color.type = "theme"
        color.theme = 1  # FFFFFF
        color.tint = -0.5
        result = _normalize_color(color)
        assert result is not None
        # Darkened: each channel * 0.5 → 127 = 7f
        r = int(result[0:2], 16)
        assert r < 0xFF  # darker than white
        assert r == int(255 * 0.5)  # 127

    def test_indexed_color_blue(self):
        """Indexed=4 resolves to blue."""
        from unittest.mock import MagicMock

        from src.extraction.stages.parsing import _normalize_color

        color = MagicMock()
        color.type = "indexed"
        color.indexed = 4
        result = _normalize_color(color)
        assert result == "0000ff"

    def test_indexed_color_out_of_range(self):
        """Indexed=50 (beyond our 8-entry palette) returns None."""
        from unittest.mock import MagicMock

        from src.extraction.stages.parsing import _normalize_color

        color = MagicMock()
        color.type = "indexed"
        color.indexed = 50
        result = _normalize_color(color)
        assert result is None

    def test_theme_accent1_detected_as_blue(self):
        """Theme accent1 blue (4472C4) passes _is_blue_font after relaxed thresholds."""
        from src.extraction.stages.parsing import _is_blue_font

        # 4472C4: r=0x44(68), g=0x72(114), b=0xC4(196)
        # With relaxed thresholds: b>=0x80, r<0x80, g<0x80, b>r, b>g
        assert _is_blue_font("4472c4") is True


# ===========================================================================
# WS-1 Hardening: Relaxed Blue Font Thresholds
# ===========================================================================


class TestRelaxedBlueThresholds:
    """Tests for the relaxed _is_blue_font thresholds."""

    def test_navy_blue_000080(self):
        """Navy blue (000080) now detected as blue (b >= 0x80)."""
        from src.extraction.stages.parsing import _is_blue_font

        assert _is_blue_font("000080") is True

    def test_accent1_blue_4472c4(self):
        """Theme accent1 (4472C4) now detected as blue (r < 0x80)."""
        from src.extraction.stages.parsing import _is_blue_font

        assert _is_blue_font("4472c4") is True

    def test_dark_blue_003399(self):
        """Dark blue (003399) still detected (regression check)."""
        from src.extraction.stages.parsing import _is_blue_font

        assert _is_blue_font("003399") is True

    def test_pure_blue_0000ff(self):
        """Pure blue still works (regression check)."""
        from src.extraction.stages.parsing import _is_blue_font

        assert _is_blue_font("0000ff") is True

    def test_purple_not_blue(self):
        """Purple (800080) is not blue — b not dominant over r."""
        from src.extraction.stages.parsing import _is_blue_font

        # r=0x80, g=0, b=0x80 → b not > r
        assert _is_blue_font("800080") is False

    def test_teal_not_blue(self):
        """Teal (008080) is not blue — g == b, blue not dominant."""
        from src.extraction.stages.parsing import _is_blue_font

        # r=0, g=0x80, b=0x80 → b not > g
        assert _is_blue_font("008080") is False

    def test_red_not_blue(self):
        """Red is still correctly not blue."""
        from src.extraction.stages.parsing import _is_blue_font

        assert _is_blue_font("ff0000") is False


# ===========================================================================
# WS-1 Hardening: Underline Detection
# ===========================================================================


class TestUnderlineDetection:
    """Tests for is_underline field and its use in section boundaries."""

    def test_underline_detected_in_cell(self):
        """Font with underline='single' → is_underline: True."""
        xlsx_bytes = _make_minimal_xlsx(
            data={"S1": [["Header"], [100]]},
            underline_cells={("S1", 1, 1)},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx_bytes)
        cells = result["sheets"][0]["rows"][0]["cells"]
        assert cells[0]["is_underline"] is True

    def test_no_underline_default(self):
        """Normal font → is_underline: False."""
        xlsx_bytes = _make_minimal_xlsx(
            data={"S1": [["Normal"], [100]]},
        )
        result = ParsingStage._excel_to_structured_repr(xlsx_bytes)
        cells = result["sheets"][0]["rows"][0]["cells"]
        assert cells[0]["is_underline"] is False

    def test_underline_section_boundary(self):
        """Underlined header after gap → detected as section boundary."""
        # Row 1: data, Row 5: underlined header (gap of 3)
        rows = [
            {
                "row_index": 1,
                "cells": [
                    {"ref": "A1", "value": "Revenue", "is_bold": False, "is_underline": False},
                ],
            },
            {
                "row_index": 5,
                "cells": [
                    {
                        "ref": "A5",
                        "value": "Balance Sheet",
                        "is_bold": False,
                        "is_underline": True,
                        "has_border_bottom": False,
                    },
                ],
            },
        ]
        boundaries = ParsingStage._detect_section_boundaries(rows)
        assert len(boundaries) == 1
        assert boundaries[0]["label"] == "Balance Sheet"
        assert boundaries[0]["row_index"] == 5


# ===========================================================================
# WS-1 Hardening: Section Boundary Cap
# ===========================================================================


class TestSectionBoundaryCap:
    """Tests for section boundary truncation at max 20."""

    def test_section_boundary_cap_at_20(self):
        """30 bold-header rows with gaps → only 20 boundaries returned."""
        rows = []
        for i in range(30):
            row_idx = i * 3 + 1  # gap of 2 between each
            rows.append(
                {
                    "row_index": row_idx,
                    "cells": [
                        {
                            "ref": f"A{row_idx}",
                            "value": f"Section {i + 1}",
                            "is_bold": True,
                            "is_underline": False,
                            "has_border_bottom": False,
                        }
                    ],
                }
            )
        boundaries = ParsingStage._detect_section_boundaries(rows)
        assert len(boundaries) == 20

    def test_section_boundary_truncated_flag(self):
        """Truncated list has 'truncated': True on last entry."""
        rows = []
        for i in range(25):
            row_idx = i * 3 + 1
            rows.append(
                {
                    "row_index": row_idx,
                    "cells": [
                        {
                            "ref": f"A{row_idx}",
                            "value": f"Section {i + 1}",
                            "is_bold": True,
                            "is_underline": False,
                            "has_border_bottom": False,
                        }
                    ],
                }
            )
        boundaries = ParsingStage._detect_section_boundaries(rows)
        assert len(boundaries) == 20
        assert boundaries[-1].get("truncated") is True
        # First entry should NOT have truncated flag
        assert "truncated" not in boundaries[0]


# ===========================================================================
# WS-1 Hardening: Label Column Reclassification
# ===========================================================================


class TestLabelColumnReclassification:
    """Tests for _reclassify_cell_types post-processing."""

    def test_cell_type_label_in_column_e(self):
        """String cells in detected label_column=E get reclassified to 'label'."""
        sheet = {
            "label_column": "E",
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "E1", "value": "Revenue", "cell_type": "value"},
                        {"ref": "F1", "value": 100, "cell_type": "value"},
                    ],
                },
            ],
        }
        ParsingStage._reclassify_cell_types(sheet)
        assert sheet["rows"][0]["cells"][0]["cell_type"] == "label"
        # Numeric cell in F should stay "value"
        assert sheet["rows"][0]["cells"][1]["cell_type"] == "value"

    def test_cell_type_no_reclassify_default_columns(self):
        """label_column=A → no reclassification needed (default handles it)."""
        sheet = {
            "label_column": "A",
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": "Revenue", "cell_type": "label"},
                    ],
                },
            ],
        }
        ParsingStage._reclassify_cell_types(sheet)
        assert sheet["rows"][0]["cells"][0]["cell_type"] == "label"

    def test_cell_type_reclassify_skips_numeric(self):
        """Numeric cell in label column stays 'value'."""
        sheet = {
            "label_column": "E",
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "E1", "value": 42.0, "cell_type": "value"},
                    ],
                },
            ],
        }
        ParsingStage._reclassify_cell_types(sheet)
        assert sheet["rows"][0]["cells"][0]["cell_type"] == "value"

    def test_reclassify_no_label_column(self):
        """No label_column → no-op."""
        sheet = {
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "D1", "value": "Label", "cell_type": "value"},
                    ],
                },
            ],
        }
        ParsingStage._reclassify_cell_types(sheet)
        assert sheet["rows"][0]["cells"][0]["cell_type"] == "value"


# ===========================================================================
# WS-1 Hardening: Tint Application
# ===========================================================================


class TestTintApplication:
    """Tests for the _apply_tint helper."""

    def test_zero_tint_no_change(self):
        """Tint=0 returns the same color."""
        from src.extraction.stages.parsing import _apply_tint

        assert _apply_tint("4472c4", 0.0) == "4472c4"

    def test_positive_tint_lightens(self):
        """Positive tint moves channels toward 255."""
        from src.extraction.stages.parsing import _apply_tint

        result = _apply_tint("000000", 1.0)
        # Full positive tint on black → white
        assert result == "ffffff"

    def test_negative_tint_darkens(self):
        """Negative tint moves channels toward 0."""
        from src.extraction.stages.parsing import _apply_tint

        result = _apply_tint("ffffff", -1.0)
        # Full negative tint on white → black
        assert result == "000000"


# ============================================================================
# LABEL COLUMN RECLASSIFICATION TESTS
# ============================================================================


class TestLabelColumnDeriveCellType:
    """Tests for _reclassify_cell_types and _derive_cell_type with label_col_index."""

    def test_derive_cell_type_with_label_col_index(self):
        """Explicit label_col_index allows string cells in that column to be 'label'."""
        from src.extraction.stages.parsing import _derive_cell_type

        cell = {"value": "Revenue"}
        # Column D (index 3) — without label_col_index, this would be "value"
        assert _derive_cell_type(cell, 3) == "value"
        # With label_col_index=3, column D is now a label column
        assert _derive_cell_type(cell, 3, label_col_index=3) == "label"

    def test_derive_cell_type_left_of_label_col(self):
        """Columns left of label_col_index also get 'label' type."""
        from src.extraction.stages.parsing import _derive_cell_type

        cell = {"value": "Category"}
        # label_col_index=4 (column E); column C (index 2) is left of it
        assert _derive_cell_type(cell, 2, label_col_index=4) == "label"

    def test_derive_cell_type_right_of_label_col(self):
        """Columns right of label_col_index stay 'value'."""
        from src.extraction.stages.parsing import _derive_cell_type

        cell = {"value": "Some text"}
        # label_col_index=3 (column D); column F (index 5) is right of it
        assert _derive_cell_type(cell, 5, label_col_index=3) == "value"

    def test_derive_cell_type_formula_takes_priority(self):
        """formula cells keep 'formula' type regardless of label_col_index."""
        from src.extraction.stages.parsing import _derive_cell_type

        cell = {"value": "Revenue", "formula": "=SUM(A1:A5)"}
        assert _derive_cell_type(cell, 0, label_col_index=3) == "formula"

    def test_reclassify_noop_for_column_a(self):
        """No change when label_column is 'A' (default heuristic covers A-C)."""
        sheet = {
            "label_column": "A",
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": "Revenue", "cell_type": "label"},
                        {"ref": "B1", "value": 100, "cell_type": "value"},
                    ],
                }
            ],
        }
        ParsingStage._reclassify_cell_types(sheet)
        assert sheet["rows"][0]["cells"][0]["cell_type"] == "label"
        assert sheet["rows"][0]["cells"][1]["cell_type"] == "value"

    def test_reclassify_labels_in_column_d(self):
        """Labels in col D get 'label' type after reclassification."""
        sheet = {
            "label_column": "D",
            "rows": [
                {
                    "row_index": 1,
                    "cells": [
                        {"ref": "A1", "value": "Category", "cell_type": "value"},
                        {"ref": "B1", "value": "Sub", "cell_type": "value"},
                        {"ref": "C1", "value": "Detail", "cell_type": "value"},
                        {"ref": "D1", "value": "Revenue", "cell_type": "value"},
                        {"ref": "E1", "value": 100, "cell_type": "value"},
                    ],
                }
            ],
        }
        ParsingStage._reclassify_cell_types(sheet)

        # Columns A-D should now be "label" (all at or left of label column D)
        assert sheet["rows"][0]["cells"][0]["cell_type"] == "label"  # A
        assert sheet["rows"][0]["cells"][1]["cell_type"] == "label"  # B
        assert sheet["rows"][0]["cells"][2]["cell_type"] == "label"  # C
        assert sheet["rows"][0]["cells"][3]["cell_type"] == "label"  # D
        # Column E should remain "value" (right of label column)
        assert sheet["rows"][0]["cells"][4]["cell_type"] == "value"  # E

    def test_col_letter_to_index(self):
        """Verify column letter → 0-based index conversion."""
        from src.extraction.stages.parsing import _col_letter_to_index

        assert _col_letter_to_index("A") == 0
        assert _col_letter_to_index("B") == 1
        assert _col_letter_to_index("C") == 2
        assert _col_letter_to_index("Z") == 25
        assert _col_letter_to_index("AA") == 26
        assert _col_letter_to_index("AB") == 27
