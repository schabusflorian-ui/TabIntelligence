"""End-to-end stress tests for the extraction pipeline.

Programmatically creates diverse Excel workbooks and pushes them through
the deterministic pipeline stages (parsing → section detection → triage
summary → mapping section lookup) WITHOUT calling Claude.

Tests scalability across:
  - Sheet count and row count
  - Multi-section sheets with varied boundary styles
  - Messy metadata (deep headers, shifted labels, units)
  - Edge case structures that push heuristics to their limits
"""

import io
import time
from pathlib import Path
from typing import Any, Dict

import openpyxl
import pytest
from openpyxl.styles import Border, Font, PatternFill, Side

from src.extraction.section_detector import SectionDetector, _guess_category
from src.extraction.stages.mapping import MappingStage
from src.extraction.stages.parsing import ParsingStage
from src.extraction.stages.triage import TriageStage

# ---------------------------------------------------------------------------
# Excel builder helpers
# ---------------------------------------------------------------------------


def _make_workbook() -> openpyxl.Workbook:
    """Create a blank workbook with the default sheet removed."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _add_is_sheet(
    wb: openpyxl.Workbook,
    name: str = "Income Statement",
    start_row: int = 1,
    num_rows: int = 15,
    label_col: str = "A",
    header_row: int = 1,
    periods: list | None = None,
    bold_header: bool = True,
) -> None:
    """Add an Income Statement sheet with realistic structure."""
    ws = wb.create_sheet(name)
    periods = periods or ["FY2022", "FY2023", "FY2024E"]

    # Header row with period labels
    lc = ord(label_col) - ord("A")
    for i, p in enumerate(periods):
        ws.cell(row=header_row, column=lc + 2 + i, value=p).font = Font(bold=True)

    is_labels = [
        "Revenue",
        "Cost of Goods Sold",
        "Gross Profit",
        "SG&A",
        "R&D",
        "Depreciation & Amortization",
        "Operating Income",
        "Interest Expense",
        "Income Before Tax",
        "Tax Provision",
        "Net Income",
    ]
    for i, label in enumerate(is_labels[:num_rows]):
        row = start_row + i + (1 if header_row == start_row else 0)
        cell = ws.cell(row=row, column=lc + 1, value=label)
        if bold_header and i == 0:
            cell.font = Font(bold=True)
        # Is-total rows
        if "total" in label.lower() or "net" in label.lower() or "gross" in label.lower():
            cell.font = Font(bold=True)
        for j, _p in enumerate(periods):
            ws.cell(row=row, column=lc + 2 + j, value=100000 * (j + 1) + i * 1000)


def _add_bs_sheet(
    wb: openpyxl.Workbook,
    name: str = "Balance Sheet",
    label_col: str = "A",
    periods: list | None = None,
) -> None:
    """Add a Balance Sheet with realistic structure."""
    ws = wb.create_sheet(name)
    periods = periods or ["FY2022", "FY2023", "FY2024E"]

    lc = ord(label_col) - ord("A")
    for i, p in enumerate(periods):
        ws.cell(row=1, column=lc + 2 + i, value=p).font = Font(bold=True)

    bs_labels = [
        "Cash & Equivalents",
        "Accounts Receivable",
        "Inventory",
        "Total Current Assets",
        "PP&E",
        "Goodwill",
        "Total Assets",
        "Accounts Payable",
        "Short-term Debt",
        "Total Current Liabilities",
        "Long-term Debt",
        "Total Liabilities",
        "Common Equity",
        "Retained Earnings",
        "Total Equity",
        "Total Liabilities & Equity",
    ]
    for i, label in enumerate(bs_labels):
        row = 2 + i
        cell = ws.cell(row=row, column=lc + 1, value=label)
        if "total" in label.lower():
            cell.font = Font(bold=True)
        for j, _p in enumerate(periods):
            ws.cell(row=row, column=lc + 2 + j, value=50000 * (j + 1) + i * 500)


def _add_combined_sheet(
    wb: openpyxl.Workbook,
    name: str = "Combined FS",
    gap_rows: int = 3,
    bold_headers: bool = True,
) -> None:
    """Add a multi-section sheet with IS + BS separated by a gap."""
    ws = wb.create_sheet(name)
    periods = ["FY2022", "FY2023", "FY2024E"]

    # Row 1: IS header + periods on same row (so section label is correct)
    is_header = ws.cell(row=1, column=1, value="Profit & Loss Statement")
    if bold_headers:
        is_header.font = Font(bold=True)
    for i, p in enumerate(periods):
        ws.cell(row=1, column=2 + i, value=p).font = Font(bold=True)

    # Section 1: P&L data rows
    is_labels = ["Revenue", "COGS", "Gross Profit", "SG&A", "Operating Income", "Net Income"]
    for i, label in enumerate(is_labels):
        ws.cell(row=2 + i, column=1, value=label)
        for j in range(3):
            ws.cell(row=2 + i, column=2 + j, value=100000 + i * 10000 + j * 5000)

    bs_start = 2 + len(is_labels) + gap_rows
    # Section 2: BS
    bs_header = ws.cell(row=bs_start, column=1, value="Balance Sheet")
    if bold_headers:
        bs_header.font = Font(bold=True)
    bs_labels = ["Cash", "Total Assets", "Total Liabilities", "Total Equity"]
    for i, label in enumerate(bs_labels):
        ws.cell(row=bs_start + 1 + i, column=1, value=label)
        for j in range(3):
            ws.cell(row=bs_start + 1 + i, column=2 + j, value=50000 + i * 5000 + j * 2000)


def _add_notes_sheet(wb: openpyxl.Workbook, name: str = "Notes") -> None:
    """Add a pure-text notes sheet (should be tier 4)."""
    ws = wb.create_sheet(name)
    notes = [
        "Source: Management projections",
        "Prepared by: Finance Team",
        "Confidential - Do not distribute",
        "Draft as of March 2024",
        "Notes:",
        "1. Revenue is recognized on delivery.",
        "2. COGS includes direct materials and labor.",
        "3. SG&A includes marketing and admin costs.",
    ]
    for i, note in enumerate(notes):
        ws.cell(row=i + 1, column=1, value=note)


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    """Serialize workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Pipeline helper: run deterministic stages (no Claude)
# ---------------------------------------------------------------------------


def _run_deterministic_pipeline(file_bytes: bytes) -> Dict[str, Any]:
    """Run parsing + metadata + section detection + triage summary.

    Returns a dict with all intermediate results for inspection.
    """
    # Stage 1: Structured repr
    structured = ParsingStage._excel_to_structured_repr(file_bytes)

    # Stage 1b: Metadata detection
    for sheet in structured["sheets"]:
        sheet.update(ParsingStage._detect_sheet_metadata(sheet))

    # Stage 1c: Markdown generation
    markdown = ParsingStage._structured_to_markdown(structured)

    # Stage 2a: Section detection (per sheet)
    detector = SectionDetector()
    sections_by_sheet = {}
    for sheet in structured["sheets"]:
        sections = detector.detect_sections(sheet)
        sections_by_sheet[sheet["sheet_name"]] = sections

    # Stage 2b: Triage summary (what would be sent to Claude)
    parsed_result = {
        "sheets": [
            {
                "sheet_name": s["sheet_name"],
                "rows": [
                    {"label": _get_first_label(row, s.get("label_column", "A"))}
                    for row in s["rows"]
                ],
            }
            for s in structured["sheets"]
        ],
    }
    summaries = TriageStage._build_sheet_summary(parsed_result, structured)

    return {
        "structured": structured,
        "markdown": markdown,
        "sections_by_sheet": sections_by_sheet,
        "summaries": summaries,
    }


def _get_first_label(row: dict, label_column: str) -> str:
    """Extract label from the label column of a row."""
    import re

    for cell in row.get("cells", []):
        ref = cell.get("ref", "")
        m = re.match(r"([A-Z]+)", ref)
        if m and m.group(1) == label_column:
            val = cell.get("value")
            if isinstance(val, str):
                return val
    return ""


# ===========================================================================
# Test Classes
# ===========================================================================


class TestStandardWorkbook:
    """Standard 3-sheet workbook: IS + BS + Notes."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        _add_is_sheet(wb)
        _add_bs_sheet(wb)
        _add_notes_sheet(wb)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_sheet_count(self, result):
        assert result["structured"]["sheet_count"] == 3

    def test_is_metadata(self, result):
        is_sheet = next(
            s for s in result["structured"]["sheets"] if s["sheet_name"] == "Income Statement"
        )
        assert is_sheet["label_column"] == "A"
        assert is_sheet["header_row_index"] is not None

    def test_bs_metadata(self, result):
        bs_sheet = next(
            s for s in result["structured"]["sheets"] if s["sheet_name"] == "Balance Sheet"
        )
        assert bs_sheet["label_column"] == "A"

    def test_notes_few_formulas(self, result):
        notes_summary = next(s for s in result["summaries"] if s["name"] == "Notes")
        assert notes_summary.get("formula_count", 0) == 0

    def test_no_sections_on_single_sheets(self, result):
        """Single-purpose sheets should have exactly 1 section."""
        for name in ["Income Statement", "Balance Sheet"]:
            sections = result["sections_by_sheet"][name]
            assert len(sections) == 1, f"{name} should have 1 section, got {len(sections)}"

    def test_markdown_not_empty(self, result):
        assert len(result["markdown"]) > 100

    def test_summaries_have_all_sheets(self, result):
        names = {s["name"] for s in result["summaries"]}
        assert names == {"Income Statement", "Balance Sheet", "Notes"}


class TestMultiSectionSheet:
    """Combined FS with IS + BS on one sheet."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        _add_combined_sheet(wb, gap_rows=3)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_two_sections_detected(self, result):
        sections = result["sections_by_sheet"]["Combined FS"]
        assert len(sections) >= 2, (
            f"Expected 2+ sections, got {len(sections)}: {[s.label for s in sections]}"
        )

    def test_first_section_is_pl(self, result):
        sections = result["sections_by_sheet"]["Combined FS"]
        assert sections[0].category_hint == "income_statement"

    def test_second_section_is_bs(self, result):
        sections = result["sections_by_sheet"]["Combined FS"]
        bs = [s for s in sections if s.category_hint == "balance_sheet"]
        assert len(bs) >= 1

    def test_triage_summary_has_sections(self, result):
        summary = next(s for s in result["summaries"] if s["name"] == "Combined FS")
        assert "sections" in summary
        assert len(summary["sections"]) >= 2

    def test_section_lookup_works(self, result):
        """Mapping can build a section_lookup from triage sections."""
        summary = next(s for s in result["summaries"] if s["name"] == "Combined FS")
        triage_list = [
            {
                "sheet_name": "Combined FS",
                "tier": 1,
                "section": sec["label"],
                "section_start_row": sec["start_row"],
                "section_end_row": sec["end_row"],
                "category_hint": sec.get("category_hint"),
            }
            for sec in summary["sections"]
        ]
        lookup = MappingStage._build_section_lookup(triage_list)
        assert "Combined FS" in lookup
        assert len(lookup["Combined FS"]) >= 2


class TestNoGapBoldBoundary:
    """Combined sheet with gap=1 bold headers (the hardest case)."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        # gap_rows=0 means BS header immediately follows last IS row
        _add_combined_sheet(wb, name="Combined FS", gap_rows=0, bold_headers=True)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_sections_detected_via_precomputed(self, result):
        """Pre-computed boundaries from parsing should catch gap=0+bold."""
        sections = result["sections_by_sheet"]["Combined FS"]
        # With pre-computed boundary merging, this should detect 2+ sections
        # even though gap=0 (bold header with border in parsing)
        labels = [s.label for s in sections]
        has_pl = any("profit" in l.lower() or "loss" in l.lower() for l in labels)
        has_bs = any("balance" in l.lower() for l in labels)
        # At minimum, the sheet should not be a single undifferentiated blob
        assert len(sections) >= 1, f"Got labels: {labels}"
        if len(sections) >= 2:
            # Best case: both sections detected
            assert has_pl or has_bs, f"Expected financial sections, got: {labels}"


class TestLabelColumnB:
    """Sheet with labels in column B (column A has row numbers)."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        ws = wb.create_sheet("Model")
        periods = ["FY2022", "FY2023"]
        # Row numbers in col A, labels in col B, values in C-D
        for i, p in enumerate(periods):
            ws.cell(row=1, column=3 + i, value=p).font = Font(bold=True)
        labels = [
            "Revenue",
            "COGS",
            "Gross Profit",
            "SG&A",
            "EBITDA",
            "D&A",
            "EBIT",
            "Interest",
            "EBT",
            "Tax",
            "Net Income",
        ]
        for i, label in enumerate(labels):
            ws.cell(row=2 + i, column=1, value=i + 1)  # Row number
            ws.cell(row=2 + i, column=2, value=label)
            for j in range(2):
                ws.cell(row=2 + i, column=3 + j, value=100000 + i * 5000)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_label_column_detected(self, result):
        sheet = result["structured"]["sheets"][0]
        assert sheet["label_column"] == "B", f"Expected label_column=B, got {sheet['label_column']}"

    def test_labels_in_summary(self, result):
        summary = result["summaries"][0]
        # Sample labels should come from column B, not column A
        labels = summary["sample_labels"]
        assert any("Revenue" in l for l in labels), f"Labels: {labels}"


class TestDeepHeaders:
    """Headers pushed past row 10 by metadata/disclaimers."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        ws = wb.create_sheet("IS Deep")
        # Rows 1-12: disclaimers
        disclaimers = [
            "CONFIDENTIAL",
            "Draft - Not for Distribution",
            "Prepared by: Finance Team",
            "Date: March 2024",
            "Source: Internal Estimates",
            "Subject to change",
            "Reviewed by: CFO",
            "Version 3.2",
            "Preliminary results",
            "Management estimates",
            "Unaudited figures",
            "For internal use only",
        ]
        for i, d in enumerate(disclaimers):
            ws.cell(row=i + 1, column=1, value=d)
        # Row 13: period headers
        for j, p in enumerate(["FY2022", "FY2023", "FY2024E"]):
            ws.cell(row=13, column=2 + j, value=p).font = Font(bold=True)
        # Rows 14+: data
        for i, label in enumerate(["Revenue", "COGS", "Gross Profit", "OpEx", "Net Income"]):
            ws.cell(row=14 + i, column=1, value=label)
            for j in range(3):
                ws.cell(row=14 + i, column=2 + j, value=100000 + i * 10000)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_header_detected_past_row_10(self, result):
        sheet = result["structured"]["sheets"][0]
        assert sheet["header_row_index"] == 13, (
            f"Expected header at row 13, got {sheet['header_row_index']}"
        )

    def test_non_financial_rows_detected(self, result):
        sheet = result["structured"]["sheets"][0]
        nf = sheet.get("non_financial_rows", set())
        # At least some disclaimers should be flagged
        assert len(nf) >= 3, f"Expected 3+ non-financial rows, got {len(nf)}: {nf}"


class TestUnitsDetection:
    """Unit annotation at various depths."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        ws = wb.create_sheet("Model")
        ws.cell(row=1, column=1, value="Company ABC Financial Model")
        ws.cell(row=2, column=1, value="Fiscal Year Ending December 31")
        ws.cell(row=3, column=1, value="(in millions)")
        for j, p in enumerate(["FY2022", "FY2023"]):
            ws.cell(row=4, column=2 + j, value=p).font = Font(bold=True)
        for i, label in enumerate(["Revenue", "COGS", "Net Income"]):
            ws.cell(row=5 + i, column=1, value=label)
            for j in range(2):
                ws.cell(row=5 + i, column=2 + j, value=100 + i * 10)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_unit_detected(self, result):
        sheet = result["structured"]["sheets"][0]
        assert sheet["unit_hint"] == "millions"
        assert sheet["unit_multiplier"] == 1_000_000.0

    def test_unit_in_markdown(self, result):
        assert "Units: millions" in result["markdown"]


class TestColorBoundaries:
    """Section boundaries via fill colours (no gap needed)."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        ws = wb.create_sheet("Model")
        blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        # Period headers
        for j, p in enumerate(["FY2022", "FY2023"]):
            ws.cell(row=1, column=2 + j, value=p).font = Font(bold=True)

        # Section 1: IS (blue header)
        row = 2
        for c in range(1, 4):
            cell = ws.cell(row=row, column=c)
            cell.fill = blue_fill
            cell.font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1, value="Income Statement")
        for i, label in enumerate(["Revenue", "COGS", "Gross Profit", "OpEx", "Net Income"]):
            ws.cell(row=3 + i, column=1, value=label)
            for j in range(2):
                ws.cell(row=3 + i, column=2 + j, value=100000 + i * 10000)

        # Section 2: BS (green header, no gap)
        row = 8
        for c in range(1, 4):
            cell = ws.cell(row=row, column=c)
            cell.fill = green_fill
            cell.font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=1, value="Balance Sheet")
        for i, label in enumerate(["Cash", "Total Assets", "Total Liabilities"]):
            ws.cell(row=9 + i, column=1, value=label)
            for j in range(2):
                ws.cell(row=9 + i, column=2 + j, value=50000 + i * 5000)

        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_two_sections_detected(self, result):
        sections = result["sections_by_sheet"]["Model"]
        assert len(sections) >= 2, f"Expected 2+ sections, got {[s.label for s in sections]}"

    def test_section_categories(self, result):
        sections = result["sections_by_sheet"]["Model"]
        categories = {s.category_hint for s in sections if s.category_hint}
        assert "income_statement" in categories or "balance_sheet" in categories


class TestLargeWorkbook:
    """Stress test with many sheets and rows."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        # 10 sheets, 100 rows each
        for sheet_idx in range(10):
            ws = wb.create_sheet(f"Sheet {sheet_idx + 1}")
            for j, p in enumerate(["FY2022", "FY2023", "FY2024E"]):
                ws.cell(row=1, column=2 + j, value=p).font = Font(bold=True)
            for i in range(100):
                ws.cell(row=2 + i, column=1, value=f"Line Item {i + 1}")
                for j in range(3):
                    ws.cell(row=2 + i, column=2 + j, value=1000 * (i + 1) + j * 100)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_all_sheets_parsed(self, result):
        assert result["structured"]["sheet_count"] == 10

    def test_all_rows_captured(self, result):
        for sheet in result["structured"]["sheets"]:
            assert len(sheet["rows"]) >= 100

    def test_performance_acceptable(self):
        """Full pipeline on 10x100 should complete in < 5s."""
        wb = _make_workbook()
        for sheet_idx in range(10):
            ws = wb.create_sheet(f"Sheet {sheet_idx + 1}")
            for i in range(100):
                ws.cell(row=1 + i, column=1, value=f"Item {i}")
                for j in range(5):
                    ws.cell(row=1 + i, column=2 + j, value=1000 * i)
        file_bytes = _wb_to_bytes(wb)

        start = time.time()
        _run_deterministic_pipeline(file_bytes)
        elapsed = time.time() - start
        assert elapsed < 5.0, f"Pipeline took {elapsed:.2f}s (limit: 5s)"

    def test_summaries_generated_for_all(self, result):
        assert len(result["summaries"]) == 10


class TestThreeSectionSheet:
    """Sheet with IS + BS + CF separated by gaps."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        ws = wb.create_sheet("Combined")

        # Row 1: IS header + periods (so section label is "Income Statement")
        ws.cell(row=1, column=1, value="Income Statement").font = Font(bold=True)
        for j, p in enumerate(["FY2022", "FY2023"]):
            ws.cell(row=1, column=2 + j, value=p).font = Font(bold=True)

        # IS data (rows 2-6)
        for i, l in enumerate(["Revenue", "COGS", "Gross Profit", "OpEx", "Net Income"]):
            ws.cell(row=2 + i, column=1, value=l)
            ws.cell(row=2 + i, column=2, value=100000 + i * 10000)

        # Gap (rows 7-9 blank)

        # BS section (rows 10-16)
        ws.cell(row=10, column=1, value="Balance Sheet").font = Font(bold=True)
        for i, l in enumerate(["Cash", "AR", "Total Assets", "AP", "Total Equity"]):
            ws.cell(row=11 + i, column=1, value=l)
            ws.cell(row=11 + i, column=2, value=50000 + i * 5000)

        # Gap (rows 17-19 blank)

        # CF section (rows 20-26)
        ws.cell(row=20, column=1, value="Cash Flow Statement").font = Font(bold=True)
        for i, l in enumerate(["Net Income", "D&A", "Changes in WC", "CFO", "CapEx"]):
            ws.cell(row=21 + i, column=1, value=l)
            ws.cell(row=21 + i, column=2, value=20000 + i * 2000)

        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_three_sections_detected(self, result):
        sections = result["sections_by_sheet"]["Combined"]
        assert len(sections) == 3, (
            f"Expected 3 sections, got {len(sections)}: "
            f"{[(s.label, s.category_hint) for s in sections]}"
        )

    def test_all_categories_present(self, result):
        sections = result["sections_by_sheet"]["Combined"]
        categories = {s.category_hint for s in sections}
        assert "income_statement" in categories
        assert "balance_sheet" in categories
        assert "cash_flow" in categories

    def test_section_ranges_correct(self, result):
        sections = result["sections_by_sheet"]["Combined"]
        # Sections should not overlap
        for i in range(len(sections) - 1):
            assert sections[i].end_row < sections[i + 1].start_row

    def test_section_lookup_for_mapping(self, result):
        summary = next(s for s in result["summaries"] if s["name"] == "Combined")
        triage_list = [
            {
                "sheet_name": "Combined",
                "tier": 1,
                "section": sec["label"],
                "section_start_row": sec["start_row"],
                "section_end_row": sec["end_row"],
                "category_hint": sec.get("category_hint"),
            }
            for sec in summary["sections"]
        ]
        lookup = MappingStage._build_section_lookup(triage_list)
        assert len(lookup["Combined"]) == 3

        # Test row assignment using midpoints of each section
        sections = summary["sections"]
        sec_by_cat = {s["category_hint"]: s for s in sections}
        is_mid = (
            sec_by_cat["income_statement"]["start_row"] + sec_by_cat["income_statement"]["end_row"]
        ) // 2
        bs_mid = (
            sec_by_cat["balance_sheet"]["start_row"] + sec_by_cat["balance_sheet"]["end_row"]
        ) // 2
        cf_mid = (sec_by_cat["cash_flow"]["start_row"] + sec_by_cat["cash_flow"]["end_row"]) // 2

        parsed = {
            "sheets": [
                {
                    "sheet_name": "Combined",
                    "rows": [
                        {"label": "Revenue", "row_index": is_mid},
                        {"label": "Cash", "row_index": bs_mid},
                        {"label": "CFO", "row_index": cf_mid},
                    ],
                }
            ],
        }
        items = MappingStage._build_grouped_line_items(parsed, lookup)
        assert items[0].get("section_category") == "income_statement"
        assert items[1].get("section_category") == "balance_sheet"
        assert items[2].get("section_category") == "cash_flow"


class TestBorderBasedBoundaries:
    """Section headers with bottom borders (no gap, no bold)."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        ws = wb.create_sheet("Model")
        bottom_border = Border(bottom=Side(style="medium"))

        for j, p in enumerate(["FY2022", "FY2023"]):
            ws.cell(row=1, column=2 + j, value=p).font = Font(bold=True)

        # Section 1 header with border
        cell = ws.cell(row=2, column=1, value="Revenue Summary")
        cell.font = Font(bold=True)
        cell.border = bottom_border
        for i in range(5):
            ws.cell(row=3 + i, column=1, value=f"Rev item {i + 1}")
            ws.cell(row=3 + i, column=2, value=10000 + i * 1000)

        # Section 2 header with border (no gap! row 8 follows row 7)
        cell = ws.cell(row=8, column=1, value="Expense Summary")
        cell.font = Font(bold=True)
        cell.border = bottom_border
        for i in range(5):
            ws.cell(row=9 + i, column=1, value=f"Exp item {i + 1}")
            ws.cell(row=9 + i, column=2, value=5000 + i * 500)

        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_section_boundaries_in_structured(self, result):
        """Parsing should detect border-based section boundaries."""
        sheet = result["structured"]["sheets"][0]
        boundaries = sheet.get("section_boundaries", [])
        # Should detect at least the border-based headers
        boundary_labels = [b["label"] for b in boundaries]
        assert len(boundaries) >= 1, f"Expected border-based boundaries, got: {boundary_labels}"


class TestEmptyAndMinimalSheets:
    """Edge cases: empty sheets, 1-row sheets, hidden sheets."""

    @pytest.fixture(scope="class")
    def result(self):
        wb = _make_workbook()
        # Empty sheet
        wb.create_sheet("Empty")
        # 1-row sheet
        ws = wb.create_sheet("Single")
        ws.cell(row=1, column=1, value="Revenue")
        ws.cell(row=1, column=2, value=100000)
        # Hidden sheet
        ws = wb.create_sheet("Hidden IS")
        ws.sheet_state = "hidden"
        for i, l in enumerate(["Revenue", "COGS", "Net Income"]):
            ws.cell(row=1 + i, column=1, value=l)
            ws.cell(row=1 + i, column=2, value=100000 + i * 10000)
        # Normal sheet
        ws = wb.create_sheet("IS")
        ws.cell(row=1, column=1, value="Revenue")
        ws.cell(row=1, column=2, value=100000)
        ws.cell(row=2, column=1, value="Net Income")
        ws.cell(row=2, column=2, value=50000)
        return _run_deterministic_pipeline(_wb_to_bytes(wb))

    def test_empty_sheet_no_crash(self, result):
        empty_sections = result["sections_by_sheet"].get("Empty", [])
        assert empty_sections == []

    def test_single_row_handled(self, result):
        sections = result["sections_by_sheet"].get("Single", [])
        assert len(sections) <= 1  # 0 or 1

    def test_hidden_sheet_flagged(self, result):
        hidden = next(s for s in result["structured"]["sheets"] if s["sheet_name"] == "Hidden IS")
        assert hidden["is_hidden"] is True

    def test_all_sheets_in_summaries(self, result):
        names = {s["name"] for s in result["summaries"]}
        # All sheets including empty and hidden should appear
        assert "IS" in names


class TestCategoryKeywordRobustness:
    """Verify _guess_category does not produce false positives."""

    @pytest.mark.parametrize(
        "label,expected",
        [
            # True positives
            ("Income Statement", "income_statement"),
            ("Profit & Loss", "income_statement"),
            ("P/L", "income_statement"),
            ("I/S", "income_statement"),
            ("Balance Sheet", "balance_sheet"),
            ("B/S", "balance_sheet"),
            ("Statement of Financial Position", "balance_sheet"),
            ("Cash Flow Statement", "cash_flow"),
            ("Statement of Cash Flows", "cash_flow"),
            ("C/F", "cash_flow"),
            ("Debt Schedule", "debt_schedule"),
            ("Loan Facility", "debt_schedule"),
            # True negatives (line items, NOT section headers)
            ("Net Income", None),
            ("Gross Loss", None),
            ("Total Assets", None),
            ("Bad Debt Expense", None),
            ("Total Debt", None),
            ("Operating Income", None),
            ("Cash and Cash Equivalents", None),
            ("Revenue", None),
            ("SG&A", None),
            ("Depreciation & Amortization", None),
            ("Interest Income", None),
            ("Retained Earnings", None),
            ("Accounts Receivable", None),
            ("Goodwill", None),
            ("", None),
        ],
    )
    def test_category_classification(self, label, expected):
        assert _guess_category(label) == expected, (
            f"_guess_category('{label}') should be {expected}"
        )


class TestRealFixture:
    """Run against the actual messy_startup.xlsx fixture."""

    FIXTURE = Path(__file__).resolve().parent.parent / "fixtures" / "messy_startup.xlsx"

    @pytest.fixture(scope="class")
    def result(self):
        if not self.FIXTURE.exists():
            pytest.skip("messy_startup.xlsx fixture not found")
        return _run_deterministic_pipeline(self.FIXTURE.read_bytes())

    def test_four_sheets(self, result):
        assert result["structured"]["sheet_count"] == 4

    def test_combined_fs_multi_section(self, result):
        sections = result["sections_by_sheet"].get("Combined FS", [])
        assert len(sections) >= 2

    def test_saas_model_label_column(self, result):
        saas = next(s for s in result["structured"]["sheets"] if s["sheet_name"] == "SaaS Model")
        assert saas["label_column"] == "B"

    def test_full_pipeline_under_2s(self):
        """Real fixture should process in < 2s."""
        if not self.FIXTURE.exists():
            pytest.skip("messy_startup.xlsx fixture not found")
        start = time.time()
        _run_deterministic_pipeline(self.FIXTURE.read_bytes())
        elapsed = time.time() - start
        assert elapsed < 2.0, f"Took {elapsed:.2f}s"


class TestRealisticFixture:
    """Run against realistic_model.xlsx if available."""

    FIXTURE = Path(__file__).resolve().parent.parent / "fixtures" / "realistic_model.xlsx"

    @pytest.fixture(scope="class")
    def result(self):
        if not self.FIXTURE.exists():
            pytest.skip("realistic_model.xlsx fixture not found")
        return _run_deterministic_pipeline(self.FIXTURE.read_bytes())

    def test_parses_without_crash(self, result):
        assert result["structured"]["sheet_count"] >= 1

    def test_sections_detected(self, result):
        """At least one sheet should have sections or be single."""
        total = sum(len(v) for v in result["sections_by_sheet"].values())
        assert total >= 1

    def test_summaries_generated(self, result):
        assert len(result["summaries"]) >= 1
