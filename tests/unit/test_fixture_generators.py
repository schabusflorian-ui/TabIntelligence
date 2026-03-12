"""Tests for fixture generators and expected JSON files."""

import json
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
FIXTURES_DIR = PROJECT_ROOT / "tests" / "fixtures"
SCRIPTS_DIR = PROJECT_ROOT / "scripts"

# All generators and their expected output files
GENERATORS = [
    ("create_saas_startup.py", "saas_startup.xlsx"),
    ("create_seed_burn.py", "seed_burn.xlsx"),
    ("create_european_model.py", "european_model.xlsx"),
    ("create_edge_cases.py", "edge_cases.xlsx"),
    ("create_large_model.py", "large_model.xlsx"),
]

# Expected JSON files
EXPECTED_FILES = [
    "saas_startup_expected.json",
    "seed_burn_expected.json",
    "european_model_expected.json",
    "edge_cases_expected.json",
    "large_model_expected.json",
    "realistic_model_expected.json",
]

# Expected sheet counts per fixture
EXPECTED_SHEETS = {
    "saas_startup.xlsx": {
        "count": 5,
        "names": ["P&L", "Balance Sheet", "SaaS Metrics", "Headcount", "Scratch"],
    },
    "seed_burn.xlsx": {
        "count": 3,
        "names": ["Monthly P&L", "Cash", "Notes"],
    },
    "european_model.xlsx": {
        "count": 6,
        "names": [
            "Profit & Loss Account",
            "Statement of Financial Position",
            "Cash Flow Statement",
            "Debt Summary",
            "Assumptions",
            "Board Notes",
        ],
    },
    "edge_cases.xlsx": {
        "count": 4,
        "names": ["Financials", "Quarterly Detail", "Sensitivities", "Old Draft v1"],
    },
    "large_model.xlsx": {
        "count": 12,
        "names": [
            "Income Statement",
            "Balance Sheet",
            "Cash Flow Statement",
            "Debt Schedule",
            "Revenue Build",
            "OpEx Build",
            "Working Capital",
            "D&A Schedule",
            "Tax Schedule",
            "Assumptions",
            "Returns Analysis",
            "Cover Page",
        ],
    },
}


class TestGeneratorsProduceValidFiles:
    """Test that each generator produces a valid xlsx file."""

    @pytest.mark.parametrize("script,output_file", GENERATORS)
    def test_generator_produces_file(self, script, output_file):
        """Generator creates an xlsx file that exists and is non-empty."""
        xlsx_path = FIXTURES_DIR / output_file
        assert xlsx_path.exists(), f"{output_file} not found — run {script} first"
        assert xlsx_path.stat().st_size > 0, f"{output_file} is empty"

    @pytest.mark.parametrize("script,output_file", GENERATORS)
    def test_fixture_under_500kb(self, script, output_file):
        """Fixture file is under 500KB."""
        xlsx_path = FIXTURES_DIR / output_file
        if not xlsx_path.exists():
            pytest.skip(f"{output_file} not found")
        size_kb = xlsx_path.stat().st_size / 1024
        assert size_kb < 500, f"{output_file} is {size_kb:.0f}KB, exceeds 500KB limit"

    @pytest.mark.parametrize("output_file", EXPECTED_SHEETS.keys())
    def test_correct_sheet_count_and_names(self, output_file):
        """Fixture has the expected number of sheets with correct names."""
        xlsx_path = FIXTURES_DIR / output_file
        if not xlsx_path.exists():
            pytest.skip(f"{output_file} not found")

        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        expected = EXPECTED_SHEETS[output_file]

        assert len(wb.sheetnames) == expected["count"], (
            f"{output_file}: expected {expected['count']}"
            f" sheets, got {len(wb.sheetnames)}:"
            f" {wb.sheetnames}"
        )

        for name in expected["names"]:
            assert name in wb.sheetnames, (
                f"{output_file}: missing sheet '{name}', have: {wb.sheetnames}"
            )
        wb.close()


class TestExpectedJsonStructure:
    """Test that expected JSON files have valid structure."""

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_required_keys(self, expected_file):
        """Expected JSON has all required top-level keys."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        required = [
            "description",
            "model_info",
            "expected_triage",
            "expected_mappings",
            "acceptable_alternatives",
        ]
        for key in required:
            assert key in data, f"{expected_file} missing key: {key}"

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_model_info_keys(self, expected_file):
        """model_info has required keys."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        info = data["model_info"]
        for key in ["file", "sheets", "company_type", "periods"]:
            assert key in info, f"{expected_file} model_info missing: {key}"

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_triage_tiers_valid(self, expected_file):
        """All triage entries have valid tiers (1-4) and decisions."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        valid_tiers = {1, 2, 3, 4}
        valid_decisions = {"PROCESS_HIGH", "PROCESS_MEDIUM", "PROCESS_LOW", "SKIP"}

        for entry in data["expected_triage"]:
            assert "sheet_name" in entry, f"{expected_file}: triage entry missing sheet_name"
            assert "tier" in entry, f"{expected_file}: triage entry missing tier"
            assert "decision" in entry, f"{expected_file}: triage entry missing decision"
            assert entry["tier"] in valid_tiers, (
                f"{expected_file}: invalid tier {entry['tier']} for {entry['sheet_name']}"
            )
            assert entry["decision"] in valid_decisions, (
                f"{expected_file}: invalid decision {entry['decision']} for {entry['sheet_name']}"
            )

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_no_duplicate_labels_per_sheet(self, expected_file):
        """No duplicate original_labels within the same sheet."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        # Group labels by sheet
        by_sheet = {}
        for m in data["expected_mappings"]:
            sheet = m.get("sheet", "Unknown")
            by_sheet.setdefault(sheet, []).append(m["original_label"])

        for sheet, labels in by_sheet.items():
            dupes = [l for l in labels if labels.count(l) > 1]
            assert not dupes, f"{expected_file}: duplicate labels in '{sheet}': {set(dupes)}"

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_mappings_have_required_fields(self, expected_file):
        """Each mapping has original_label, canonical_name, and sheet."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        for m in data["expected_mappings"]:
            assert "original_label" in m, f"{expected_file}: mapping missing original_label"
            assert "canonical_name" in m, f"{expected_file}: mapping missing canonical_name"
            assert "sheet" in m, (
                f"{expected_file}: mapping '{m.get('original_label')}' missing sheet"
            )


class TestCanonicalNamesValid:
    """Validate that all canonical_names in expected JSONs exist in taxonomy."""

    @pytest.fixture(scope="class")
    def all_canonical_names(self):
        """Load all valid canonical names from taxonomy."""
        taxonomy_path = PROJECT_ROOT / "data" / "taxonomy.json"
        with open(taxonomy_path) as f:
            taxonomy = json.load(f)

        names = set()
        for items in taxonomy.get("categories", {}).values():
            for item in items:
                names.add(item["canonical_name"])
        return names

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_canonical_names_in_taxonomy(self, expected_file, all_canonical_names):
        """All canonical_names used in expected JSON exist in taxonomy."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        missing = []
        for m in data["expected_mappings"]:
            cn = m["canonical_name"]
            if cn not in all_canonical_names:
                missing.append(f"{m['original_label']} -> {cn}")

        assert not missing, f"{expected_file}: canonical names not in taxonomy:\n" + "\n".join(
            f"  - {m}" for m in missing
        )

    @pytest.mark.parametrize("expected_file", EXPECTED_FILES)
    def test_acceptable_alternatives_in_taxonomy(self, expected_file, all_canonical_names):
        """All acceptable alternative values exist in taxonomy."""
        path = FIXTURES_DIR / expected_file
        if not path.exists():
            pytest.skip(f"{expected_file} not found")

        with open(path) as f:
            data = json.load(f)

        missing = []
        for canonical, alts in data.get("acceptable_alternatives", {}).items():
            for alt in alts:
                if alt not in all_canonical_names:
                    missing.append(f"{canonical}: alt '{alt}' not in taxonomy")

        assert not missing, f"{expected_file}: invalid alternatives:\n" + "\n".join(
            f"  - {m}" for m in missing
        )
