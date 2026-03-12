#!/usr/bin/env python3
"""Create tests/fixtures/saas_startup.xlsx — Series B SaaS company model.

Sheet 1 (P&L): Income statement with subscription + services revenue split
Sheet 2 (Balance Sheet): Standard BS with current/non-current breakdown
Sheet 3 (SaaS Metrics): ARR, MRR, NRR, churn, CAC, LTV, customers
Sheet 4 (Headcount): Team size by department with rev-per-employee
Sheet 5 (Scratch): Pure text notes (should be tier 4 skip)
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

OUTPUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "saas_startup.xlsx"


def _build_sheet1_pl(wb: openpyxl.Workbook) -> None:
    """Sheet 1: P&L — labels in col A, spacer col B, data in C-E."""
    ws = wb.create_sheet("P&L")
    bold = Font(bold=True)

    # Row 1: company name merged A1:E1
    ws.merge_cells("A1:E1")
    ws["A1"] = "AcmeSoft Inc - P&L"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2: unit annotation
    ws["A2"] = "(in thousands)"
    ws["A2"].font = Font(italic=True, color="888888")

    # Row 3: headers
    ws["A3"] = ""
    ws["B3"] = ""
    ws["C3"] = "FY2023"
    ws["D3"] = "FY2024"
    ws["E3"] = "FY2025E"
    for col in ["C", "D", "E"]:
        ws[f"{col}3"].font = bold

    # Data rows: (row, label, bold_flag, c_val, d_val, e_val)
    # Col A = label, Col B = spacer (empty), Col C/D/E = data
    data = [
        (4, "Rev", True, 4200, 5880, 7644),
        (5, "  Subscription Revenue", False, 3570, 5058, 6879),
        (6, "  Services Revenue", False, 630, 822, 765),
        (7, None, False, None, None, None),  # spacer between rev detail and COGS
    ]

    # COGS at row 8
    cogs_row = 8
    data.append((cogs_row, "COGS", False, -1260, -1646, -1911))

    # GP at row 9 — will add formula below
    gp_row = 9
    data.append((gp_row, "GP", True, 2940, 4234, 5733))

    data.append((10, None, False, None, None, None))  # blank row

    # Operating Expenses section
    data.append((11, "Operating Expenses", True, None, None, None))  # section header
    data.append((12, "  S&M", False, -1680, -2117, -2446))
    data.append((13, "  R&D", False, -1050, -1411, -1681))
    data.append((14, "  G&A", False, -420, -529, -612))
    data.append((15, "Total Opex", True, -3150, -4057, -4739))

    data.append((16, None, False, None, None, None))  # blank row

    data.append((17, "EBITDA", True, -210, 177, 994))
    data.append((18, "D&A", False, -168, -235, -306))
    data.append((19, "EBIT", True, -378, -58, 688))

    for row_num, label, is_bold, *vals in data:
        if label is not None:
            cell = ws.cell(row=row_num, column=1, value=label)
            if is_bold:
                cell.font = bold
        for i, v in enumerate(vals):
            if v is not None:
                c = ws.cell(row=row_num, column=3 + i, value=v)
                c.number_format = "#,##0"

    # GP formulas: GP = Rev + COGS
    ws.cell(row=gp_row, column=3, value="=C4+C8").number_format = "#,##0"
    ws.cell(row=gp_row, column=4, value="=D4+D8").number_format = "#,##0"
    ws.cell(row=gp_row, column=5, value="=E4+E8").number_format = "#,##0"


def _build_sheet2_balance_sheet(wb: openpyxl.Workbook) -> None:
    """Sheet 2: Balance Sheet — labels in col A, data in B-D."""
    ws = wb.create_sheet("Balance Sheet")
    bold = Font(bold=True)

    # Row 1: title
    ws["A1"] = "Balance Sheet"
    ws["A1"].font = Font(bold=True, size=12)

    # Row 2: headers
    ws["A2"] = ""
    ws["B2"] = "FY2023"
    ws["C2"] = "FY2024"
    ws["D2"] = "FY2025E"
    for col in ["B", "C", "D"]:
        ws[f"{col}2"].font = bold

    data = [
        (3, "Cash & Equivalents", True, 8500, 6200, 5800),
        (4, "Accounts Receivable", False, 1050, 1470, 1911),
        (5, "Prepaid Expenses", False, 210, 294, 382),
        (6, "Total Current Assets", True, 9760, 7964, 8093),
        (7, None, False, None, None, None),  # blank
        (8, "PP&E, Net", False, 840, 1176, 1529),
        (9, "Total Assets", True, 10600, 9140, 9622),
        (10, None, False, None, None, None),  # blank
        (11, "Accounts Payable", False, 315, 412, 478),
        (12, "Deferred Revenue", False, 700, 980, 1274),
        (13, "Total Current Liabilities", True, 1015, 1392, 1752),
        (14, "Convertible Notes", False, 3000, 3000, 0),
        (15, "Total Liabilities", True, 4015, 4392, 1752),
        (16, None, False, None, None, None),  # blank
        (17, "Total Equity", True, 6585, 4748, 7870),
        (18, "Total Liabilities & Equity", True, 10600, 9140, 9622),
    ]

    for row_num, label, is_bold, *vals in data:
        if label is not None:
            cell = ws.cell(row=row_num, column=1, value=label)
            if is_bold:
                cell.font = bold
        for i, v in enumerate(vals):
            if v is not None:
                ws.cell(row=row_num, column=2 + i, value=v).number_format = "#,##0"


def _build_sheet3_metrics(wb: openpyxl.Workbook) -> None:
    """Sheet 3: SaaS Metrics — ARR, MRR, NRR, churn, CAC, LTV, customers."""
    ws = wb.create_sheet("SaaS Metrics")
    bold = Font(bold=True)

    # Row 1: title
    ws["A1"] = "Key Performance Metrics"
    ws["A1"].font = Font(bold=True, size=12)

    # Row 2: headers
    ws["A2"] = ""
    ws["B2"] = "FY2023"
    ws["C2"] = "FY2024"
    ws["D2"] = "FY2025E"
    for col in ["B", "C", "D"]:
        ws[f"{col}2"].font = bold

    # Metrics data: (row, label, val1, val2, val3, format)
    metrics = [
        (3, "ARR ($000)", 4284, 6070, 8255, "#,##0"),
        (4, "MRR", 357, 506, 688, "#,##0"),
        (5, "NRR", 1.12, 1.15, 1.18, "0%"),
        (6, "Logo Churn", 0.08, 0.065, 0.05, "0%"),
        (7, "CAC", 850, 780, 720, "#,##0"),
        (8, "LTV", 3400, 4100, 5040, "#,##0"),
        (9, "LTV/CAC", 4.0, 5.3, 7.0, "0.0"),
        (10, "Burn Rate", 210, 0, 0, "#,##0"),
        (11, "Runway (months)", 41, 0, 0, "#,##0"),
        (12, "Customers", 420, 580, 750, "#,##0"),
        (13, "New Customers", 180, 220, 250, "#,##0"),
    ]

    for row_num, label, v1, v2, v3, fmt in metrics:
        ws.cell(row=row_num, column=1, value=label)
        for i, v in enumerate([v1, v2, v3]):
            c = ws.cell(row=row_num, column=2 + i, value=v)
            c.number_format = fmt

    # Annotation rows
    # Row 14: blank
    # Row 15: note
    ws.cell(row=15, column=1, value="Note: NRR based on dollar retention")
    # Row 16: separator
    ws.cell(row=16, column=1, value="===")
    # Row 17: source
    ws.cell(row=17, column=1, value="Source: Internal metrics dashboard")


def _build_sheet4_headcount(wb: openpyxl.Workbook) -> None:
    """Sheet 4: Headcount by department with rev per employee."""
    ws = wb.create_sheet("Headcount")
    bold = Font(bold=True)

    # Row 1: title
    ws["A1"] = "Team Size"
    ws["A1"].font = Font(bold=True, size=12)

    # Row 2: headers
    ws["A2"] = ""
    ws["B2"] = "FY2023"
    ws["C2"] = "FY2024"
    ws["D2"] = "FY2025E"
    for col in ["B", "C", "D"]:
        ws[f"{col}2"].font = bold

    data = [
        (3, "Engineering", False, 35, 48, 60, "#,##0"),
        (4, "Sales", False, 22, 30, 38, "#,##0"),
        (5, "G&A", False, 12, 15, 18, "#,##0"),
        (6, "Total HC", True, 69, 93, 116, "#,##0"),
        (7, "Rev per Employee", False, 60.9, 63.2, 65.9, "0.0"),
    ]

    for row_num, label, is_bold, v1, v2, v3, fmt in data:
        cell = ws.cell(row=row_num, column=1, value=label)
        if is_bold:
            cell.font = bold
        for i, v in enumerate([v1, v2, v3]):
            c = ws.cell(row=row_num, column=2 + i, value=v)
            c.number_format = fmt


def _build_sheet5_scratch(wb: openpyxl.Workbook) -> None:
    """Sheet 5: Scratch notes — pure text, should be tier 4 skip."""
    ws = wb.create_sheet("Scratch")

    notes = [
        "TODO: Update revenue model for Q3",
        "TODO: Validate CAC assumptions",
        "v2.3 - Updated March 2025",
        "Draft - do not distribute",
    ]

    for i, text in enumerate(notes, start=1):
        ws.cell(row=i, column=1, value=text)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_sheet1_pl(wb)
    _build_sheet2_balance_sheet(wb)
    _build_sheet3_metrics(wb)
    _build_sheet4_headcount(wb)
    _build_sheet5_scratch(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"Created {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
