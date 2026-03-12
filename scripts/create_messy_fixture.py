"""Generate tests/fixtures/messy_startup.xlsx — a realistic messy startup model.

Sheet 1 (SaaS Model): Labels in column B, periods in row 3, unit annotation
Sheet 2 (Combined FS): P&L rows 1-25, blank 26-28, BS rows 29-50 on ONE sheet
Sheet 3 (Metrics & Notes): SaaS metrics mixed with annotations and separators
Sheet 4 (Notes): Pure text notes (should be tier 4 skip)
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

OUTPUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "messy_startup.xlsx"


def _build_sheet1_saas_model(wb: openpyxl.Workbook) -> None:
    """Sheet 1: Labels in column B, periods in row 3, '(in thousands)' annotation."""
    ws = wb.create_sheet("SaaS Model")
    bold = Font(bold=True)

    # Row 1: company name (merged across B-F)
    ws.merge_cells("B1:F1")
    ws["B1"] = "Acme SaaS Inc."
    ws["B1"].font = Font(bold=True, size=14)

    # Row 2: unit annotation
    ws["B2"] = "(in thousands)"
    ws["B2"].font = Font(italic=True, color="888888")

    # Row 3: period headers — column A is numbering, B is labels, C-F are periods
    ws["A3"] = "#"
    ws["B3"] = ""
    ws["C3"] = "FY2022"
    ws["D3"] = "FY2023"
    ws["E3"] = "FY2024E"
    ws["F3"] = "FY2025E"
    for col in ["C", "D", "E", "F"]:
        ws[f"{col}3"].font = bold

    # Data rows (labels in col B, row numbers in col A)
    data = [
        (4, 1, "Revenue", 1200, 1800, 2500, 3200),
        (5, 2, "Growth Rate", None, 0.50, 0.389, 0.28),
        (6, 3, "COGS", -360, -504, -625, -768),
        (7, 4, "Gross Profit", 840, 1296, 1875, 2432),
        (8, None, None, None, None, None, None),  # blank separator
        (9, 5, "Sales & Marketing", -480, -630, -750, -864),
        (10, 6, "R&D", -240, -324, -400, -480),
        (11, 7, "G&A", -120, -162, -200, -240),
        (12, 8, "Total OpEx", -840, -1116, -1350, -1584),
        (13, None, None, None, None, None, None),  # blank separator
        (14, 9, "EBITDA", 0, 180, 525, 848),
        (15, 10, "D&A", -60, -72, -100, -128),
        (16, 11, "EBIT", -60, 108, 425, 720),
        (17, None, "Source: Management projections", None, None, None, None),
    ]

    for row_num, num, label, *vals in data:
        if num is not None:
            ws.cell(row=row_num, column=1, value=num)
        if label is not None:
            cell = ws.cell(row=row_num, column=2, value=label)
            if label in ("Revenue", "Gross Profit", "Total OpEx", "EBITDA", "EBIT"):
                cell.font = bold
        for i, v in enumerate(vals):
            if v is not None:
                c = ws.cell(row=row_num, column=3 + i, value=v)
                if isinstance(v, float) and abs(v) < 1:
                    c.number_format = "0.0%"
                else:
                    c.number_format = "#,##0"

    # Formulas
    ws.cell(row=7, column=3, value="=C4+C6").number_format = "#,##0"
    ws.cell(row=12, column=3, value="=SUM(C9:C11)").number_format = "#,##0"
    ws.cell(row=14, column=3, value="=C7+C12").number_format = "#,##0"


def _build_sheet2_combined_fs(wb: openpyxl.Workbook) -> None:
    """Sheet 2: P&L (rows 1-25) + Balance Sheet (rows 29-50) on one sheet."""
    ws = wb.create_sheet("Combined FS")
    bold = Font(bold=True)
    header_font = Font(bold=True, size=12)

    # ---- P&L Section (rows 1-25) ----
    ws["A1"] = "Profit & Loss Statement"
    ws["A1"].font = header_font

    ws["A2"] = ""
    ws["B2"] = "FY2022"
    ws["C2"] = "FY2023"
    ws["D2"] = "FY2024E"
    for col in ["B", "C", "D"]:
        ws[f"{col}2"].font = bold

    pl_data = [
        (3, "Revenue", 5000, 6200, 7800),
        (4, "  Product Revenue", 3500, 4340, 5460),
        (5, "  Service Revenue", 1500, 1860, 2340),
        (6, "Cost of Revenue", -2000, -2418, -2964),
        (7, "Gross Profit", 3000, 3782, 4836),
        (8, None, None, None, None),  # blank
        (9, "Operating Expenses", None, None, None),
        (10, "  Sales & Marketing", -1500, -1798, -2106),
        (11, "  Research & Development", -750, -930, -1170),
        (12, "  General & Administrative", -450, -558, -702),
        (13, "Total Operating Expenses", -2700, -3286, -3978),
        (14, None, None, None, None),  # blank
        (15, "Operating Income", 300, 496, 858),
        (16, "Interest Expense", -50, -45, -35),
        (17, "Other Income", 10, 15, 20),
        (18, "Income Before Tax", 260, 466, 843),
        (19, "Income Tax", -65, -117, -211),
        (20, "Net Income", 195, 350, 632),
        (21, None, None, None, None),  # blank
        (22, "Depreciation & Amortization", 200, 248, 312),
        (23, "EBITDA", 500, 744, 1170),
        (24, None, None, None, None),  # blank
        (25, "---", None, None, None),  # separator
    ]

    for row_num, label, *vals in pl_data:
        if label is not None:
            cell = ws.cell(row=row_num, column=1, value=label)
            if label in (
                "Revenue",
                "Gross Profit",
                "Operating Expenses",
                "Total Operating Expenses",
                "Operating Income",
                "Net Income",
                "EBITDA",
            ):
                cell.font = bold
        for i, v in enumerate(vals):
            if v is not None:
                ws.cell(row=row_num, column=2 + i, value=v).number_format = "#,##0"

    # Formulas for P&L
    ws.cell(row=7, column=2, value="=B3+B6").number_format = "#,##0"
    ws.cell(row=13, column=2, value="=SUM(B10:B12)").number_format = "#,##0"
    ws.cell(row=15, column=2, value="=B7+B13").number_format = "#,##0"
    ws.cell(row=20, column=2, value="=B18+B19").number_format = "#,##0"

    # ---- Blank gap rows 26-28 ----
    # (left empty — 3 blank rows = section boundary)

    # ---- Balance Sheet Section (rows 29-50) ----
    ws["A29"] = "Balance Sheet"
    ws["A29"].font = header_font

    ws["A30"] = ""
    ws["B30"] = "FY2022"
    ws["C30"] = "FY2023"
    ws["D30"] = "FY2024E"
    for col in ["B", "C", "D"]:
        ws[f"{col}30"].font = bold

    bs_data = [
        (31, "Assets", None, None, None),
        (32, "  Cash & Equivalents", 2000, 2800, 3500),
        (33, "  Accounts Receivable", 800, 992, 1248),
        (34, "  Inventory", 200, 248, 312),
        (35, "  Prepaid Expenses", 100, 124, 156),
        (36, "Total Current Assets", 3100, 4164, 5216),
        (37, None, None, None, None),  # blank
        (38, "  Property & Equipment", 1500, 1800, 2200),
        (39, "  Intangible Assets", 500, 450, 400),
        (40, "  Goodwill", 1000, 1000, 1000),
        (41, "Total Non-Current Assets", 3000, 3250, 3600),
        (42, "Total Assets", 6100, 7414, 8816),
        (43, None, None, None, None),  # blank
        (44, "Liabilities", None, None, None),
        (45, "  Accounts Payable", 400, 484, 593),
        (46, "  Accrued Expenses", 300, 372, 468),
        (47, "Total Current Liabilities", 700, 856, 1061),
        (48, "  Long-term Debt", 1000, 800, 600),
        (49, "Total Liabilities", 1700, 1656, 1661),
        (50, "Total Equity", 4400, 5758, 7155),
    ]

    for row_num, label, *vals in bs_data:
        if label is not None:
            cell = ws.cell(row=row_num, column=1, value=label)
            if label in (
                "Assets",
                "Total Current Assets",
                "Total Non-Current Assets",
                "Total Assets",
                "Liabilities",
                "Total Current Liabilities",
                "Total Liabilities",
                "Total Equity",
            ):
                cell.font = bold
        for i, v in enumerate(vals):
            if v is not None:
                ws.cell(row=row_num, column=2 + i, value=v).number_format = "#,##0"

    # Formulas for BS
    ws.cell(row=36, column=2, value="=SUM(B32:B35)").number_format = "#,##0"
    ws.cell(row=42, column=2, value="=B36+B41").number_format = "#,##0"


def _build_sheet3_metrics(wb: openpyxl.Workbook) -> None:
    """Sheet 3: SaaS metrics mixed with annotations."""
    ws = wb.create_sheet("Metrics & Notes")
    Font(bold=True)

    ws["A1"] = "Key SaaS Metrics"
    ws["A1"].font = Font(bold=True, size=12)

    ws["B1"] = "FY2022"
    ws["C1"] = "FY2023"
    ws["D1"] = "FY2024E"

    metrics_data = [
        (2, "ARR ($000)", 1200, 1800, 2500),
        (3, "MRR ($000)", 100, 150, 208),
        (4, "Net Revenue Retention", 1.15, 1.18, 1.20),
        (5, "Gross Margin", 0.70, 0.72, 0.75),
        (6, "CAC Payback (months)", 18, 15, 12),
        (7, "LTV/CAC", 3.2, 4.1, 5.0),
        (8, None, None, None, None),
        (9, "Note: ARR based on annualised December MRR", None, None, None),
        (10, "Note: NRR excludes churned customers from base", None, None, None),
        (11, "===", None, None, None),
        (12, "Prepared by: Finance Team", None, None, None),
    ]

    for row_num, label, *vals in metrics_data:
        if label is not None:
            ws.cell(row=row_num, column=1, value=label)
        for i, v in enumerate(vals):
            if v is not None:
                c = ws.cell(row=row_num, column=2 + i, value=v)
                if isinstance(v, float) and v < 2:
                    c.number_format = "0.0%"
                elif isinstance(v, float):
                    c.number_format = "0.0"
                else:
                    c.number_format = "#,##0"


def _build_sheet4_notes(wb: openpyxl.Workbook) -> None:
    """Sheet 4: Pure text notes — should be tier 4."""
    ws = wb.create_sheet("Notes")

    notes = [
        "Model Assumptions and Notes",
        "",
        "1. Revenue growth based on pipeline analysis and Q3 run rate",
        "2. COGS assumes 30% margin on product, 20% margin on services",
        "3. OpEx scaling assumes hiring plan attached in HR tab",
        "4. Tax rate: 25% effective, no NOL carryforwards",
        "5. Capex: $200K base + $50K per 100 new customers",
        "",
        "Revision History:",
        "v1.0 - Initial draft (Jan 2024)",
        "v1.1 - Updated revenue assumptions (Feb 2024)",
        "v2.0 - Board presentation version (Mar 2024)",
    ]

    for i, text in enumerate(notes, start=1):
        ws.cell(row=i, column=1, value=text)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_sheet1_saas_model(wb)
    _build_sheet2_combined_fs(wb)
    _build_sheet3_metrics(wb)
    _build_sheet4_notes(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"Created {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
