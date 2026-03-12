#!/usr/bin/env python3
"""Create tests/fixtures/european_model.xlsx — European mid-market IFRS model."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

OUTPUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "european_model.xlsx"

# =============================================================================
# Constants
# =============================================================================

PERIODS = ["FY2022A", "FY2023A", "FY2024E", "FY2025E"]

BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
TITLE_FONT_14 = Font(name="Calibri", size=14, bold=True)
TITLE_FONT_12 = Font(name="Calibri", size=12, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="808080")

NUMBER_FMT = "#,##0.0"
PCT_FMT = "0.0%"


# =============================================================================
# Helpers
# =============================================================================


def _write_title(ws, title, size=14):
    """Write bold title in row 1 and italic gray subtitle in row 2."""
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT_14 if size == 14 else TITLE_FONT_12
    cell.alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:E2")
    sub = ws.cell(row=2, column=1, value="(EUR millions)")
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(horizontal="left")


def _write_headers(ws, row=3):
    """Write period headers in the given row."""
    ws.cell(row=row, column=1, value="").font = BOLD_FONT
    for col_idx, period in enumerate(PERIODS, 2):
        cell = ws.cell(row=row, column=col_idx, value=period)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 42
    for col_letter in ("B", "C", "D", "E"):
        ws.column_dimensions[col_letter].width = 14


def _write_row(ws, row, label, values, bold=False, fmt=NUMBER_FMT):
    """Write a label + numeric values row."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = BOLD_FONT if bold else NORMAL_FONT

    if values is None:
        return

    for col_idx, val in enumerate(values, 2):
        if val is None:
            continue
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
        if bold:
            cell.font = BOLD_FONT


def _write_section_header(ws, row, label):
    """Write a bold section header with no values."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = BOLD_FONT


# =============================================================================
# Sheet builders
# =============================================================================


def _build_profit_and_loss(wb):
    """Sheet 1: Profit & Loss Account (income_statement)."""
    ws = wb.active
    ws.title = "Profit & Loss Account"

    _write_title(ws, "Consolidated Profit & Loss Account", size=14)
    _write_headers(ws, row=3)

    r = 4  # starting data row

    # Line items
    turnover = [148.5, 159.2, 170.8, 182.4]
    cost_of_sales = [-87.8, -93.4, -99.6, -105.5]
    gross_profit = [round(turnover[i] + cost_of_sales[i], 1) for i in range(4)]
    # gross_profit => [60.7, 65.8, 71.2, 76.9]

    distribution_costs = [-18.6, -19.9, -21.0, -22.0]
    admin_expenses = [-22.3, -23.5, -24.8, -26.1]
    other_operating_income = [1.2, 1.5, 1.8, 2.0]
    operating_profit = [
        round(
            gross_profit[i] + distribution_costs[i] + admin_expenses[i] + other_operating_income[i],
            1,
        )
        for i in range(4)
    ]
    # operating_profit => [21.0, 23.9, 27.2, 30.8]

    finance_costs = [-4.8, -4.5, -4.1, -3.7]
    finance_income = [0.3, 0.4, 0.5, 0.6]
    profit_before_tax = [
        round(operating_profit[i] + finance_costs[i] + finance_income[i], 1) for i in range(4)
    ]
    # profit_before_tax => [16.5, 19.8, 23.6, 27.7]

    taxation = [-4.1, -5.0, -5.9, -6.9]
    profit_for_year = [round(profit_before_tax[i] + taxation[i], 1) for i in range(4)]
    # profit_for_year => [12.4, 14.8, 17.7, 20.8]

    _write_row(ws, r, "Turnover", turnover, bold=True)
    r += 1
    _write_row(ws, r, "Cost of Sales", cost_of_sales)
    r += 1
    _write_row(ws, r, "Gross Profit", gross_profit, bold=True)
    r += 1
    r += 1  # blank row
    _write_row(ws, r, "Distribution Costs", distribution_costs)
    r += 1
    _write_row(ws, r, "Administrative Expenses", admin_expenses)
    r += 1
    _write_row(ws, r, "Other Operating Income", other_operating_income)
    r += 1
    _write_row(ws, r, "Operating Profit", operating_profit, bold=True)
    r += 1
    r += 1  # blank row
    _write_row(ws, r, "Finance Costs", finance_costs)
    r += 1
    _write_row(ws, r, "Finance Income", finance_income)
    r += 1
    _write_row(ws, r, "Profit Before Tax", profit_before_tax, bold=True)
    r += 1
    _write_row(ws, r, "Taxation", taxation)
    r += 1
    _write_row(ws, r, "Profit for the Year", profit_for_year, bold=True)
    r += 1


def _build_balance_sheet(wb):
    """Sheet 2: Statement of Financial Position (balance_sheet)."""
    ws = wb.create_sheet("Statement of Financial Position")

    _write_title(ws, "Consolidated Statement of Financial Position", size=14)
    _write_headers(ws, row=3)

    r = 4

    # --- Non-Current Assets ---
    _write_section_header(ws, r, "Non-Current Assets")
    r += 1
    ppe = [45.2, 46.8, 48.5, 50.1]
    intangibles = [12.5, 11.8, 11.1, 10.4]
    goodwill = [28.0, 28.0, 28.0, 28.0]
    total_nca = [round(ppe[i] + intangibles[i] + goodwill[i], 1) for i in range(4)]
    # total_nca => [85.7, 86.6, 87.6, 88.5]

    _write_row(ws, r, "  Property, Plant and Equipment", ppe)
    r += 1
    _write_row(ws, r, "  Intangible Assets", intangibles)
    r += 1
    _write_row(ws, r, "  Goodwill", goodwill)
    r += 1
    _write_row(ws, r, "Total Non-Current Assets", total_nca, bold=True)
    r += 1
    r += 1  # blank

    # --- Current Assets ---
    _write_section_header(ws, r, "Current Assets")
    r += 1
    stocks = [14.8, 15.5, 16.2, 17.0]
    trade_debtors = [24.8, 26.5, 28.5, 30.4]
    other_debtors = [3.2, 3.4, 3.6, 3.8]
    cash_at_bank = [18.5, 22.1, 28.0, 36.8]
    total_ca = [
        round(stocks[i] + trade_debtors[i] + other_debtors[i] + cash_at_bank[i], 1)
        for i in range(4)
    ]
    # total_ca => [61.3, 67.5, 76.3, 88.0]

    _write_row(ws, r, "  Stocks", stocks)
    r += 1
    _write_row(ws, r, "  Trade Debtors", trade_debtors)
    r += 1
    _write_row(ws, r, "  Other Debtors", other_debtors)
    r += 1
    _write_row(ws, r, "  Cash at Bank", cash_at_bank)
    r += 1
    _write_row(ws, r, "Total Current Assets", total_ca, bold=True)
    r += 1
    r += 1  # blank

    total_assets = [round(total_nca[i] + total_ca[i], 1) for i in range(4)]
    # total_assets => [147.0, 154.1, 163.9, 176.5]
    _write_row(ws, r, "Total Assets", total_assets, bold=True)
    r += 1
    r += 1  # blank

    # --- Current Liabilities ---
    _write_section_header(ws, r, "Current Liabilities")
    r += 1
    trade_creditors = [16.2, 17.0, 18.0, 19.1]
    accruals = [8.5, 9.1, 9.7, 10.4]
    provisions = [3.5, 3.8, 4.0, 4.2]
    current_borrowings = [5.0, 5.0, 5.0, 5.0]
    total_cl = [
        round(trade_creditors[i] + accruals[i] + provisions[i] + current_borrowings[i], 1)
        for i in range(4)
    ]
    # total_cl => [33.2, 34.9, 36.7, 38.7]

    _write_row(ws, r, "  Trade Creditors", trade_creditors)
    r += 1
    _write_row(ws, r, "  Accruals", accruals)
    r += 1
    _write_row(ws, r, "  Provisions", provisions)
    r += 1
    _write_row(ws, r, "  Current Portion of Borrowings", current_borrowings)
    r += 1
    _write_row(ws, r, "Total Current Liabilities", total_cl, bold=True)
    r += 1
    r += 1  # blank

    # --- Non-Current Liabilities ---
    _write_section_header(ws, r, "Non-Current Liabilities")
    r += 1
    bank_borrowings = [40.0, 35.0, 30.0, 25.0]
    other_ncl = [2.5, 2.3, 2.1, 1.9]
    total_ncl = [round(bank_borrowings[i] + other_ncl[i], 1) for i in range(4)]
    # total_ncl => [42.5, 37.3, 32.1, 26.9]

    _write_row(ws, r, "  Bank Borrowings", bank_borrowings)
    r += 1
    _write_row(ws, r, "  Other Non-Current Liabilities", other_ncl)
    r += 1
    _write_row(ws, r, "Total Non-Current Liabilities", total_ncl, bold=True)
    r += 1
    r += 1  # blank

    total_liabilities = [round(total_cl[i] + total_ncl[i], 1) for i in range(4)]
    # total_liabilities => [75.7, 72.2, 68.8, 65.6]
    _write_row(ws, r, "Total Liabilities", total_liabilities, bold=True)
    r += 1
    r += 1  # blank

    # --- Shareholders' Funds ---
    _write_section_header(ws, r, "Shareholders' Funds")
    r += 1
    share_capital = [10.0, 10.0, 10.0, 10.0]
    retained_profit = [61.3, 71.9, 85.1, 100.9]
    total_equity = [round(share_capital[i] + retained_profit[i], 1) for i in range(4)]
    # total_equity => [71.3, 81.9, 95.1, 110.9]

    _write_row(ws, r, "  Share Capital", share_capital)
    r += 1
    _write_row(ws, r, "  Retained Profit", retained_profit)
    r += 1
    _write_row(ws, r, "Total Shareholders' Funds", total_equity, bold=True)
    r += 1
    r += 1  # blank

    total_le = [round(total_liabilities[i] + total_equity[i], 1) for i in range(4)]
    # total_le => [147.0, 154.1, 163.9, 176.5]
    _write_row(ws, r, "Total Liabilities and Shareholders' Funds", total_le, bold=True)


def _build_cash_flow(wb):
    """Sheet 3: Cash Flow Statement (cash_flow)."""
    ws = wb.create_sheet("Cash Flow Statement")

    _write_title(ws, "Consolidated Cash Flow Statement", size=14)
    _write_headers(ws, row=3)

    r = 4

    cash_gen_ops = [18.0, 20.7, 24.1, 27.8]
    taxation_paid = [-3.8, -4.1, -5.1, -5.5]
    net_operating = [round(cash_gen_ops[i] + taxation_paid[i], 1) for i in range(4)]
    # net_operating => [14.2, 16.6, 19.0, 22.3]

    _write_row(ws, r, "Cash Generated from Operations", cash_gen_ops, bold=True)
    r += 1
    _write_row(ws, r, "Taxation Paid", taxation_paid)
    r += 1
    _write_row(ws, r, "Net Cash from Operating Activities", net_operating, bold=True)
    r += 1
    r += 1  # blank

    purchase_ppe = [-6.0, -6.8, -6.8, -7.5]
    purchase_intangibles = [-1.0, -1.0, -1.0, -1.0]
    net_investing = [round(purchase_ppe[i] + purchase_intangibles[i], 1) for i in range(4)]
    # net_investing => [-7.0, -7.8, -7.8, -8.5]

    _write_row(ws, r, "Purchase of Property, Plant and Equipment", purchase_ppe)
    r += 1
    _write_row(ws, r, "Purchase of Intangible Assets", purchase_intangibles)
    r += 1
    _write_row(ws, r, "Net Cash Used in Investing Activities", net_investing, bold=True)
    r += 1
    r += 1  # blank

    repayment = [-5.0, -5.0, -5.0, -5.0]
    dividends_paid = [0.0, -0.2, -0.3, 0.0]
    net_financing = [round(repayment[i] + dividends_paid[i], 1) for i in range(4)]
    # net_financing => [-5.0, -5.2, -5.3, -5.0]

    _write_row(ws, r, "Repayment of Borrowings", repayment)
    r += 1
    _write_row(ws, r, "Dividends Paid", dividends_paid)
    r += 1
    _write_row(ws, r, "Net Cash Used in Financing Activities", net_financing, bold=True)
    r += 1
    r += 1  # blank

    net_increase = [
        round(net_operating[i] + net_investing[i] + net_financing[i], 1) for i in range(4)
    ]
    # net_increase => [2.2, 3.6, 5.9, 8.8]
    beg_cash = [16.3, 18.5, 22.1, 28.0]
    end_cash = [round(beg_cash[i] + net_increase[i], 1) for i in range(4)]
    # end_cash => [18.5, 22.1, 28.0, 36.8]

    _write_row(ws, r, "Net Increase / (Decrease) in Cash", net_increase, bold=True)
    r += 1
    _write_row(ws, r, "Cash at Beginning of Year", beg_cash)
    r += 1
    _write_row(ws, r, "Cash at End of Year", end_cash, bold=True)
    r += 1


def _build_debt_summary(wb):
    """Sheet 4: Debt Summary (debt_schedule)."""
    ws = wb.create_sheet("Debt Summary")

    _write_title(ws, "Debt Summary", size=12)
    _write_headers(ws, row=3)

    r = 4

    rcf = [0.0, 0.0, 0.0, 0.0]
    term_loan = [45.0, 40.0, 35.0, 30.0]
    total_borrowings = [round(rcf[i] + term_loan[i], 1) for i in range(4)]
    # total_borrowings => [45.0, 40.0, 35.0, 30.0]

    interest_rate = [0.035, 0.038, 0.040, 0.042]
    interest_paid = [1.6, 1.5, 1.4, 1.3]

    _write_row(ws, r, "Revolving Credit Facility", rcf)
    r += 1
    _write_row(ws, r, "Term Loan", term_loan)
    r += 1
    _write_row(ws, r, "Total Borrowings", total_borrowings, bold=True)
    r += 1
    r += 1  # blank
    _write_row(ws, r, "Interest Rate", interest_rate, fmt=PCT_FMT)
    r += 1
    _write_row(ws, r, "Interest Paid", interest_paid)
    r += 1


def _build_assumptions(wb):
    """Sheet 5: Assumptions."""
    ws = wb.create_sheet("Assumptions")

    _write_title(ws, "Key Assumptions", size=12)
    _write_headers(ws, row=3)

    r = 4

    revenue_growth = [None, 0.072, 0.073, 0.068]
    gross_margin = [0.409, 0.413, 0.417, 0.422]
    tax_rate = [0.250, 0.250, 0.250, 0.250]
    capex_pct = [0.040, 0.043, 0.044, 0.044]
    dividend_payout = [0.0, 0.014, 0.017, 0.0]

    _write_row(ws, r, "Revenue Growth", revenue_growth, fmt=PCT_FMT)
    r += 1
    _write_row(ws, r, "Gross Margin", gross_margin, fmt=PCT_FMT)
    r += 1
    _write_row(ws, r, "Tax Rate", tax_rate, fmt=PCT_FMT)
    r += 1
    _write_row(ws, r, "Capex % Revenue", capex_pct, fmt=PCT_FMT)
    r += 1
    _write_row(ws, r, "Dividend Payout Ratio", dividend_payout, fmt=PCT_FMT)
    r += 1


def _build_board_notes(wb):
    """Sheet 6: Board Notes (skip — pure text)."""
    ws = wb.create_sheet("Board Notes")

    lines = [
        "Board Meeting - Q4 2024 Update",
        "",
        "Agenda Items:",
        "1. Review FY2024 performance vs budget",
        "2. Approve FY2025 plan",
        "3. Debt refinancing discussion",
        "4. Capital allocation strategy",
        "",
        "Confidential - Board Members Only",
    ]

    for idx, line in enumerate(lines, 1):
        ws.cell(row=idx, column=1, value=line).font = NORMAL_FONT

    ws.column_dimensions["A"].width = 50


# =============================================================================
# Main
# =============================================================================


def main():
    """Create the European mid-market IFRS model workbook."""
    wb = Workbook()

    _build_profit_and_loss(wb)
    _build_balance_sheet(wb)
    _build_cash_flow(wb)
    _build_debt_summary(wb)
    _build_assumptions(wb)
    _build_board_notes(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    print(f"Created: {OUTPUT}")
    print(f"Sheets:  {wb.sheetnames}")


if __name__ == "__main__":
    main()
