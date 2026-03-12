#!/usr/bin/env python3
"""Create tests/fixtures/edge_cases.xlsx — structural edge case model.

Generates a 4-sheet Excel workbook with structural challenges:
  - Sheet 1 (Financials): Combined IS + BS on one sheet, separated by blank rows
  - Sheet 2 (Quarterly Detail): Labels in column B, mixed quarterly + annual periods
  - Sheet 3 (Sensitivities): Sensitivity matrix (non-time-series)
  - Sheet 4 (Old Draft v1): Stale draft data that should be skipped

All values are deterministic. GP = Rev - COGS, EBIT = GP - OpEx, etc.

Usage:
    python scripts/create_edge_cases.py

Creates: tests/fixtures/edge_cases.xlsx
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

OUTPUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "edge_cases.xlsx"

NUMBER_FMT = "#,##0"


def _build_sheet1_financials(wb: openpyxl.Workbook) -> None:
    """Sheet 1: Combined IS + BS on one sheet (tier 1).

    Income Statement and Balance Sheet separated by 3 blank rows.
    This is the key edge case for the pipeline.
    """
    ws = wb.active
    ws.title = "Financials"
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)

    periods = ["FY2022", "FY2023", "FY2024E"]

    # --- Income Statement Section ---

    # Row 1: section title
    ws.cell(row=1, column=1, value="Income Statement").font = title_font

    # Row 2: headers
    ws.cell(row=2, column=1, value="").font = header_font
    for col_idx, period in enumerate(periods, 2):
        ws.cell(row=2, column=col_idx, value=period).font = header_font

    # P&L data (rows 3-18)
    # Revenue
    revenue = [50000, 55000, 60500]
    product_rev = [35000, 38500, 42350]
    service_rev = [15000, 16500, 18150]
    cogs = [-22500, -24750, -27225]
    gross_profit = [revenue[i] + cogs[i] for i in range(3)]  # 27500, 30250, 33275

    selling = [-7500, -8250, -9075]
    admin = [-5000, -5500, -6050]
    total_opex = [selling[i] + admin[i] for i in range(3)]  # -12500, -13750, -15125

    operating_income = [gross_profit[i] + total_opex[i] for i in range(3)]  # 15000, 16500, 18150
    interest_exp = [-2000, -1800, -1600]
    ebt = [operating_income[i] + interest_exp[i] for i in range(3)]  # 13000, 14700, 16550
    income_tax = [-3250, -3675, -4138]
    net_income = [9750, 11025, 12413]  # per spec (FY2024E has $1 rounding vs EBT-Tax)

    pl_rows = [
        (3, "Revenue", revenue, True),
        (4, "  Product Revenue", product_rev, False),
        (5, "  Service Revenue", service_rev, False),
        (6, "Cost of Goods Sold", cogs, False),
        (7, "Gross Profit", gross_profit, True),
        (8, None, None, False),  # blank
        (9, "  Selling Expenses", selling, False),
        (10, "  Administrative", admin, False),
        (11, "Total Operating Expenses", total_opex, True),
        (12, None, None, False),  # blank
        (13, "Operating Income", operating_income, True),
        (14, "Interest Expense", interest_exp, False),
        (15, "Earnings Before Tax", ebt, True),
        (16, "Income Tax", income_tax, False),
        (17, "Net Income", net_income, True),
    ]

    for row_num, label, values, is_bold in pl_rows:
        if label is not None:
            cell = ws.cell(row=row_num, column=1, value=label)
            if is_bold:
                cell.font = bold
        if values is not None:
            for col_idx, val in enumerate(values, 2):
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.number_format = NUMBER_FMT

    # Rows 18-20: BLANK (section separator — 3 blank rows)
    # (left empty)

    # --- Balance Sheet Section ---

    # Row 21: section title
    ws.cell(row=21, column=1, value="Balance Sheet").font = title_font

    # Row 22: headers
    ws.cell(row=22, column=1, value="").font = header_font
    for col_idx, period in enumerate(periods, 2):
        ws.cell(row=22, column=col_idx, value=period).font = header_font

    # BS data (rows 23 onward)
    cash = [8000, 12000, 18000]
    ar = [6850, 7534, 8288]
    inventory = [3800, 4100, 4400]
    total_ca = [cash[i] + ar[i] + inventory[i] for i in range(3)]  # 18650, 23634, 30688

    fixed_assets = [25000, 26000, 27000]
    intangibles = [5000, 4500, 4000]
    total_assets = [total_ca[i] + fixed_assets[i] + intangibles[i] for i in range(3)]  # 48650, 54134, 61688

    ap = [4500, 4950, 5445]
    accrued = [2500, 2750, 3025]
    total_cl = [ap[i] + accrued[i] for i in range(3)]  # 7000, 7700, 8470
    lt_debt = [15000, 13000, 11000]
    total_liab = [total_cl[i] + lt_debt[i] for i in range(3)]  # 22000, 20700, 19470

    total_equity = [total_assets[i] - total_liab[i] for i in range(3)]  # 26650, 33434, 42218
    total_le = [total_liab[i] + total_equity[i] for i in range(3)]  # = total_assets

    bs_rows = [
        (23, "Assets", None, True),
        (24, "  Cash", cash, False),
        (25, "  Accounts Receivable", ar, False),
        (26, "  Inventory", inventory, False),
        (27, "Total Current Assets", total_ca, True),
        (28, None, None, False),  # blank
        (29, "  Fixed Assets, Net", fixed_assets, False),
        (30, "  Intangibles", intangibles, False),
        (31, "Total Assets", total_assets, True),
        (32, None, None, False),  # blank
        (33, "Liabilities", None, True),
        (34, "  Accounts Payable", ap, False),
        (35, "  Accrued Expenses", accrued, False),
        (36, "Total Current Liabilities", total_cl, True),
        (37, "  Long-Term Debt", lt_debt, False),
        (38, "Total Liabilities", total_liab, True),
        (39, None, None, False),  # blank
        (40, "Total Equity", total_equity, True),
        (41, "Total Liabilities & Equity", total_le, True),
    ]

    for row_num, label, values, is_bold in bs_rows:
        if label is not None:
            cell = ws.cell(row=row_num, column=1, value=label)
            if is_bold:
                cell.font = bold
        if values is not None:
            for col_idx, val in enumerate(values, 2):
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.number_format = NUMBER_FMT

    # Set column widths
    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 14


def _build_sheet2_quarterly(wb: openpyxl.Workbook) -> None:
    """Sheet 2: Quarterly Detail (tier 2, income_statement).

    Labels in column B (col A has row numbers), mixed quarterly + annual periods.
    """
    ws = wb.create_sheet("Quarterly Detail")
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)

    # Row 1: title
    ws.cell(row=1, column=1, value="Quarterly Revenue Detail").font = title_font

    # Row 2: headers
    headers = ["#", "", "Q1 FY24", "Q2 FY24", "Q3 FY24", "Q4 FY24", "FY2024"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=header).font = bold

    # Data rows: col A = row number, col B = label, cols C-G = data
    revenue_q = [14000, 14500, 15500, 16500]
    revenue_annual = sum(revenue_q)  # 60500

    cogs_q = [-6300, -6525, -6975, -7425]
    cogs_annual = sum(cogs_q)  # -27225

    gp_q = [revenue_q[i] + cogs_q[i] for i in range(4)]  # 7700, 7975, 8525, 9075
    gp_annual = sum(gp_q)  # 33275

    opex_q = [-3500, -3600, -3800, -4225]
    opex_annual = sum(opex_q)  # -15125

    ebit_q = [gp_q[i] + opex_q[i] for i in range(4)]  # 4200, 4375, 4725, 4850
    ebit_annual = sum(ebit_q)  # 18150

    data_rows = [
        (3, 1, "Revenue", revenue_q + [revenue_annual], False),
        (4, 2, "COGS", cogs_q + [cogs_annual], False),
        (5, 3, "Gross Profit", gp_q + [gp_annual], True),
        (6, 4, "OpEx", opex_q + [opex_annual], False),
        (7, 5, "EBIT", ebit_q + [ebit_annual], True),
    ]

    for row_num, num, label, values, is_bold in data_rows:
        ws.cell(row=row_num, column=1, value=num)
        cell = ws.cell(row=row_num, column=2, value=label)
        if is_bold:
            cell.font = bold
        for col_idx, val in enumerate(values, 3):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.number_format = NUMBER_FMT

    # Set column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 14


def _build_sheet3_sensitivities(wb: openpyxl.Workbook) -> None:
    """Sheet 3: Sensitivities (tier 3, non-standard layout).

    Sensitivity matrix of EBITDA values, not a time series.
    """
    ws = wb.create_sheet("Sensitivities")
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)

    # Row 1: title
    ws.cell(row=1, column=1, value="Revenue Growth Sensitivity").font = title_font

    # Row 2: blank

    # Row 3: column headers
    row3_data = ["Revenue Growth \u2192", "", "5%", "8%", "10%", "12%", "15%"]
    for col_idx, val in enumerate(row3_data, 1):
        ws.cell(row=3, column=col_idx, value=val)

    # Row 4: row header label
    ws.cell(row=4, column=1, value="Margin \u2193").font = bold

    # Rows 5-9: matrix data
    matrix = [
        ("35%", [1058, 1134, 1188, 1243, 1328]),
        ("38%", [1150, 1233, 1292, 1352, 1444]),
        ("40%", [1210, 1298, 1360, 1423, 1520]),
        ("42%", [1271, 1363, 1428, 1494, 1596]),
        ("45%", [1363, 1462, 1531, 1601, 1710]),
    ]

    for row_offset, (margin_label, values) in enumerate(matrix):
        row_num = 5 + row_offset
        ws.cell(row=row_num, column=1, value=margin_label)
        for col_idx, val in enumerate(values, 3):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.number_format = NUMBER_FMT

    # Set column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 4
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 12


def _build_sheet4_old_draft(wb: openpyxl.Workbook) -> None:
    """Sheet 4: Old Draft v1 (tier 4, skip).

    Sheet name contains 'Draft' -> should be classified as tier 4.
    Only 1 column of data, stale financials.
    """
    ws = wb.create_sheet("Old Draft v1")
    red_bold = Font(bold=True, color="FF0000")

    # Row 1: warning header
    ws.cell(row=1, column=1, value="DRAFT - DO NOT USE").font = red_bold

    # Row 2: description
    ws.cell(row=2, column=1, value="This is an old version from Q1 2023")

    # Row 3: blank

    # Rows 4-10: stale data (single column)
    stale_data = [
        (4, "Revenue", 45000),
        (5, "COGS", -20000),
        (6, "Gross Profit", 25000),
        (7, None, None),
        (8, "Net Income", 8000),
    ]

    for row_num, label, value in stale_data:
        if label is not None:
            ws.cell(row=row_num, column=1, value=label)
        if value is not None:
            c = ws.cell(row=row_num, column=2, value=value)
            c.number_format = NUMBER_FMT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14


def main() -> None:
    """Create the edge cases workbook."""
    wb = openpyxl.Workbook()

    _build_sheet1_financials(wb)
    _build_sheet2_quarterly(wb)
    _build_sheet3_sensitivities(wb)
    _build_sheet4_old_draft(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))

    print(f"Created {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")
    print(f"Sheets: {wb.sheetnames}")

    # Verify internal consistency
    print("\nConsistency checks:")
    # IS checks
    assert 50000 - 22500 == 27500, "GP = Rev - COGS (FY2022)"
    assert 55000 - 24750 == 30250, "GP = Rev - COGS (FY2023)"
    assert 60500 - 27225 == 33275, "GP = Rev - COGS (FY2024E)"
    print("  IS: GP = Rev - COGS ... OK")

    assert 27500 - 12500 == 15000, "OI = GP - OpEx (FY2022)"
    assert 30250 - 13750 == 16500, "OI = GP - OpEx (FY2023)"
    assert 33275 - 15125 == 18150, "OI = GP - OpEx (FY2024E)"
    print("  IS: Operating Income = GP - OpEx ... OK")

    assert 15000 - 2000 == 13000, "EBT = OI - Interest (FY2022)"
    assert 16500 - 1800 == 14700, "EBT = OI - Interest (FY2023)"
    assert 18150 - 1600 == 16550, "EBT = OI - Interest (FY2024E)"
    print("  IS: EBT = OI - Interest ... OK")

    assert 13000 - 3250 == 9750, "NI = EBT - Tax (FY2022)"
    assert 14700 - 3675 == 11025, "NI = EBT - Tax (FY2023)"
    # FY2024E: 16550 - 4138 = 12412, spec rounds to 12413 ($1 rounding)
    print("  IS: Net Income = EBT - Tax ... OK (FY2024E has $1 rounding)")

    # BS checks
    assert 8000 + 6850 + 3800 == 18650, "Total CA (FY2022)"
    assert 18650 + 25000 + 5000 == 48650, "Total Assets (FY2022)"
    assert 4500 + 2500 == 7000, "Total CL (FY2022)"
    assert 7000 + 15000 == 22000, "Total Liab (FY2022)"
    assert 48650 - 22000 == 26650, "Equity = TA - TL (FY2022)"
    assert 22000 + 26650 == 48650, "TL&E = TA (FY2022)"
    print("  BS: Assets = Liabilities + Equity (FY2022) ... OK")

    assert 12000 + 7534 + 4100 == 23634, "Total CA (FY2023)"
    assert 23634 + 26000 + 4500 == 54134, "Total Assets (FY2023)"
    assert 4950 + 2750 == 7700, "Total CL (FY2023)"
    assert 7700 + 13000 == 20700, "Total Liab (FY2023)"
    assert 54134 - 20700 == 33434, "Equity = TA - TL (FY2023)"
    assert 20700 + 33434 == 54134, "TL&E = TA (FY2023)"
    print("  BS: Assets = Liabilities + Equity (FY2023) ... OK")

    assert 18000 + 8288 + 4400 == 30688, "Total CA (FY2024E)"
    assert 30688 + 27000 + 4000 == 61688, "Total Assets (FY2024E)"
    assert 5445 + 3025 == 8470, "Total CL (FY2024E)"
    assert 8470 + 11000 == 19470, "Total Liab (FY2024E)"
    assert 61688 - 19470 == 42218, "Equity = TA - TL (FY2024E)"
    assert 19470 + 42218 == 61688, "TL&E = TA (FY2024E)"
    print("  BS: Assets = Liabilities + Equity (FY2024E) ... OK")

    # Quarterly checks
    assert 14000 + 14500 + 15500 + 16500 == 60500, "Q sum = Annual Revenue"
    assert -6300 + -6525 + -6975 + -7425 == -27225, "Q sum = Annual COGS"
    assert 7700 + 7975 + 8525 + 9075 == 33275, "Q sum = Annual GP"
    assert -3500 + -3600 + -3800 + -4225 == -15125, "Q sum = Annual OpEx"
    assert 4200 + 4375 + 4725 + 4850 == 18150, "Q sum = Annual EBIT"
    print("  Quarterly: Q1+Q2+Q3+Q4 = Annual ... OK")

    print("\nAll consistency checks passed.")


if __name__ == "__main__":
    main()
