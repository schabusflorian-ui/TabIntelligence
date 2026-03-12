#!/usr/bin/env python3
"""
Create a realistic mid-market LBO financial model for benchmarking.

Generates an 8-sheet Excel workbook with:
  - Income Statement, Balance Sheet, Cash Flow Statement
  - Debt Schedule, Working Capital, Assumptions
  - D&A Schedule, Scratch - Notes (junk sheet)

Data represents a ~$250M revenue mid-market company across 5 periods
(FY2022A through FY2026E) with internally consistent financials.

Usage:
    python scripts/create_realistic_model.py

Creates: tests/fixtures/realistic_model.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


# =============================================================================
# Financial Data
# =============================================================================

PERIODS = ["FY2022A", "FY2023A", "FY2024E", "FY2025E", "FY2026E"]

# Assumptions (used to derive everything else)
ASSUMPTIONS = {
    "Revenue Growth Rate":       [None,   0.08,   0.10,   0.09,   0.07],
    "Gross Margin":              [0.545,  0.550,  0.555,  0.560,  0.565],
    "SG&A % of Revenue":        [0.120,  0.118,  0.115,  0.113,  0.110],
    "R&D % of Revenue":         [0.060,  0.062,  0.065,  0.063,  0.060],
    "Other OpEx % of Revenue":  [0.025,  0.024,  0.023,  0.022,  0.020],
    "Capex % of Revenue":       [0.045,  0.048,  0.050,  0.048,  0.045],
    "Tax Rate":                 [0.250,  0.250,  0.250,  0.250,  0.250],
    "Term Loan A Rate":         [0.045,  0.050,  0.055,  0.052,  0.048],
    "Senior Notes Rate":        [0.065,  0.065,  0.065,  0.065,  0.065],
    "DSO (days)":               [52,     50,     48,     47,     45],
    "DIO (days)":               [38,     36,     35,     34,     33],
    "DPO (days)":               [42,     44,     45,     46,     48],
}

# Base revenue for FY2022A
BASE_REVENUE = 232_000

# Derive revenue series
def _revenue_series():
    rev = [BASE_REVENUE]
    for i, g in enumerate(ASSUMPTIONS["Revenue Growth Rate"]):
        if i == 0:
            continue
        rev.append(round(rev[-1] * (1 + g)))
    return rev

REVENUE = _revenue_series()  # e.g. [232000, 250560, 275616, 300421, 321451]


def _build_income_statement():
    """Build income statement rows from assumptions."""
    rows = []

    # Revenue breakdown
    product_pct = [0.72, 0.71, 0.70, 0.69, 0.68]
    product_rev = [round(REVENUE[i] * product_pct[i]) for i in range(5)]
    service_rev = [REVENUE[i] - product_rev[i] for i in range(5)]

    cogs = [round(REVENUE[i] * (1 - ASSUMPTIONS["Gross Margin"][i])) for i in range(5)]
    gross_profit = [REVENUE[i] - cogs[i] for i in range(5)]

    sga = [round(REVENUE[i] * ASSUMPTIONS["SG&A % of Revenue"][i]) for i in range(5)]
    rd = [round(REVENUE[i] * ASSUMPTIONS["R&D % of Revenue"][i]) for i in range(5)]
    other_opex = [round(REVENUE[i] * ASSUMPTIONS["Other OpEx % of Revenue"][i]) for i in range(5)]
    total_opex = [sga[i] + rd[i] + other_opex[i] for i in range(5)]

    ebitda = [gross_profit[i] - total_opex[i] for i in range(5)]

    # D&A computed from D&A schedule
    da_values = _da_values()
    ebit = [ebitda[i] - da_values[i] for i in range(5)]

    # Interest from debt schedule
    interest = _interest_values()
    other_inc_exp = [round(REVENUE[i] * 0.003) for i in range(5)]  # small other expense

    pretax = [ebit[i] - interest[i] - other_inc_exp[i] for i in range(5)]
    tax = [round(pretax[i] * ASSUMPTIONS["Tax Rate"][i]) for i in range(5)]
    net_income = [pretax[i] - tax[i] for i in range(5)]

    rows.append(("Net Revenues", REVENUE, "bold"))
    rows.append(("  Product Revenue", product_rev, "indent"))
    rows.append(("  Service Revenue", service_rev, "indent"))
    rows.append(("Cost of Sales", cogs, "normal"))
    rows.append(("Gross Profit", gross_profit, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Operating Expenses:", [None]*5, "section"))
    rows.append(("  SG&A", sga, "indent"))
    rows.append(("  R&D Expenses", rd, "indent"))
    rows.append(("  Other Operating Expenses", other_opex, "indent"))
    rows.append(("Total Operating Expenses", total_opex, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("EBITDA", ebitda, "bold"))
    rows.append(("Depreciation & Amortization", da_values, "normal"))
    rows.append(("EBIT", ebit, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Interest Expense", interest, "normal"))
    rows.append(("Other (Income) / Expense, Net", other_inc_exp, "normal"))
    rows.append(("Pre-Tax Income", pretax, "bold"))
    rows.append(("Income Tax Provision", tax, "normal"))
    rows.append(("Net Income", net_income, "bold"))

    return rows, {
        "revenue": REVENUE, "cogs": cogs, "gross_profit": gross_profit,
        "sga": sga, "rd": rd, "other_opex": other_opex, "total_opex": total_opex,
        "ebitda": ebitda, "da": da_values, "ebit": ebit,
        "interest": interest, "other_inc_exp": other_inc_exp,
        "pretax": pretax, "tax": tax, "net_income": net_income,
    }


def _da_values():
    """Depreciation & Amortization schedule values."""
    # Existing asset base D&A declining; new capex D&A growing
    existing_da = [8200, 7800, 7400, 7000, 6600]
    capex = [round(REVENUE[i] * ASSUMPTIONS["Capex % of Revenue"][i]) for i in range(5)]
    # Assume capex depreciated straight-line over 7 years, partial year in year of purchase
    new_da = [0, 0, 0, 0, 0]
    for yr in range(5):
        for prev_yr in range(yr + 1):
            new_da[yr] += round(capex[prev_yr] / 7 * (1 if prev_yr < yr else 0.5))
    total_da = [existing_da[i] + new_da[i] for i in range(5)]
    return total_da


def _interest_values():
    """Interest expense from debt schedule."""
    tla_balance = [95000, 88000, 81000, 74000, 67000]
    senior_balance = [50000, 50000, 50000, 50000, 50000]

    tla_interest = [round(tla_balance[i] * ASSUMPTIONS["Term Loan A Rate"][i]) for i in range(5)]
    senior_interest = [round(senior_balance[i] * ASSUMPTIONS["Senior Notes Rate"][i]) for i in range(5)]
    total_interest = [tla_interest[i] + senior_interest[i] for i in range(5)]
    return total_interest


def _build_balance_sheet(is_data):
    """Build balance sheet from income statement data."""
    net_income = is_data["net_income"]

    # Current assets
    ar = [round(REVENUE[i] * ASSUMPTIONS["DSO (days)"][i] / 365) for i in range(5)]
    inventory = [round(is_data["cogs"][i] * ASSUMPTIONS["DIO (days)"][i] / 365) for i in range(5)]
    prepaid = [round(REVENUE[i] * 0.012) for i in range(5)]
    other_ca = [round(REVENUE[i] * 0.008) for i in range(5)]

    # PP&E
    capex = [round(REVENUE[i] * ASSUMPTIONS["Capex % of Revenue"][i]) for i in range(5)]
    ppe_net = [42000, 42000, 42000, 42000, 42000]  # initial
    for i in range(5):
        if i > 0:
            ppe_net[i] = ppe_net[i-1] + capex[i] - is_data["da"][i]
        else:
            ppe_net[i] = 42000

    goodwill = [85000] * 5
    intangibles_net = [18000, 16500, 15000, 13500, 12000]
    other_nca = [3500, 3200, 2900, 2600, 2300]

    total_ca = [ar[i] + inventory[i] + prepaid[i] + other_ca[i] for i in range(5)]

    # We need cash to make BS balance -- compute last
    # Non-current assets subtotal (before cash determination)
    total_nca_excl = [ppe_net[i] + goodwill[i] + intangibles_net[i] + other_nca[i] for i in range(5)]

    # Liabilities
    ap = [round(is_data["cogs"][i] * ASSUMPTIONS["DPO (days)"][i] / 365) for i in range(5)]
    accrued = [round(REVENUE[i] * 0.035) for i in range(5)]
    current_ltd = [7000] * 5  # annual TLA principal
    other_cl = [round(REVENUE[i] * 0.015) for i in range(5)]
    total_cl = [ap[i] + accrued[i] + current_ltd[i] + other_cl[i] for i in range(5)]

    tla_balance = [95000, 88000, 81000, 74000, 67000]
    senior_notes = [50000] * 5
    total_ltd = [tla_balance[i] + senior_notes[i] for i in range(5)]

    total_liab = [total_cl[i] + total_ltd[i] for i in range(5)]

    # Equity
    common_stock = [25000] * 5
    retained = [45000]
    for i in range(1, 5):
        retained.append(retained[-1] + net_income[i])

    total_equity = [common_stock[i] + retained[i] for i in range(5)]

    # Cash = Total L&E - Total CA (excl cash) - Total NCA
    # Total Assets = Total L&E, so Cash = Total L&E - (CA excl cash) - NCA
    total_le = [total_liab[i] + total_equity[i] for i in range(5)]
    cash = [total_le[i] - total_ca[i] - total_nca_excl[i] for i in range(5)]

    # Now add cash to get proper total CA
    total_ca_with_cash = [cash[i] + total_ca[i] for i in range(5)]
    total_assets = [total_ca_with_cash[i] + total_nca_excl[i] for i in range(5)]

    rows = []
    rows.append(("ASSETS", [None]*5, "section"))
    rows.append(("Current Assets", [None]*5, "subsection"))
    rows.append(("  Cash & Equivalents", cash, "indent"))
    rows.append(("  Accounts Receivable", ar, "indent"))
    rows.append(("  Inventory", inventory, "indent"))
    rows.append(("  Prepaid Expenses", prepaid, "indent"))
    rows.append(("  Other Current Assets", other_ca, "indent"))
    rows.append(("Total Current Assets", total_ca_with_cash, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Non-Current Assets", [None]*5, "subsection"))
    rows.append(("  PP&E (Net)", ppe_net, "indent"))
    rows.append(("  Goodwill", goodwill, "indent"))
    rows.append(("  Other Intangibles (Net)", intangibles_net, "indent"))
    rows.append(("  Other Non-Current Assets", other_nca, "indent"))
    rows.append(("Total Assets", total_assets, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("LIABILITIES", [None]*5, "section"))
    rows.append(("Current Liabilities", [None]*5, "subsection"))
    rows.append(("  Accounts Payable", ap, "indent"))
    rows.append(("  Accrued Liabilities", accrued, "indent"))
    rows.append(("  Current Portion of LT Debt", current_ltd, "indent"))
    rows.append(("  Other Current Liabilities", other_cl, "indent"))
    rows.append(("Total Current Liabilities", total_cl, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Long-Term Debt", [None]*5, "subsection"))
    rows.append(("  Term Loan A", tla_balance, "indent"))
    # Term Loan B is zero for this company
    rows.append(("  Senior Notes", senior_notes, "indent"))
    rows.append(("Total Long-Term Debt", total_ltd, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Total Liabilities", total_liab, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("SHAREHOLDERS' EQUITY", [None]*5, "section"))
    rows.append(("  Common Stock", common_stock, "indent"))
    rows.append(("  Retained Earnings", retained, "indent"))
    rows.append(("Total Shareholders' Equity", total_equity, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Total Liabilities & Shareholders' Equity", total_le, "bold"))

    return rows, {
        "cash": cash, "ar": ar, "inventory": inventory, "prepaid": prepaid,
        "other_ca": other_ca, "total_ca": total_ca_with_cash,
        "ppe_net": ppe_net, "goodwill": goodwill, "intangibles_net": intangibles_net,
        "other_nca": other_nca, "total_assets": total_assets,
        "ap": ap, "accrued": accrued, "current_ltd": current_ltd,
        "other_cl": other_cl, "total_cl": total_cl,
        "tla_balance": tla_balance, "senior_notes": senior_notes,
        "total_ltd": total_ltd, "total_liab": total_liab,
        "common_stock": common_stock, "retained": retained,
        "total_equity": total_equity, "total_le": total_le,
    }


def _build_cash_flow(is_data, bs_data):
    """Build cash flow statement."""
    net_income = is_data["net_income"]
    da = is_data["da"]

    # Working capital changes (delta in current assets/liabilities)
    chg_ar = [0] + [-(bs_data["ar"][i] - bs_data["ar"][i-1]) for i in range(1, 5)]
    chg_inv = [0] + [-(bs_data["inventory"][i] - bs_data["inventory"][i-1]) for i in range(1, 5)]
    chg_ap = [0] + [(bs_data["ap"][i] - bs_data["ap"][i-1]) for i in range(1, 5)]
    chg_other_wc = [0] + [
        -(bs_data["prepaid"][i] - bs_data["prepaid"][i-1])
        - (bs_data["other_ca"][i] - bs_data["other_ca"][i-1])
        + (bs_data["accrued"][i] - bs_data["accrued"][i-1])
        + (bs_data["other_cl"][i] - bs_data["other_cl"][i-1])
        for i in range(1, 5)
    ]
    total_wc_chg = [chg_ar[i] + chg_inv[i] + chg_ap[i] + chg_other_wc[i] for i in range(5)]

    cfo = [net_income[i] + da[i] + total_wc_chg[i] for i in range(5)]

    # Investing
    capex = [round(REVENUE[i] * ASSUMPTIONS["Capex % of Revenue"][i]) for i in range(5)]
    capex_neg = [-c for c in capex]
    cfi = capex_neg  # only capex for simplicity

    # Financing
    debt_repay = [0, -7000, -7000, -7000, -7000]  # TLA amortization
    cff = debt_repay

    net_change = [cfo[i] + cfi[i] + cff[i] for i in range(5)]

    # Cash reconciliation
    beg_cash = bs_data["cash"][:1] + bs_data["cash"][:4]  # beginning = prior period ending
    # Actually, for FY2022A we need a "prior" value; use the computed cash
    beg_cash = [bs_data["cash"][0] - net_change[0]] + bs_data["cash"][:4]
    end_cash = [beg_cash[i] + net_change[i] for i in range(5)]

    fcf = [cfo[i] + capex_neg[i] for i in range(5)]

    rows = []
    rows.append(("Cash from Operating Activities", [None]*5, "section"))
    rows.append(("  Net Income", net_income, "indent"))
    rows.append(("  Depreciation & Amortization", da, "indent"))
    rows.append(("  Changes in Accounts Receivable", chg_ar, "indent"))
    rows.append(("  Changes in Inventory", chg_inv, "indent"))
    rows.append(("  Changes in Accounts Payable", chg_ap, "indent"))
    rows.append(("  Other Working Capital Changes", chg_other_wc, "indent"))
    rows.append(("Cash from Operations", cfo, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Cash from Investing Activities", [None]*5, "section"))
    rows.append(("  Capital Expenditures", capex_neg, "indent"))
    rows.append(("Cash from Investing", cfi, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Cash from Financing Activities", [None]*5, "section"))
    rows.append(("  Debt Repayment", debt_repay, "indent"))
    rows.append(("Cash from Financing", cff, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Net Change in Cash", net_change, "bold"))
    rows.append(("Beginning Cash", beg_cash, "normal"))
    rows.append(("Ending Cash", end_cash, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Free Cash Flow", fcf, "bold"))

    return rows


def _build_debt_schedule():
    """Build debt schedule."""
    tla_beg = [100000, 95000, 88000, 81000, 74000]
    tla_repay = [-5000, -7000, -7000, -7000, -7000]
    tla_end = [tla_beg[i] + tla_repay[i] for i in range(5)]

    revolver_beg = [0, 0, 0, 0, 0]
    revolver_draw = [0, 0, 0, 0, 0]
    revolver_end = [0, 0, 0, 0, 0]

    senior_beg = [50000] * 5
    senior_end = [50000] * 5

    tla_rate = ASSUMPTIONS["Term Loan A Rate"]
    sr_rate = ASSUMPTIONS["Senior Notes Rate"]

    tla_interest = [round((tla_beg[i] + tla_end[i]) / 2 * tla_rate[i]) for i in range(5)]
    sr_interest = [round(senior_beg[i] * sr_rate[i]) for i in range(5)]
    total_interest = [tla_interest[i] + sr_interest[i] for i in range(5)]

    total_debt = [tla_end[i] + revolver_end[i] + senior_end[i] for i in range(5)]

    rows = []
    rows.append(("Term Loan A", [None]*5, "section"))
    rows.append(("  Beginning Balance", tla_beg, "indent"))
    rows.append(("  Mandatory Repayment", tla_repay, "indent"))
    rows.append(("  Ending Balance", tla_end, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Revolver", [None]*5, "section"))
    rows.append(("  Beginning Balance", revolver_beg, "indent"))
    rows.append(("  Draws / (Repayments)", revolver_draw, "indent"))
    rows.append(("  Ending Balance", revolver_end, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Senior Notes", [None]*5, "section"))
    rows.append(("  Beginning Balance", senior_beg, "indent"))
    rows.append(("  Ending Balance", senior_end, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Interest Expense", [None]*5, "subsection"))
    rows.append(("  Term Loan A Interest", tla_interest, "indent"))
    rows.append(("  Senior Notes Interest", sr_interest, "indent"))
    rows.append(("  Total Interest Expense", total_interest, "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Total Debt Outstanding", total_debt, "bold"))

    return rows


def _build_working_capital(is_data, bs_data):
    """Build working capital schedule."""
    dso = ASSUMPTIONS["DSO (days)"]
    dio = ASSUMPTIONS["DIO (days)"]
    dpo = ASSUMPTIONS["DPO (days)"]

    nwc = [
        bs_data["ar"][i] + bs_data["inventory"][i] - bs_data["ap"][i]
        for i in range(5)
    ]
    chg_nwc = [0] + [nwc[i] - nwc[i-1] for i in range(1, 5)]

    rows = []
    rows.append(("Days Sales Outstanding", dso, "normal"))
    rows.append(("Days Inventory Outstanding", dio, "normal"))
    rows.append(("Days Payable Outstanding", dpo, "normal"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Accounts Receivable", bs_data["ar"], "normal"))
    rows.append(("Inventory", bs_data["inventory"], "normal"))
    rows.append(("Accounts Payable", bs_data["ap"], "normal"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Net Working Capital", nwc, "bold"))
    rows.append(("Change in Net Working Capital", chg_nwc, "bold"))

    return rows


def _build_assumptions():
    """Build assumptions sheet."""
    rows = []
    for label, values in ASSUMPTIONS.items():
        if "Rate" in label or "%" in label or "Margin" in label:
            rows.append((label, values, "pct"))
        elif "days" in label.lower() or "DSO" in label or "DIO" in label or "DPO" in label:
            rows.append((label, values, "decimal"))
        else:
            rows.append((label, values, "pct"))
    return rows


def _build_da_schedule():
    """Build D&A schedule."""
    capex = [round(REVENUE[i] * ASSUMPTIONS["Capex % of Revenue"][i]) for i in range(5)]
    existing_da = [8200, 7800, 7400, 7000, 6600]
    new_da = [0, 0, 0, 0, 0]
    for yr in range(5):
        for prev_yr in range(yr + 1):
            new_da[yr] += round(capex[prev_yr] / 7 * (1 if prev_yr < yr else 0.5))
    total_da = [existing_da[i] + new_da[i] for i in range(5)]

    existing_ppe = [60000, 55000, 50000, 45000, 40000]
    amort = [1800, 1500, 1500, 1500, 1500]  # intangible amortization

    rows = []
    rows.append(("Existing PP&E (Gross)", existing_ppe, "normal"))
    rows.append(("  Depreciation - Existing Assets", existing_da, "indent"))
    rows.append(("New Capital Expenditures", capex, "normal"))
    rows.append(("  Depreciation - New Capex", new_da, "indent"))
    rows.append(("Total Depreciation", [existing_da[i] + new_da[i] - amort[i] for i in range(5)], "bold"))
    rows.append(("", [None]*5, "spacer"))
    rows.append(("Intangible Amortization", amort, "normal"))
    rows.append(("Total D&A", total_da, "bold"))

    return rows


def _build_scratch():
    """Build scratch/notes sheet (should be classified as SKIP)."""
    return [
        ("Notes and Working Calculations", "normal"),
        ("Last updated: 2024-01-15 by J.Smith", "normal"),
        ("TODO: Verify management fee assumptions with sponsor", "normal"),
        ("Check: PIK toggle on mezz tranche", "normal"),
        ("Sensitivity: +/- 100bps on base rate", "normal"),
        ("See email from 12/3 re: add-back adjustments", "normal"),
    ]


# =============================================================================
# Excel Formatting
# =============================================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
INDENT_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC")
)
TOTAL_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000"),
)

NUMBER_FMT = '#,##0'
PCT_FMT = '0.0%'
DECIMAL_FMT = '0.0'


def _write_sheet(ws, title, rows, is_numeric=True, is_pct_sheet=False):
    """Write a standard financial sheet with headers and data."""
    # Title row (merged)
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left")

    # Period headers
    header_row = 3
    ws.cell(row=header_row, column=1, value="").font = BOLD_FONT
    for col_idx, period in enumerate(PERIODS, 2):
        cell = ws.cell(row=header_row, column=col_idx, value=period)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.column_dimensions['A'].width = 38

    # Data rows
    start_row = 4
    for row_idx, row_data in enumerate(rows, start_row):
        label = row_data[0]
        values = row_data[1]
        style = row_data[2] if len(row_data) > 2 else "normal"

        # Label cell
        label_cell = ws.cell(row=row_idx, column=1, value=label)

        if style == "section":
            label_cell.font = SECTION_FONT
        elif style == "subsection":
            label_cell.font = BOLD_FONT
        elif style == "bold":
            label_cell.font = BOLD_FONT
        elif style == "indent":
            label_cell.font = INDENT_FONT
            label_cell.alignment = Alignment(indent=2)
        elif style == "spacer":
            continue
        else:
            label_cell.font = NORMAL_FONT

        if values is None or all(v is None for v in values):
            continue

        for col_idx, value in enumerate(values, 2):
            if value is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="right")

            if style == "pct" or is_pct_sheet:
                cell.number_format = PCT_FMT
            elif style == "decimal":
                cell.number_format = DECIMAL_FMT
            elif isinstance(value, float):
                cell.number_format = PCT_FMT if abs(value) < 1 else NUMBER_FMT
            else:
                cell.number_format = NUMBER_FMT

            if style == "bold":
                cell.font = BOLD_FONT
                cell.border = THIN_BORDER


def _write_scratch_sheet(ws, title, items):
    """Write the scratch/notes sheet (simple text, no financial data)."""
    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT

    for row_idx, item in enumerate(items, 3):
        text = item[0]
        ws.cell(row=row_idx, column=1, value=text).font = NORMAL_FONT

    ws.column_dimensions['A'].width = 55


# =============================================================================
# Main
# =============================================================================

def create_realistic_model():
    """Create the realistic financial model workbook."""
    wb = Workbook()

    # Build data
    is_rows, is_data = _build_income_statement()
    bs_rows, bs_data = _build_balance_sheet(is_data)
    cf_rows = _build_cash_flow(is_data, bs_data)
    debt_rows = _build_debt_schedule()
    wc_rows = _build_working_capital(is_data, bs_data)
    assumption_rows = _build_assumptions()
    da_rows = _build_da_schedule()
    scratch_items = _build_scratch()

    # Sheet 1: Income Statement
    ws_is = wb.active
    ws_is.title = "Income Statement"
    _write_sheet(ws_is, "Income Statement", is_rows)

    # Sheet 2: Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    _write_sheet(ws_bs, "Balance Sheet", bs_rows)

    # Sheet 3: Cash Flow Statement
    ws_cf = wb.create_sheet("Cash Flow Statement")
    _write_sheet(ws_cf, "Cash Flow Statement", cf_rows)

    # Sheet 4: Debt Schedule
    ws_debt = wb.create_sheet("Debt Schedule")
    _write_sheet(ws_debt, "Debt Schedule", debt_rows)

    # Sheet 5: Working Capital
    ws_wc = wb.create_sheet("Working Capital")
    _write_sheet(ws_wc, "Working Capital", wc_rows)

    # Sheet 6: Assumptions
    ws_assum = wb.create_sheet("Assumptions")
    _write_sheet(ws_assum, "Assumptions", assumption_rows, is_pct_sheet=False)

    # Sheet 7: D&A Schedule
    ws_da = wb.create_sheet("D&A Schedule")
    _write_sheet(ws_da, "D&A Schedule", da_rows)

    # Sheet 8: Scratch - Notes
    ws_scratch = wb.create_sheet("Scratch - Notes")
    _write_scratch_sheet(ws_scratch, "Scratch - Notes", scratch_items)

    # Save
    output_dir = Path("tests/fixtures")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "realistic_model.xlsx"
    wb.save(output_path)

    print(f"Created: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Revenue series: {REVENUE}")

    # Verify balance sheet balances
    for i, period in enumerate(PERIODS):
        ta = bs_data["total_assets"][i]
        tle = bs_data["total_le"][i]
        balanced = "OK" if ta == tle else f"MISMATCH: {ta} != {tle}"
        print(f"  {period}: Total Assets = {ta:,}, Total L&E = {tle:,} [{balanced}]")

    return output_path


if __name__ == "__main__":
    create_realistic_model()
