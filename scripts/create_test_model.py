#!/usr/bin/env python3
"""
Create a sample financial model for testing.

Usage:
    python scripts/create_test_model.py

Creates: tests/fixtures/sample_model.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def create_sample_model():
    """Create a simple 3-statement financial model for testing."""

    wb = Workbook()

    # =========================================================================
    # INCOME STATEMENT
    # =========================================================================
    ws_pl = wb.active
    ws_pl.title = "Income Statement"

    # Headers
    headers = ["", "FY2022", "FY2023", "FY2024E"]
    for col, header in enumerate(headers, 1):
        cell = ws_pl.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data
    pl_data = [
        ("Revenue", 100000, 115000, 132000),
        ("  Product Revenue", 70000, 80000, 92000),
        ("  Service Revenue", 30000, 35000, 40000),
        ("Cost of Goods Sold", 40000, 46000, 53000),
        ("Gross Profit", 60000, 69000, 79000),
        ("", None, None, None),
        ("Operating Expenses", None, None, None),
        ("  Sales & Marketing", 15000, 17000, 19000),
        ("  R&D", 10000, 12000, 14000),
        ("  G&A", 8000, 9000, 10000),
        ("Total Operating Expenses", 33000, 38000, 43000),
        ("", None, None, None),
        ("EBITDA", 27000, 31000, 36000),
        ("Depreciation & Amortization", 5000, 6000, 7000),
        ("EBIT", 22000, 25000, 29000),
        ("Interest Expense", 2000, 2500, 3000),
        ("EBT", 20000, 22500, 26000),
        ("Tax Expense", 5000, 5625, 6500),
        ("Net Income", 15000, 16875, 19500),
    ]

    for row_idx, (label, *values) in enumerate(pl_data, 2):
        ws_pl.cell(row=row_idx, column=1, value=label)
        for col_idx, value in enumerate(values, 2):
            if value is not None:
                ws_pl.cell(row=row_idx, column=col_idx, value=value)

    # Format headers
    ws_pl.column_dimensions["A"].width = 30

    # =========================================================================
    # BALANCE SHEET
    # =========================================================================
    ws_bs = wb.create_sheet("Balance Sheet")

    for col, header in enumerate(headers, 1):
        cell = ws_bs.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    bs_data = [
        ("ASSETS", None, None, None),
        ("Current Assets", None, None, None),
        ("  Cash & Cash Equivalents", 25000, 30000, 35000),
        ("  Accounts Receivable", 15000, 18000, 21000),
        ("  Inventory", 10000, 12000, 14000),
        ("Total Current Assets", 50000, 60000, 70000),
        ("", None, None, None),
        ("Non-Current Assets", None, None, None),
        ("  Property, Plant & Equipment", 40000, 45000, 50000),
        ("  Intangible Assets", 15000, 14000, 13000),
        ("  Goodwill", 20000, 20000, 20000),
        ("Total Non-Current Assets", 75000, 79000, 83000),
        ("", None, None, None),
        ("Total Assets", 125000, 139000, 153000),
        ("", None, None, None),
        ("LIABILITIES", None, None, None),
        ("Current Liabilities", None, None, None),
        ("  Accounts Payable", 8000, 10000, 12000),
        ("  Accrued Expenses", 5000, 6000, 7000),
        ("  Short-term Debt", 10000, 10000, 10000),
        ("Total Current Liabilities", 23000, 26000, 29000),
        ("", None, None, None),
        ("Non-Current Liabilities", None, None, None),
        ("  Long-term Debt", 30000, 35000, 40000),
        ("Total Non-Current Liabilities", 30000, 35000, 40000),
        ("", None, None, None),
        ("Total Liabilities", 53000, 61000, 69000),
        ("", None, None, None),
        ("EQUITY", None, None, None),
        ("  Common Stock", 50000, 50000, 50000),
        ("  Retained Earnings", 22000, 28000, 34000),
        ("Total Equity", 72000, 78000, 84000),
        ("", None, None, None),
        ("Total Liabilities & Equity", 125000, 139000, 153000),
    ]

    for row_idx, (label, *values) in enumerate(bs_data, 2):
        ws_bs.cell(row=row_idx, column=1, value=label)
        for col_idx, value in enumerate(values, 2):
            if value is not None:
                ws_bs.cell(row=row_idx, column=col_idx, value=value)

    ws_bs.column_dimensions["A"].width = 30

    # =========================================================================
    # CASH FLOW
    # =========================================================================
    ws_cf = wb.create_sheet("Cash Flow")

    for col, header in enumerate(headers, 1):
        cell = ws_cf.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    cf_data = [
        ("Cash from Operating Activities", None, None, None),
        ("  Net Income", 15000, 16875, 19500),
        ("  Depreciation & Amortization", 5000, 6000, 7000),
        ("  Changes in Working Capital", -2000, -3000, -4000),
        ("Cash from Operations", 18000, 19875, 22500),
        ("", None, None, None),
        ("Cash from Investing Activities", None, None, None),
        ("  Capital Expenditures", -8000, -10000, -12000),
        ("  Acquisitions", 0, 0, 0),
        ("Cash from Investing", -8000, -10000, -12000),
        ("", None, None, None),
        ("Cash from Financing Activities", None, None, None),
        ("  Debt Proceeds", 5000, 5000, 5000),
        ("  Dividends", -3000, -4000, -5000),
        ("Cash from Financing", 2000, 1000, 0),
        ("", None, None, None),
        ("Net Change in Cash", 12000, 10875, 10500),
        ("Beginning Cash", 13000, 25000, 35875),
        ("Ending Cash", 25000, 35875, 46375),
    ]

    for row_idx, (label, *values) in enumerate(cf_data, 2):
        ws_cf.cell(row=row_idx, column=1, value=label)
        for col_idx, value in enumerate(values, 2):
            if value is not None:
                ws_cf.cell(row=row_idx, column=col_idx, value=value)

    ws_cf.column_dimensions["A"].width = 30

    # =========================================================================
    # SCRATCH SHEET (should be skipped)
    # =========================================================================
    ws_scratch = wb.create_sheet("Scratch - Working")
    ws_scratch.cell(row=1, column=1, value="Random notes and calculations")
    ws_scratch.cell(row=2, column=1, value="TODO: fix formula")
    ws_scratch.cell(row=3, column=1, value="Check with John")

    # =========================================================================
    # SAVE
    # =========================================================================
    output_dir = Path("tests/fixtures")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "sample_model.xlsx"

    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"Sheets: {wb.sheetnames}")

    return output_path


if __name__ == "__main__":
    create_sample_model()
