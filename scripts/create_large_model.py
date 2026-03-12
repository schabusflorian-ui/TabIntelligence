#!/usr/bin/env python3
"""Create tests/fixtures/large_model.xlsx -- large 12-sheet corporate model (~200 items)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "large_model.xlsx"
PERIODS = ["FY2022A", "FY2023A", "FY2024E", "FY2025E", "FY2026E"]

# ---------------------------------------------------------------------------
# Shared helper
# ---------------------------------------------------------------------------

BOLD_FONT = Font(name="Calibri", size=12, bold=True)
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
NUMBER_FMT = "#,##0"


def _write_financial_sheet(ws, title, headers, rows_data):
    """Write a standard financial sheet.

    Row 1: title (bold, size 12)
    Row 2: headers (bold)
    Subsequent rows: label in col A, values in cols B-F.
    Items with "Total" in the label or tagged bold get bold font.
    All numeric cells use '#,##0' format.
    """
    # Row 1 -- title
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = BOLD_FONT

    # Row 2 -- headers
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = HEADER_FONT

    # Set column widths
    ws.column_dimensions["A"].width = 38
    for col_letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 14

    # Data rows starting at row 3
    for row_offset, item in enumerate(rows_data):
        row_num = row_offset + 3
        if isinstance(item, tuple) and len(item) == 3:
            label, values, is_bold = item
        elif isinstance(item, tuple) and len(item) == 2:
            label, values = item
            is_bold = False
        else:
            continue

        # Determine boldness: explicit flag or "Total" in label
        make_bold = is_bold or "Total" in label

        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = BOLD_FONT if make_bold else NORMAL_FONT

        if values is None:
            continue
        for col_idx, val in enumerate(values, 2):
            if val is None:
                continue
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.font = BOLD_FONT if make_bold else NORMAL_FONT
            if isinstance(val, (int, float)):
                c.number_format = NUMBER_FMT


# ============================================================================
# INTEGRATED FINANCIAL MODEL
#
# Build order:
#   1. Income Statement (standalone)
#   2. Balance Sheet non-cash items (standalone assumptions)
#   3. Cash Flow Statement (derived from IS + BS deltas)
#   4. BS Cash = CF ending cash
#   5. APIC = plug to make BS balance
# ============================================================================

# ---------------------------------------------------------------------------
# Income Statement  (sheet 1, ~22 named items)
# ---------------------------------------------------------------------------

IS_NET_REVENUE = [500000, 530000, 561800, 595508, 631238]
IS_PRODUCT_A = [200000, 212000, 224720, 238203, 252495]
IS_PRODUCT_B = [150000, 159000, 168540, 178652, 189371]
IS_PRODUCT_C = [100000, 106000, 112360, 119102, 126248]
IS_LICENSING = [50000, 53000, 56180, 59551, 63124]
IS_COGS = [-225000, -238500, -252810, -267978, -284057]
IS_GROSS_PROFIT = [275000, 291500, 308990, 327530, 347181]
IS_SM = [-50000, -53000, -56180, -59551, -63124]
IS_RD = [-35000, -37100, -39326, -41685, -44186]
IS_GA = [-30000, -31800, -33708, -35730, -37874]
IS_DEPR = [-15000, -15900, -16854, -17865, -18937]
IS_AMORT = [-5000, -5000, -5000, -5000, -5000]
IS_TOTAL_OPEX = [-135000, -142800, -151068, -159831, -169121]
IS_OPER_INCOME = [140000, 148700, 157922, 167699, 178060]
IS_INTEREST = [-12000, -11000, -10000, -9000, -8000]
IS_OTHER_INCOME = [2000, 2100, 2200, 2300, 2400]
IS_IBT = [130000, 139800, 150122, 160999, 172460]
IS_TAX = [-32500, -34950, -37531, -40250, -43115]
IS_NET_INCOME = [97500, 104850, 112591, 120749, 129345]

# Non-cash items for CF
SBC = [8000, 8500, 9000, 9500, 10000]

# ---------------------------------------------------------------------------
# Balance Sheet non-cash items (fixed assumptions)
# ---------------------------------------------------------------------------

# Current Assets (excl cash)
BS_AR = [68493, 72603, 76959, 81581, 86477]
BS_INVENTORY = [30822, 32671, 34633, 36692, 38854]
BS_PREPAID = [5000, 5300, 5618, 5955, 6312]
BS_OTHER_CA = [3000, 3180, 3371, 3573, 3787]

# Non-Current Assets
BS_PPE = [120000, 122100, 125246, 128381, 131444]
BS_GOODWILL = [150000, 150000, 155000, 155000, 155000]
BS_INTANGIBLES = [40000, 35000, 30000, 25000, 20000]
BS_DTA = [5000, 5000, 5000, 5000, 5000]
BS_ROU = [25000, 23000, 21000, 19000, 17000]
BS_OTHER_NCA = [8000, 7500, 7000, 6500, 6000]

# Current Liabilities
BS_AP = [30822, 32671, 34633, 36692, 38854]
BS_ACCRUED = [25000, 26500, 28090, 29775, 31562]
BS_CURRENT_DEBT = [10000, 10000, 10000, 10000, 10000]
BS_DEFERRED_REV = [15000, 15900, 16854, 17865, 18937]
BS_LEASE_CL = [5000, 5000, 5000, 5000, 5000]
BS_OTHER_CL = [4000, 4240, 4494, 4764, 5049]

# Non-Current Liabilities
BS_LTD = [200000, 190000, 180000, 170000, 160000]
BS_DTL = [12000, 12000, 12000, 12000, 12000]
BS_LEASE_NCL = [20000, 18000, 16000, 14000, 12000]
BS_OTHER_NCL = [6000, 5500, 5000, 4500, 4000]

# Equity (excl APIC which is the plug)
BS_COMMON = [10000, 10000, 10000, 10000, 10000]

# Dividends
DIVIDENDS = [-20000, -22000, -24000, -26000, -28000]

# Retained earnings: start at 50000, roll forward with NI + dividends
BS_RETAINED = [50000]
for _i in range(1, 5):
    BS_RETAINED.append(BS_RETAINED[-1] + IS_NET_INCOME[_i] + DIVIDENDS[_i])

# ---------------------------------------------------------------------------
# Cash Flow Statement  (sheet 3, ~25 items)
# Derived from IS + BS delta movements
# ---------------------------------------------------------------------------

# D&A add-back (positive)
CF_DA = [abs(IS_DEPR[i]) + abs(IS_AMORT[i]) for i in range(5)]

# Working capital changes (year-over-year deltas)
CF_CHG_AR = [0] + [-(BS_AR[i] - BS_AR[i - 1]) for i in range(1, 5)]
CF_CHG_INV = [0] + [-(BS_INVENTORY[i] - BS_INVENTORY[i - 1]) for i in range(1, 5)]
CF_CHG_AP = [0] + [(BS_AP[i] - BS_AP[i - 1]) for i in range(1, 5)]
CF_CHG_ACCRUED = [0] + [(BS_ACCRUED[i] - BS_ACCRUED[i - 1]) for i in range(1, 5)]
CF_CHG_DEFREV = [0] + [(BS_DEFERRED_REV[i] - BS_DEFERRED_REV[i - 1]) for i in range(1, 5)]

CF_CFO = [
    IS_NET_INCOME[i] + CF_DA[i] + SBC[i]
    + CF_CHG_AR[i] + CF_CHG_INV[i] + CF_CHG_AP[i]
    + CF_CHG_ACCRUED[i] + CF_CHG_DEFREV[i]
    for i in range(5)
]

# Investing -- capex = net PP&E change + depreciation
CAPEX = [0] + [-(BS_PPE[i] - BS_PPE[i - 1] + abs(IS_DEPR[i])) for i in range(1, 5)]
CAPEX[0] = -18000  # initial year capex assumption
ACQUISITIONS = [0, 0, -5000, 0, 0]  # goodwill jump in FY2024E
CF_CFI = [CAPEX[i] + ACQUISITIONS[i] for i in range(5)]

# Financing
DEBT_REPAY = [-10000, -10000, -10000, -10000, -10000]
SHARE_REPURCHASE = [0, 0, 0, 0, 0]
CF_CFF = [DEBT_REPAY[i] + DIVIDENDS[i] + SHARE_REPURCHASE[i] for i in range(5)]

CF_NET_CHANGE = [CF_CFO[i] + CF_CFI[i] + CF_CFF[i] for i in range(5)]

# Cash from CF: start at 25000, roll forward
_STARTING_CASH = 25000
CF_BEG_CASH = [_STARTING_CASH]
CF_END_CASH = [CF_BEG_CASH[0] + CF_NET_CHANGE[0]]
for _i in range(1, 5):
    CF_BEG_CASH.append(CF_END_CASH[-1])
    CF_END_CASH.append(CF_BEG_CASH[-1] + CF_NET_CHANGE[_i])

# BS Cash = CF Ending Cash (this is the key linkage)
BS_CASH = CF_END_CASH[:]

# ---------------------------------------------------------------------------
# Balance Sheet totals (with cash now known)
# ---------------------------------------------------------------------------

BS_TOTAL_CA_EX_CASH = [
    BS_AR[i] + BS_INVENTORY[i] + BS_PREPAID[i] + BS_OTHER_CA[i]
    for i in range(5)
]
BS_TOTAL_CA = [BS_CASH[i] + BS_TOTAL_CA_EX_CASH[i] for i in range(5)]
BS_TOTAL_NCA = [
    BS_PPE[i] + BS_GOODWILL[i] + BS_INTANGIBLES[i] + BS_DTA[i]
    + BS_ROU[i] + BS_OTHER_NCA[i]
    for i in range(5)
]
BS_TOTAL_ASSETS = [BS_TOTAL_CA[i] + BS_TOTAL_NCA[i] for i in range(5)]

BS_TOTAL_CL = [
    BS_AP[i] + BS_ACCRUED[i] + BS_CURRENT_DEBT[i] + BS_DEFERRED_REV[i]
    + BS_LEASE_CL[i] + BS_OTHER_CL[i]
    for i in range(5)
]
BS_TOTAL_NCL = [
    BS_LTD[i] + BS_DTL[i] + BS_LEASE_NCL[i] + BS_OTHER_NCL[i]
    for i in range(5)
]
BS_TOTAL_LIAB = [BS_TOTAL_CL[i] + BS_TOTAL_NCL[i] for i in range(5)]

# APIC = plug: Total Assets - Total Liabilities - Common - Retained
BS_APIC = [
    BS_TOTAL_ASSETS[i] - BS_TOTAL_LIAB[i] - BS_COMMON[i] - BS_RETAINED[i]
    for i in range(5)
]
BS_TOTAL_EQUITY = [BS_COMMON[i] + BS_APIC[i] + BS_RETAINED[i] for i in range(5)]
BS_TOTAL_LE = [BS_TOTAL_LIAB[i] + BS_TOTAL_EQUITY[i] for i in range(5)]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_income_statement(ws):
    rows = [
        ("Net Revenue", IS_NET_REVENUE, True),
        ("  Product A", IS_PRODUCT_A, False),
        ("  Product B", IS_PRODUCT_B, False),
        ("  Product C", IS_PRODUCT_C, False),
        ("  Licensing Revenue", IS_LICENSING, False),
        ("", None, False),
        ("Cost of Goods Sold", IS_COGS, False),
        ("Gross Profit", IS_GROSS_PROFIT, True),
        ("", None, False),
        ("  Sales & Marketing", IS_SM, False),
        ("  Research & Development", IS_RD, False),
        ("  General & Administrative", IS_GA, False),
        ("  Depreciation", IS_DEPR, False),
        ("  Amortization", IS_AMORT, False),
        ("Total Operating Expenses", IS_TOTAL_OPEX, True),
        ("", None, False),
        ("Operating Income", IS_OPER_INCOME, True),
        ("Interest Expense", IS_INTEREST, False),
        ("Other Income", IS_OTHER_INCOME, False),
        ("Income Before Tax", IS_IBT, True),
        ("Income Tax Expense", IS_TAX, False),
        ("Net Income", IS_NET_INCOME, True),
    ]
    _write_financial_sheet(ws, "Income Statement", ["", *PERIODS], rows)


def _build_balance_sheet(ws):
    rows = [
        ("ASSETS", None, True),
        ("Current Assets:", None, True),
        ("  Cash & Cash Equivalents", BS_CASH, False),
        ("  Accounts Receivable", BS_AR, False),
        ("  Inventory", BS_INVENTORY, False),
        ("  Prepaid Expenses", BS_PREPAID, False),
        ("  Other Current Assets", BS_OTHER_CA, False),
        ("Total Current Assets", BS_TOTAL_CA, True),
        ("", None, False),
        ("Non-Current Assets:", None, True),
        ("  Net PP&E", BS_PPE, False),
        ("  Goodwill", BS_GOODWILL, False),
        ("  Intangible Assets", BS_INTANGIBLES, False),
        ("  Deferred Tax Asset", BS_DTA, False),
        ("  Right-of-Use Asset", BS_ROU, False),
        ("  Other Non-Current Assets", BS_OTHER_NCA, False),
        ("Total Non-Current Assets", BS_TOTAL_NCA, True),
        ("Total Assets", BS_TOTAL_ASSETS, True),
        ("", None, False),
        ("LIABILITIES", None, True),
        ("Current Liabilities:", None, True),
        ("  Accounts Payable", BS_AP, False),
        ("  Accrued Expenses", BS_ACCRUED, False),
        ("  Current Portion of Debt", BS_CURRENT_DEBT, False),
        ("  Deferred Revenue", BS_DEFERRED_REV, False),
        ("  Lease Liability - Current", BS_LEASE_CL, False),
        ("  Other Current Liabilities", BS_OTHER_CL, False),
        ("Total Current Liabilities", BS_TOTAL_CL, True),
        ("", None, False),
        ("Non-Current Liabilities:", None, True),
        ("  Long-term Debt", BS_LTD, False),
        ("  Deferred Tax Liability", BS_DTL, False),
        ("  Lease Liability - Non-Current", BS_LEASE_NCL, False),
        ("  Other Non-Current Liabilities", BS_OTHER_NCL, False),
        ("Total Non-Current Liabilities", BS_TOTAL_NCL, True),
        ("Total Liabilities", BS_TOTAL_LIAB, True),
        ("", None, False),
        ("SHAREHOLDERS' EQUITY", None, True),
        ("  Common Stock", BS_COMMON, False),
        ("  Additional Paid-In Capital", BS_APIC, False),
        ("  Retained Earnings", BS_RETAINED, False),
        ("Total Shareholders' Equity", BS_TOTAL_EQUITY, True),
        ("Total Liabilities & Equity", BS_TOTAL_LE, True),
    ]
    _write_financial_sheet(ws, "Balance Sheet", ["", *PERIODS], rows)


def _build_cash_flow(ws):
    rows = [
        ("Operating Activities:", None, True),
        ("  Net Income", IS_NET_INCOME, False),
        ("  Depreciation & Amortization", CF_DA, False),
        ("  Stock-Based Compensation", SBC, False),
        ("  Change in Accounts Receivable", CF_CHG_AR, False),
        ("  Change in Inventory", CF_CHG_INV, False),
        ("  Change in Accounts Payable", CF_CHG_AP, False),
        ("  Change in Accrued Expenses", CF_CHG_ACCRUED, False),
        ("  Change in Deferred Revenue", CF_CHG_DEFREV, False),
        ("Cash from Operations", CF_CFO, True),
        ("", None, False),
        ("Investing Activities:", None, True),
        ("  Capital Expenditures", CAPEX, False),
        ("  Acquisitions", ACQUISITIONS, False),
        ("Cash from Investing", CF_CFI, True),
        ("", None, False),
        ("Financing Activities:", None, True),
        ("  Debt Repayment", DEBT_REPAY, False),
        ("  Dividends Paid", DIVIDENDS, False),
        ("  Share Repurchases", SHARE_REPURCHASE, False),
        ("Cash from Financing", CF_CFF, True),
        ("", None, False),
        ("Net Change in Cash", CF_NET_CHANGE, True),
        ("Beginning Cash", CF_BEG_CASH, False),
        ("Ending Cash", CF_END_CASH, True),
    ]
    _write_financial_sheet(ws, "Cash Flow Statement", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Debt Schedule  (sheet 4, ~26 items)
# ---------------------------------------------------------------------------

TLA_BEG = [120000, 110000, 100000, 90000, 80000]
TLA_DRAWS = [0, 0, 0, 0, 0]
TLA_REPAY = [-10000, -10000, -10000, -10000, -10000]
TLA_END = [TLA_BEG[i] + TLA_DRAWS[i] + TLA_REPAY[i] for i in range(5)]
TLA_RATE = [0.045, 0.050, 0.055, 0.050, 0.048]
TLA_INTEREST = [round((TLA_BEG[i] + TLA_END[i]) / 2 * TLA_RATE[i]) for i in range(5)]

TLB_BEG = [80000, 80000, 80000, 80000, 80000]
TLB_DRAWS = [0, 0, 0, 0, 0]
TLB_REPAY = [0, 0, 0, 0, 0]
TLB_END = [80000, 80000, 80000, 80000, 80000]
TLB_RATE = [0.065, 0.065, 0.065, 0.065, 0.065]
TLB_INTEREST = [round(TLB_BEG[i] * TLB_RATE[i]) for i in range(5)]

RCF_BEG = [0, 0, 0, 0, 0]
RCF_DRAWS = [0, 0, 0, 0, 0]
RCF_REPAY = [0, 0, 0, 0, 0]
RCF_END = [0, 0, 0, 0, 0]
RCF_RATE = [0.040, 0.045, 0.050, 0.045, 0.043]
RCF_INTEREST = [0, 0, 0, 0, 0]

DEBT_TOTAL = [TLA_END[i] + TLB_END[i] + RCF_END[i] for i in range(5)]
DEBT_TOTAL_INTEREST = [TLA_INTEREST[i] + TLB_INTEREST[i] + RCF_INTEREST[i] for i in range(5)]


def _build_debt_schedule(ws):
    rows = [
        ("Term Loan A", None, True),
        ("  Beginning Balance", TLA_BEG, False),
        ("  Draws", TLA_DRAWS, False),
        ("  Repayments", TLA_REPAY, False),
        ("  Ending Balance", TLA_END, True),
        ("  Interest Rate", TLA_RATE, False),
        ("  Interest Expense", TLA_INTEREST, False),
        ("", None, False),
        ("Term Loan B", None, True),
        ("  Beginning Balance", TLB_BEG, False),
        ("  Draws", TLB_DRAWS, False),
        ("  Repayments", TLB_REPAY, False),
        ("  Ending Balance", TLB_END, True),
        ("  Interest Rate", TLB_RATE, False),
        ("  Interest Expense", TLB_INTEREST, False),
        ("", None, False),
        ("Revolving Credit Facility", None, True),
        ("  Beginning Balance", RCF_BEG, False),
        ("  Draws", RCF_DRAWS, False),
        ("  Repayments", RCF_REPAY, False),
        ("  Ending Balance", RCF_END, True),
        ("  Interest Rate", RCF_RATE, False),
        ("  Interest Expense", RCF_INTEREST, False),
        ("", None, False),
        ("Total Debt", DEBT_TOTAL, True),
        ("Total Interest Expense", DEBT_TOTAL_INTEREST, True),
    ]
    _write_financial_sheet(ws, "Debt Schedule", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Revenue Build  (sheet 5, ~21 items)
# ---------------------------------------------------------------------------

PA_NA = [120000, 127200, 134832, 142922, 151497]
PA_EU = [50000, 53000, 56180, 59551, 63124]
PA_ASIA = [30000, 31800, 33708, 35730, 37874]
PA_TOTAL = [PA_NA[i] + PA_EU[i] + PA_ASIA[i] for i in range(5)]

PB_NA = [90000, 95400, 101124, 107191, 113623]
PB_EU = [40000, 42400, 44944, 47641, 50499]
PB_ASIA = [20000, 21200, 22472, 23820, 25249]
PB_TOTAL = [PB_NA[i] + PB_EU[i] + PB_ASIA[i] for i in range(5)]

PC_DOM = [70000, 74200, 78652, 83371, 88373]
PC_INTL = [30000, 31800, 33708, 35731, 37875]
PC_TOTAL = [PC_DOM[i] + PC_INTL[i] for i in range(5)]

LIC_TOTAL = IS_LICENSING[:]

RB_TOTAL = [PA_TOTAL[i] + PB_TOTAL[i] + PC_TOTAL[i] + LIC_TOTAL[i] for i in range(5)]


def _build_revenue_build(ws):
    rows = [
        ("Product A", None, True),
        ("  North America", PA_NA, False),
        ("  Europe", PA_EU, False),
        ("  Asia", PA_ASIA, False),
        ("  Total Product A", PA_TOTAL, True),
        ("", None, False),
        ("Product B", None, True),
        ("  North America", PB_NA, False),
        ("  Europe", PB_EU, False),
        ("  Asia", PB_ASIA, False),
        ("  Total Product B", PB_TOTAL, True),
        ("", None, False),
        ("Product C", None, True),
        ("  Domestic", PC_DOM, False),
        ("  International", PC_INTL, False),
        ("  Total Product C", PC_TOTAL, True),
        ("", None, False),
        ("Licensing", None, True),
        ("  Total Licensing", LIC_TOTAL, True),
        ("", None, False),
        ("Total Revenue", RB_TOTAL, True),
    ]
    _write_financial_sheet(ws, "Revenue Build", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# OpEx Build  (sheet 6, ~24 items)
# ---------------------------------------------------------------------------

SM_HC = [250, 260, 270, 280, 290]
SM_COMP = [-30000, -31800, -33708, -35730, -37874]
SM_COMM = [-8000, -8480, -8989, -9528, -10100]
SM_TRAVEL = [-4000, -4240, -4494, -4764, -5049]
SM_MKTG = [-8000, -8480, -8989, -9529, -10101]
SM_TOTAL = [SM_COMP[i] + SM_COMM[i] + SM_TRAVEL[i] + SM_MKTG[i] for i in range(5)]

RD_HC = [175, 182, 190, 198, 206]
RD_COMP = [-25000, -26500, -28090, -29775, -31562]
RD_CLOUD = [-6000, -6360, -6742, -7146, -7575]
RD_TOOLS = [-4000, -4240, -4494, -4764, -5049]
RD_TOTAL = [RD_COMP[i] + RD_CLOUD[i] + RD_TOOLS[i] for i in range(5)]

GA_HC = [60, 62, 64, 66, 68]
GA_COMP = [-16000, -16960, -17978, -19057, -20200]
GA_PROF = [-6000, -6360, -6742, -7146, -7575]
GA_FAC = [-5000, -5300, -5618, -5955, -6312]
GA_INS = [-3000, -3180, -3370, -3572, -3787]
GA_TOTAL = [GA_COMP[i] + GA_PROF[i] + GA_FAC[i] + GA_INS[i] for i in range(5)]

OPEX_CASH_TOTAL = [SM_TOTAL[i] + RD_TOTAL[i] + GA_TOTAL[i] for i in range(5)]


def _build_opex_build(ws):
    rows = [
        ("Sales & Marketing", None, True),
        ("  Headcount", SM_HC, False),
        ("  Compensation", SM_COMP, False),
        ("  Commissions", SM_COMM, False),
        ("  Travel & Entertainment", SM_TRAVEL, False),
        ("  Marketing Spend", SM_MKTG, False),
        ("  Total Sales & Marketing", SM_TOTAL, True),
        ("", None, False),
        ("Research & Development", None, True),
        ("  Headcount", RD_HC, False),
        ("  Compensation", RD_COMP, False),
        ("  Cloud / Hosting", RD_CLOUD, False),
        ("  Tools & Licenses", RD_TOOLS, False),
        ("  Total R&D", RD_TOTAL, True),
        ("", None, False),
        ("General & Administrative", None, True),
        ("  Headcount", GA_HC, False),
        ("  Compensation", GA_COMP, False),
        ("  Professional Fees", GA_PROF, False),
        ("  Facilities", GA_FAC, False),
        ("  Insurance", GA_INS, False),
        ("  Total G&A", GA_TOTAL, True),
        ("", None, False),
        ("Total Cash Operating Expenses", OPEX_CASH_TOTAL, True),
    ]
    _write_financial_sheet(ws, "OpEx Build", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Working Capital  (sheet 7, ~13 items)
# ---------------------------------------------------------------------------

WC_AR_DAYS = [50, 50, 50, 50, 50]
WC_INV_DAYS = [50, 50, 50, 50, 50]
WC_AP_DAYS = [50, 50, 50, 50, 50]

WC_NWC = [
    BS_AR[i] + BS_INVENTORY[i] + BS_PREPAID[i]
    - BS_AP[i] - BS_ACCRUED[i] - BS_DEFERRED_REV[i]
    for i in range(5)
]
WC_CHG = [0] + [WC_NWC[i] - WC_NWC[i - 1] for i in range(1, 5)]


def _build_working_capital(ws):
    rows = [
        ("Accounts Receivable", BS_AR, False),
        ("  AR Days", WC_AR_DAYS, False),
        ("Inventory", BS_INVENTORY, False),
        ("  Inventory Days", WC_INV_DAYS, False),
        ("Prepaid Expenses", BS_PREPAID, False),
        ("", None, False),
        ("Accounts Payable", BS_AP, False),
        ("  AP Days", WC_AP_DAYS, False),
        ("Accrued Expenses", BS_ACCRUED, False),
        ("Deferred Revenue", BS_DEFERRED_REV, False),
        ("", None, False),
        ("Net Working Capital", WC_NWC, True),
        ("Change in NWC", WC_CHG, True),
    ]
    _write_financial_sheet(ws, "Working Capital", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# D&A Schedule  (sheet 8, ~10 items)
# ---------------------------------------------------------------------------

DA_EXIST_DEPR = [12000, 11500, 11000, 10500, 10000]
DA_NEW_DEPR = [3000, 4400, 5854, 7365, 8937]
DA_TOTAL_DEPR = [DA_EXIST_DEPR[i] + DA_NEW_DEPR[i] for i in range(5)]
DA_EXIST_AMORT = [4000, 4000, 4000, 4000, 4000]
DA_NEW_AMORT = [1000, 1000, 1000, 1000, 1000]
DA_TOTAL_AMORT = [DA_EXIST_AMORT[i] + DA_NEW_AMORT[i] for i in range(5)]
DA_TOTAL = [DA_TOTAL_DEPR[i] + DA_TOTAL_AMORT[i] for i in range(5)]
DA_USEFUL_LIFE = [7, 7, 7, 7, 7]


def _build_da_schedule(ws):
    rows = [
        ("Existing Asset Depreciation", DA_EXIST_DEPR, False),
        ("New Capex Depreciation", DA_NEW_DEPR, False),
        ("Total Depreciation", DA_TOTAL_DEPR, True),
        ("", None, False),
        ("Existing Amortization", DA_EXIST_AMORT, False),
        ("New Amortization", DA_NEW_AMORT, False),
        ("Total Amortization", DA_TOTAL_AMORT, True),
        ("", None, False),
        ("Total D&A", DA_TOTAL, True),
        ("Useful Life (years)", DA_USEFUL_LIFE, False),
    ]
    _write_financial_sheet(ws, "D&A Schedule", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Tax Schedule  (sheet 9, ~9 items)
# ---------------------------------------------------------------------------

TAX_PRETAX = IS_IBT[:]
TAX_STAT_RATE = [0.25, 0.25, 0.25, 0.25, 0.25]
TAX_CURRENT = [round(IS_IBT[i] * 0.22) for i in range(5)]
TAX_DEFERRED = [IS_TAX[i] - (-TAX_CURRENT[i]) for i in range(5)]
TAX_STATE = [round(IS_IBT[i] * 0.02) for i in range(5)]
TAX_RD_CREDIT = [round(IS_IBT[i] * -0.01) for i in range(5)]
TAX_OTHER = [round(IS_IBT[i] * 0.01) for i in range(5)]
TAX_EFF_RATE = [0.25, 0.25, 0.25, 0.25, 0.25]
TAX_TOTAL = IS_TAX[:]


def _build_tax_schedule(ws):
    rows = [
        ("Pretax Income", TAX_PRETAX, True),
        ("Statutory Rate", TAX_STAT_RATE, False),
        ("Current Tax", TAX_CURRENT, False),
        ("Deferred Tax", TAX_DEFERRED, False),
        ("  State & Local", TAX_STATE, False),
        ("  R&D Credits", TAX_RD_CREDIT, False),
        ("  Other Adjustments", TAX_OTHER, False),
        ("Effective Tax Rate", TAX_EFF_RATE, False),
        ("Total Tax Expense", TAX_TOTAL, True),
    ]
    _write_financial_sheet(ws, "Tax Schedule", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Assumptions  (sheet 10, ~13 items)
# ---------------------------------------------------------------------------

ASSUM_REV_GROWTH = [None, 0.06, 0.06, 0.06, 0.06]
ASSUM_GROSS_MARGIN = [0.55, 0.55, 0.55, 0.55, 0.55]
ASSUM_OPEX_PCT = [0.27, 0.27, 0.269, 0.268, 0.268]
ASSUM_CAPEX_PCT = [0.036, 0.036, 0.036, 0.035, 0.035]
ASSUM_TAX_RATE = [0.25, 0.25, 0.25, 0.25, 0.25]
ASSUM_INT_RATE = [0.055, 0.058, 0.060, 0.058, 0.056]
ASSUM_DSO = [50, 50, 50, 50, 50]
ASSUM_DIO = [50, 50, 50, 50, 50]
ASSUM_DPO = [50, 50, 50, 50, 50]
ASSUM_DIV_PAYOUT = [0.21, 0.21, 0.21, 0.22, 0.22]


def _build_assumptions(ws):
    rows = [
        ("Revenue Growth Rate", ASSUM_REV_GROWTH, False),
        ("Gross Margin", ASSUM_GROSS_MARGIN, False),
        ("OpEx % of Revenue", ASSUM_OPEX_PCT, False),
        ("Capex % of Revenue", ASSUM_CAPEX_PCT, False),
        ("Tax Rate", ASSUM_TAX_RATE, False),
        ("Interest Rate (blended)", ASSUM_INT_RATE, False),
        ("", None, False),
        ("Working Capital Ratios:", None, True),
        ("DSO (days)", ASSUM_DSO, False),
        ("DIO (days)", ASSUM_DIO, False),
        ("DPO (days)", ASSUM_DPO, False),
        ("", None, False),
        ("Dividend Payout Ratio", ASSUM_DIV_PAYOUT, False),
    ]
    _write_financial_sheet(ws, "Assumptions", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Returns Analysis  (sheet 11, ~8 items)
# ---------------------------------------------------------------------------

RA_ENTRY_EQUITY = [BS_TOTAL_EQUITY[0]] * 5
RA_EXIT_EQUITY = BS_TOTAL_EQUITY[:]
RA_DIVS_RECEIVED = [
    sum(abs(DIVIDENDS[j]) for j in range(i + 1)) for i in range(5)
]
RA_HOLDING = [1, 2, 3, 4, 5]
RA_MOIC = [
    round((RA_EXIT_EQUITY[i] + RA_DIVS_RECEIVED[i]) / RA_ENTRY_EQUITY[i], 2)
    for i in range(5)
]
RA_ROE = [
    round(IS_NET_INCOME[i] / BS_TOTAL_EQUITY[i], 4) for i in range(5)
]
RA_ROIC = [
    round(
        IS_OPER_INCOME[i] * (1 - 0.25)
        / (BS_TOTAL_EQUITY[i] + BS_LTD[i] + BS_CURRENT_DEBT[i] - BS_CASH[i]),
        4,
    )
    for i in range(5)
]


def _build_returns_analysis(ws):
    rows = [
        ("Entry Equity", RA_ENTRY_EQUITY, False),
        ("Exit Equity", RA_EXIT_EQUITY, False),
        ("Cumulative Dividends Received", RA_DIVS_RECEIVED, False),
        ("Holding Period (years)", RA_HOLDING, False),
        ("", None, False),
        ("MOIC", RA_MOIC, True),
        ("ROE", RA_ROE, False),
        ("ROIC", RA_ROIC, False),
    ]
    _write_financial_sheet(ws, "Returns Analysis", ["", *PERIODS], rows)


# ---------------------------------------------------------------------------
# Cover Page  (sheet 12, tier 4 / skip)
# ---------------------------------------------------------------------------


def _build_cover_page(ws):
    """Cover page -- informational only, no financial data."""
    ws.column_dimensions["A"].width = 55
    items = [
        (2, "Project Alpha - Confidential Information Memorandum"),
        (4, "Prepared by: Investment Banking Division"),
        (5, "Date: March 2025"),
        (7, "DRAFT - Subject to Change"),
    ]
    for row_num, text in items:
        c = ws.cell(row=row_num, column=1, value=text)
        c.font = BOLD_FONT if row_num == 2 else NORMAL_FONT


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    """Create the large 12-sheet corporate financial model workbook."""
    wb = Workbook()

    # Sheet 1 -- Income Statement
    ws = wb.active
    ws.title = "Income Statement"
    _build_income_statement(ws)

    # Sheet 2 -- Balance Sheet
    _build_balance_sheet(wb.create_sheet("Balance Sheet"))

    # Sheet 3 -- Cash Flow Statement
    _build_cash_flow(wb.create_sheet("Cash Flow Statement"))

    # Sheet 4 -- Debt Schedule
    _build_debt_schedule(wb.create_sheet("Debt Schedule"))

    # Sheet 5 -- Revenue Build
    _build_revenue_build(wb.create_sheet("Revenue Build"))

    # Sheet 6 -- OpEx Build
    _build_opex_build(wb.create_sheet("OpEx Build"))

    # Sheet 7 -- Working Capital
    _build_working_capital(wb.create_sheet("Working Capital"))

    # Sheet 8 -- D&A Schedule
    _build_da_schedule(wb.create_sheet("D&A Schedule"))

    # Sheet 9 -- Tax Schedule
    _build_tax_schedule(wb.create_sheet("Tax Schedule"))

    # Sheet 10 -- Assumptions
    _build_assumptions(wb.create_sheet("Assumptions"))

    # Sheet 11 -- Returns Analysis
    _build_returns_analysis(wb.create_sheet("Returns Analysis"))

    # Sheet 12 -- Cover Page
    _build_cover_page(wb.create_sheet("Cover Page"))

    # Save
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    # -- Verification --
    print(f"Created: {OUTPUT}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Count total data rows across all sheets
    total_items = 0
    for sname in wb.sheetnames:
        sheet = wb[sname]
        for row in sheet.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0] and str(row[0]).strip():
                total_items += 1
    print(f"Total line items: ~{total_items}")

    # Cross-check 1: Balance Sheet balances
    print("\nBalance Sheet check (Total Assets = Total L&E):")
    all_ok = True
    for i, p in enumerate(PERIODS):
        ta = BS_TOTAL_ASSETS[i]
        tle = BS_TOTAL_LE[i]
        ok = "OK" if ta == tle else f"MISMATCH {ta} != {tle}"
        if ta != tle:
            all_ok = False
        print(f"  {p}: Total Assets = {ta:,}  Total L&E = {tle:,}  [{ok}]")

    # Cross-check 2: CF ending cash = BS cash
    print("\nCash Flow -> BS Cash check:")
    for i, p in enumerate(PERIODS):
        cf = CF_END_CASH[i]
        bs = BS_CASH[i]
        ok = "OK" if cf == bs else f"MISMATCH {cf} != {bs}"
        if cf != bs:
            all_ok = False
        print(f"  {p}: CF Ending Cash = {cf:,}  BS Cash = {bs:,}  [{ok}]")

    # Cross-check 3: Revenue Build total = IS Net Revenue
    print("\nRevenue Build -> IS check:")
    for i, p in enumerate(PERIODS):
        rb = RB_TOTAL[i]
        isr = IS_NET_REVENUE[i]
        ok = "OK" if rb == isr else f"MISMATCH {rb} != {isr}"
        if rb != isr:
            all_ok = False
        print(f"  {p}: Revenue Build = {rb:,}  IS Revenue = {isr:,}  [{ok}]")

    # Cross-check 4: OpEx Build total = IS cash opex (IS total opex minus D&A)
    print("\nOpEx Build -> IS check (cash opex, excl D&A):")
    is_cash_opex = [IS_SM[i] + IS_RD[i] + IS_GA[i] for i in range(5)]
    for i, p in enumerate(PERIODS):
        ob = OPEX_CASH_TOTAL[i]
        iso = is_cash_opex[i]
        ok = "OK" if ob == iso else f"MISMATCH {ob} != {iso}"
        if ob != iso:
            all_ok = False
        print(f"  {p}: OpEx Build = {ob:,}  IS Cash OpEx = {iso:,}  [{ok}]")

    # File size check
    size_kb = OUTPUT.stat().st_size / 1024
    size_ok = size_kb < 500
    if not size_ok:
        all_ok = False
    print(f"\nFile size: {size_kb:.1f} KB {'(OK < 500KB)' if size_ok else '(WARNING: > 500KB)'}")

    if all_ok:
        print("\nAll cross-checks PASSED.")
    else:
        print("\nWARNING: Some cross-checks FAILED.")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
