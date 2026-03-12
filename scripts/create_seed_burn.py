#!/usr/bin/env python3
"""Create tests/fixtures/seed_burn.xlsx — pre-revenue seed-stage monthly burn model."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

OUTPUT = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "seed_burn.xlsx"

MONTHS = [
    "Jan-25",
    "Feb-25",
    "Mar-25",
    "Apr-25",
    "May-25",
    "Jun-25",
    "Jul-25",
    "Aug-25",
    "Sep-25",
    "Oct-25",
    "Nov-25",
    "Dec-25",
]

# ---------------------------------------------------------------------------
# P&L data
# ---------------------------------------------------------------------------
REVENUE = [0, 0, 0, 0, 2, 5, 8, 12, 18, 25, 35, 48]
COGS = [0, 0, 0, 0, -1, -2, -3, -4, -6, -8, -11, -14]
GROSS_PROFIT = [REVENUE[i] + COGS[i] for i in range(12)]

PAYROLL = [-85, -85, -95, -95, -95, -100, -100, -100, -110, -110, -110, -120]
RENT = [-8] * 12
SOFTWARE = [-5, -5, -5, -5, -5, -6, -6, -6, -6, -6, -7, -7]
LEGAL = [-12, -3, -3, -15, -3, -3, -3, -3, -3, -3, -3, -3]
TOTAL_EXPENSES = [PAYROLL[i] + RENT[i] + SOFTWARE[i] + LEGAL[i] for i in range(12)]
NET_INCOME = [GROSS_PROFIT[i] + TOTAL_EXPENSES[i] for i in range(12)]

# ---------------------------------------------------------------------------
# Cash data
# ---------------------------------------------------------------------------
BEGINNING_CASH = [2000]
ENDING_CASH = []
for i in range(12):
    end = BEGINNING_CASH[i] + NET_INCOME[i]
    ENDING_CASH.append(end)
    if i < 11:
        BEGINNING_CASH.append(end)

RUNWAY = [round(ENDING_CASH[i] / abs(NET_INCOME[i]), 1) for i in range(12)]


def _build_sheet1_monthly_pl(wb: openpyxl.Workbook) -> None:
    """Sheet 1: Monthly P&L (tier 1, income_statement)."""
    ws = wb.create_sheet("Monthly P&L")
    bold = Font(bold=True)

    # Row 1: title
    ws["A1"] = "Seed Co - Monthly P&L"
    ws["A1"].font = Font(bold=True, size=14)

    # Row 2: unit annotation
    ws["A2"] = "(in thousands)"
    ws["A2"].font = Font(italic=True, color="888888")

    # Row 3: headers
    ws.cell(row=3, column=1, value="")
    for col_idx, month in enumerate(MONTHS, 2):
        cell = ws.cell(row=3, column=col_idx, value=month)
        cell.font = bold

    # Data rows
    rows = [
        ("Revenue", REVENUE, False),
        ("Cost of Revenue", COGS, False),
        ("Gross Profit", GROSS_PROFIT, True),
        (None, None, False),  # blank row
        ("Payroll", PAYROLL, False),
        ("Rent", RENT, False),
        ("Software", SOFTWARE, False),
        ("Legal & Professional", LEGAL, False),
        ("Total Expenses", TOTAL_EXPENSES, True),
        (None, None, False),  # blank row
        ("Net Income", NET_INCOME, True),
    ]

    row_num = 4
    for label, values, is_bold in rows:
        if label is None:
            row_num += 1
            continue
        cell = ws.cell(row=row_num, column=1, value=label)
        if is_bold:
            cell.font = bold
        for col_idx, val in enumerate(values, 2):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.number_format = "#,##0"
            if is_bold:
                c.font = bold
        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, 14):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10


def _build_sheet2_cash(wb: openpyxl.Workbook) -> None:
    """Sheet 2: Cash tracker (tier 1, cash flow / balance sheet hybrid)."""
    ws = wb.create_sheet("Cash")
    bold = Font(bold=True)

    # Row 1: title
    ws["A1"] = "Cash Tracker"
    ws["A1"].font = Font(bold=True, size=12)

    # Row 2: headers
    ws.cell(row=2, column=1, value="")
    for col_idx, month in enumerate(MONTHS, 2):
        cell = ws.cell(row=2, column=col_idx, value=month)
        cell.font = bold

    # Data rows
    rows = [
        ("Beginning Cash", BEGINNING_CASH, "#,##0", False),
        ("Revenue", REVENUE, "#,##0", False),
        ("Total Expenses", TOTAL_EXPENSES, "#,##0", False),
        ("Net Burn", NET_INCOME, "#,##0", False),
        ("Ending Cash", ENDING_CASH, "#,##0", True),
        ("Runway (months)", RUNWAY, "0.0", False),
    ]

    row_num = 3
    for label, values, fmt, is_bold in rows:
        cell = ws.cell(row=row_num, column=1, value=label)
        if is_bold:
            cell.font = bold
        for col_idx, val in enumerate(values, 2):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.number_format = fmt
            if is_bold:
                c.font = bold
        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, 14):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10


def _build_sheet3_notes(wb: openpyxl.Workbook) -> None:
    """Sheet 3: Notes (tier 4, skip)."""
    ws = wb.create_sheet("Notes")

    notes = [
        "Seed Funding: $2M received January 2025",
        "",
        "Key Assumptions:",
        "- Hiring plan: start with 6 engineers, grow to 9 by Q4",
        "- Revenue: B2B SaaS pilot with 3 initial customers Q2",
        "- Runway target: 18 months minimum",
        "",
        "Revision History:",
        "v1.0 - Board deck (Dec 2024)",
        "v1.1 - Updated for actual Jan spend (Feb 2025)",
    ]

    for i, text in enumerate(notes, start=1):
        ws.cell(row=i, column=1, value=text)

    ws.column_dimensions["A"].width = 55


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_sheet1_monthly_pl(wb)
    _build_sheet2_cash(wb)
    _build_sheet3_notes(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"Created {OUTPUT} ({OUTPUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
