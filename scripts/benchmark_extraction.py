#!/usr/bin/env python3
"""
Benchmark the DebtFund extraction pipeline with real Claude API calls.

Runs the full 5-stage extraction pipeline on a financial model Excel file
and measures timing, token usage, cost, and optionally accuracy against
expected results.

Usage:
    # Single file mode
    python scripts/benchmark_extraction.py tests/fixtures/realistic_model.xlsx
    python scripts/benchmark_extraction.py tests/fixtures/realistic_model.xlsx \\
        --expected tests/fixtures/realistic_model_expected.json --save

    # Multi-fixture mode (run all fixtures in a directory)
    python scripts/benchmark_extraction.py --fixture-dir tests/fixtures/ --save

Requires:
    ANTHROPIC_API_KEY environment variable set

Options:
    --expected FILE      Path to expected results JSON for accuracy measurement
    --save               Save detailed results to data/benchmark_results/
    --quiet              Suppress per-stage log output (show summary only)
    --fixture-dir DIR    Run all fixtures in DIR that have matching *_expected.json
    --compare [PATH]     Compare against a previous run ('latest' or file path)
"""

import argparse
import asyncio
import json
import os
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

# Ensure project root is on the path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Load .env if python-dotenv is available
try:
    from dotenv import load_dotenv

    load_dotenv(PROJECT_ROOT / ".env")
except ImportError:
    pass


def _check_api_key() -> str:
    """Validate that ANTHROPIC_API_KEY is set and looks real."""
    key = os.getenv("ANTHROPIC_API_KEY", "")
    if not key:
        print("ERROR: ANTHROPIC_API_KEY environment variable is not set.")
        print("  export ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)
    if "your-api-key" in key or not key.startswith("sk-ant-"):
        print("ERROR: ANTHROPIC_API_KEY appears to be a placeholder.")
        print(f"  Current value starts with: {key[:12]}...")
        sys.exit(1)
    return key


def _read_file(path: str) -> tuple:
    """Read file and return (bytes, size_kb, filename)."""
    file_path = Path(path)
    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)
    file_bytes = file_path.read_bytes()
    size_kb = len(file_bytes) / 1024
    return file_bytes, size_kb, file_path.name


def _count_sheets_and_rows(file_bytes: bytes) -> tuple:
    """Quick count of sheets and rows for display."""
    try:
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_count = len(wb.sheetnames)
        total_rows = 0
        for ws in wb:
            total_rows += ws.max_row or 0
        wb.close()
        return sheet_count, total_rows
    except Exception:
        return 0, 0


async def _run_extraction(file_bytes: bytes) -> dict:
    """Run the extraction pipeline with lineage DB save patched out."""
    from src.extraction.orchestrator import extract

    with patch("src.lineage.tracker.LineageTracker.save_to_db"):
        result = await extract(
            file_bytes=file_bytes,
            file_id=str(uuid.uuid4()),
        )
    return result


def _calculate_stage_metrics(result: dict) -> list:
    """Extract per-stage timing and token data from lineage summary.

    Note: The orchestrator does not currently expose per-stage timing/tokens
    in the result dict. We derive what we can from the result structure.
    The lineage summary has event counts but not detailed metrics.

    For a proper per-stage breakdown, we instrument the pipeline at runtime.
    """
    # Placeholder -- the orchestrator aggregates tokens but doesn't break
    # them down by stage in the result. Return a summary row.
    return [
        {
            "stage": "total",
            "tokens_in": 0,
            "tokens_out": 0,
            "tokens_total": result.get("tokens_used", 0),
            "cost_usd": result.get("cost_usd", 0),
        }
    ]


def _evaluate_triage_accuracy(result: dict, expected: dict) -> dict:
    """Compare triage results against expected tiers."""
    actual_triage = {t["sheet_name"]: t for t in result.get("triage", [])}
    expected_triage = expected.get("expected_triage", [])

    correct = 0
    total = len(expected_triage)
    details = []

    for exp in expected_triage:
        sheet = exp["sheet_name"]
        exp_tier = exp["tier"]
        actual = actual_triage.get(sheet, {})
        act_tier = actual.get("tier")

        is_correct = act_tier == exp_tier
        if is_correct:
            correct += 1

        details.append(
            {
                "sheet_name": sheet,
                "expected_tier": exp_tier,
                "actual_tier": act_tier,
                "correct": is_correct,
            }
        )

    return {
        "correct": correct,
        "total": total,
        "accuracy": correct / max(total, 1),
        "details": details,
    }


def _evaluate_mapping_accuracy(result: dict, expected: dict) -> dict:
    """Compare mapping results against expected canonical names."""
    # Build lookup from result line items
    actual_mappings = {}
    for item in result.get("line_items", []):
        label = item.get("original_label", "")
        canonical = item.get("canonical_name", "unmapped")
        if label and label not in actual_mappings:
            actual_mappings[label] = canonical

    expected_mappings = expected.get("expected_mappings", [])
    acceptable_alts = expected.get("acceptable_alternatives", {})

    correct = 0
    total = len(expected_mappings)
    unmapped = 0
    mismatches = []
    unmapped_labels = []

    for exp in expected_mappings:
        label = exp["original_label"]
        exp_canonical = exp["canonical_name"]
        act_canonical = actual_mappings.get(label)

        if act_canonical is None:
            # Label not found in results at all (maybe not extracted)
            unmapped += 1
            unmapped_labels.append(label)
            continue

        if act_canonical == "unmapped":
            unmapped += 1
            unmapped_labels.append(label)
            continue

        # Check if actual matches expected or an acceptable alternative
        alts = acceptable_alts.get(exp_canonical, [exp_canonical])
        if act_canonical in alts:
            correct += 1
        else:
            mismatches.append(
                {
                    "label": label,
                    "expected": exp_canonical,
                    "actual": act_canonical,
                    "acceptable": alts,
                }
            )

    matched_total = total - unmapped

    # Per-statement accuracy breakdown
    per_statement = {}
    for exp in expected_mappings:
        sheet = exp.get("sheet", "Unknown")
        per_statement.setdefault(sheet, {"correct": 0, "total": 0, "mismatches": []})
        per_statement[sheet]["total"] += 1

        label = exp["original_label"]
        exp_canonical = exp["canonical_name"]
        act_canonical = actual_mappings.get(label)

        if act_canonical is None or act_canonical == "unmapped":
            continue

        alts = acceptable_alts.get(exp_canonical, [exp_canonical])
        if act_canonical in alts:
            per_statement[sheet]["correct"] += 1
        else:
            per_statement[sheet]["mismatches"].append(
                {
                    "label": label,
                    "expected": exp_canonical,
                    "actual": act_canonical,
                }
            )

    for sheet, data in per_statement.items():
        data["accuracy"] = data["correct"] / max(data["total"], 1)

    return {
        "correct": correct,
        "total": total,
        "matched_total": matched_total,
        "accuracy": correct / max(total, 1),
        "accuracy_of_matched": correct / max(matched_total, 1),
        "unmapped_count": unmapped,
        "unmapped_labels": unmapped_labels,
        "mismatches": mismatches,
        "per_statement": per_statement,
    }


def _print_header(filename: str, size_kb: float, sheets: int, rows: int):
    """Print the benchmark header."""
    print()
    print("DebtFund Extraction Benchmark")
    print("=" * 50)
    print(f"File: {filename} ({size_kb:.0f} KB)")
    print(f"Sheets: {sheets} | Rows: ~{rows}")
    print()


def _print_result_summary(result: dict, duration: float):
    """Print extraction result summary."""
    print("Extraction Results:")
    print(f"  Duration:       {duration:.1f}s")
    print(f"  Sheets found:   {len(result.get('sheets', []))}")
    print(f"  Triage entries: {len(result.get('triage', []))}")
    print(f"  Line items:     {len(result.get('line_items', []))}")
    print(f"  Tokens used:    {result.get('tokens_used', 0):,}")
    print(f"  Cost:           ${result.get('cost_usd', 0):.4f}")

    validation = result.get("validation", {})
    if validation:
        print(f"  Validation confidence: {validation.get('overall_confidence', 0):.1%}")
        flags = validation.get("flags", [])
        if flags:
            errors = sum(1 for f in flags if f.get("severity") == "error")
            warnings = sum(1 for f in flags if f.get("severity") == "warning")
            print(f"  Validation flags: {len(flags)} ({errors} errors, {warnings} warnings)")

    completeness = validation.get("completeness", {})
    if completeness:
        print(f"  Completeness score: {completeness.get('overall_score', 0):.1%}")
        detected = completeness.get("detected_statements", [])
        if detected:
            print(f"  Detected statements: {', '.join(detected)}")
        found = completeness.get("total_found", 0)
        expected = completeness.get("total_expected", 0)
        missing = completeness.get("total_missing", 0)
        print(f"  Items: {found}/{expected} found, {missing} missing")

    quality = validation.get("quality", {})
    if quality:
        print(
            f"  Quality grade: {quality.get('letter_grade', '?')} ({quality.get('numeric_score', 0):.2f})"
        )

    lineage = result.get("lineage_summary", {})
    if lineage:
        print(f"  Lineage events: {lineage.get('total_events', 0)}")
        print(f"  Stages covered: {sorted(lineage.get('stages', []))}")

    print()


def _print_triage_details(result: dict):
    """Print triage decision details."""
    triage = result.get("triage", [])
    if not triage:
        return

    print("Triage Decisions:")
    for t in triage:
        tier = t.get("tier", "?")
        sheet = t.get("sheet_name", "Unknown")
        decision = t.get("decision", "")
        print(f"  Tier {tier}: {sheet:<30s} [{decision}]")
    print()


def _print_accuracy(triage_acc: dict, mapping_acc: dict):
    """Print accuracy comparison results."""
    print("Accuracy (vs expected):")
    print(
        f"  Triage:  {triage_acc['correct']}/{triage_acc['total']} correct "
        f"({triage_acc['accuracy']:.1%})"
    )
    print(
        f"  Mapping: {mapping_acc['correct']}/{mapping_acc['total']} correct "
        f"({mapping_acc['accuracy']:.1%})"
    )
    if mapping_acc["unmapped_count"] > 0:
        print(f"  Unmapped: {mapping_acc['unmapped_count']} items")
    print()

    # Show triage mismatches
    triage_mismatches = [d for d in triage_acc["details"] if not d["correct"]]
    if triage_mismatches:
        print("Triage Mismatches:")
        for d in triage_mismatches:
            print(
                f"  - {d['sheet_name']}: expected tier {d['expected_tier']}, "
                f"got tier {d['actual_tier']}"
            )
        print()

    # Show mapping mismatches
    if mapping_acc["mismatches"]:
        print("Mapping Mismatches:")
        for m in mapping_acc["mismatches"]:
            print(f'  - "{m["label"]}": expected {m["expected"]}, got {m["actual"]}')
        print()

    # Show unmapped labels
    if mapping_acc["unmapped_labels"]:
        print("Unmapped Labels:")
        for label in mapping_acc["unmapped_labels"]:
            print(f'  - "{label}"')
        print()

    # Per-statement breakdown
    per_stmt = mapping_acc.get("per_statement", {})
    if per_stmt:
        print("Per-Statement Mapping Accuracy:")
        for sheet, data in sorted(per_stmt.items()):
            print(f"  {sheet:<30s} {data['correct']}/{data['total']} ({data['accuracy']:.1%})")
        print()


def _save_results(
    result: dict, duration: float, filename: str, triage_acc: dict = None, mapping_acc: dict = None
):
    """Save detailed results to data/benchmark_results/."""
    results_dir = PROJECT_ROOT / "data" / "benchmark_results"
    results_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = results_dir / f"{timestamp}.json"

    output = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "file": filename,
        "duration_seconds": round(duration, 2),
        "sheets": result.get("sheets", []),
        "triage": result.get("triage", []),
        "line_items_count": len(result.get("line_items", [])),
        "tokens_used": result.get("tokens_used", 0),
        "cost_usd": result.get("cost_usd", 0),
        "validation": result.get("validation", {}),
        "lineage_summary": result.get("lineage_summary", {}),
    }

    if triage_acc:
        output["triage_accuracy"] = {
            "correct": triage_acc["correct"],
            "total": triage_acc["total"],
            "accuracy": round(triage_acc["accuracy"], 4),
            "details": triage_acc["details"],
        }

    if mapping_acc:
        output["mapping_accuracy"] = {
            "correct": mapping_acc["correct"],
            "total": mapping_acc["total"],
            "accuracy": round(mapping_acc["accuracy"], 4),
            "unmapped_count": mapping_acc["unmapped_count"],
            "unmapped_labels": mapping_acc["unmapped_labels"],
            "mismatches": mapping_acc["mismatches"],
            "per_statement": mapping_acc.get("per_statement", {}),
        }

    # Completeness tracking
    completeness = result.get("validation", {}).get("completeness", {})
    if completeness:
        output["completeness"] = completeness

    # Quality score
    quality = result.get("validation", {}).get("quality", {})
    if quality:
        output["quality"] = quality

    # Include sampled line items (first 20) for inspection
    output["sample_line_items"] = result.get("line_items", [])[:20]

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"Results saved to: {output_path}")
    return output_path


def _load_previous_result(compare_arg: str) -> dict | None:
    """Load a previous benchmark result for comparison.

    If *compare_arg* is ``"latest"``, glob ``data/benchmark_results/*.json``
    and pick the most recent file. Otherwise treat as a file path.
    """
    if compare_arg == "latest":
        results_dir = PROJECT_ROOT / "data" / "benchmark_results"
        if not results_dir.exists():
            print("No previous results found (directory missing).")
            return None
        files = sorted(results_dir.glob("*.json"))
        if not files:
            print("No previous results found (no JSON files).")
            return None
        target = files[-1]
    else:
        target = Path(compare_arg)

    if not target.exists():
        print(f"Previous result not found: {target}")
        return None

    with open(target) as f:
        data = json.load(f)
    print(f"Loaded previous result: {target.name}")
    return data


def _compare_results(current: dict, previous: dict) -> dict:
    """Compute deltas between current and previous benchmark runs."""
    diff = {}

    # Mapping accuracy delta
    curr_map = current.get("mapping_accuracy", {})
    prev_map = previous.get("mapping_accuracy", {})
    if curr_map and prev_map:
        diff["mapping_accuracy_delta"] = round(
            curr_map.get("accuracy", 0) - prev_map.get("accuracy", 0), 4
        )

        # New and resolved mismatches
        curr_mismatches = {m["label"] for m in curr_map.get("mismatches", [])}
        prev_mismatches = {m["label"] for m in prev_map.get("mismatches", [])}
        diff["new_mismatches"] = sorted(curr_mismatches - prev_mismatches)
        diff["resolved_mismatches"] = sorted(prev_mismatches - curr_mismatches)

    # Triage accuracy delta
    curr_tri = current.get("triage_accuracy", {})
    prev_tri = previous.get("triage_accuracy", {})
    if curr_tri and prev_tri:
        diff["triage_accuracy_delta"] = round(
            curr_tri.get("accuracy", 0) - prev_tri.get("accuracy", 0), 4
        )

    # Quality score delta
    curr_quality = current.get("quality", {})
    prev_quality = previous.get("quality", {})
    if curr_quality and prev_quality:
        diff["quality_score_delta"] = round(
            curr_quality.get("numeric_score", 0) - prev_quality.get("numeric_score", 0), 4
        )

    # Completeness score delta
    curr_comp = current.get("completeness", {})
    prev_comp = previous.get("completeness", {})
    if curr_comp and prev_comp:
        diff["completeness_score_delta"] = round(
            curr_comp.get("overall_score", 0) - prev_comp.get("overall_score", 0), 4
        )

    # Duration delta
    curr_dur = current.get("duration_seconds", 0)
    prev_dur = previous.get("duration_seconds", 0)
    if curr_dur and prev_dur:
        diff["duration_delta"] = round(curr_dur - prev_dur, 2)

    # Token delta
    diff["token_delta"] = current.get("tokens_used", 0) - previous.get("tokens_used", 0)

    return diff


def _print_comparison(diff: dict):
    """Print comparison deltas with colour indicators."""
    print("Comparison with Previous Run:")
    print("-" * 40)

    def _fmt(value, higher_is_better=True):
        if value > 0:
            arrow = "+" if higher_is_better else "+"
            indicator = "(better)" if higher_is_better else "(worse)"
        elif value < 0:
            arrow = ""
            indicator = "(worse)" if higher_is_better else "(better)"
        else:
            return "  no change"
        return f"  {arrow}{value:+.4f} {indicator}"

    if "mapping_accuracy_delta" in diff:
        print(f"  Mapping accuracy: {_fmt(diff['mapping_accuracy_delta'])}")
    if "triage_accuracy_delta" in diff:
        print(f"  Triage accuracy:  {_fmt(diff['triage_accuracy_delta'])}")
    if "quality_score_delta" in diff:
        print(f"  Quality score:    {_fmt(diff['quality_score_delta'])}")
    if "completeness_score_delta" in diff:
        print(f"  Completeness:     {_fmt(diff['completeness_score_delta'])}")
    if "duration_delta" in diff:
        d = diff["duration_delta"]
        print(
            f"  Duration:         {d:+.1f}s {'(faster)' if d < 0 else '(slower)' if d > 0 else ''}"
        )
    if "token_delta" in diff:
        t = diff["token_delta"]
        print(f"  Tokens:           {t:+,}")

    if diff.get("new_mismatches"):
        print(f"\n  New mismatches ({len(diff['new_mismatches'])}):")
        for label in diff["new_mismatches"]:
            print(f'    - "{label}"')
    if diff.get("resolved_mismatches"):
        print(f"\n  Resolved mismatches ({len(diff['resolved_mismatches'])}):")
        for label in diff["resolved_mismatches"]:
            print(f'    + "{label}"')

    print()


def _discover_fixtures(fixture_dir: str) -> list:
    """Discover xlsx files with matching *_expected.json in a directory.

    Returns list of (xlsx_path, expected_path) tuples sorted by filename.
    """
    dir_path = Path(fixture_dir)
    if not dir_path.is_dir():
        print(f"ERROR: Fixture directory not found: {dir_path}")
        sys.exit(1)

    pairs = []
    for xlsx_file in sorted(dir_path.glob("*.xlsx")):
        expected_file = xlsx_file.parent / f"{xlsx_file.stem}_expected.json"
        if expected_file.exists():
            pairs.append((xlsx_file, expected_file))

    return pairs


def _run_single_fixture(xlsx_path: Path, expected_path: Path, save: bool, quiet: bool) -> dict:
    """Run extraction on a single fixture and return results dict.

    Returns a dict with keys: fixture_name, duration, result, triage_acc,
    mapping_acc, tokens_used, cost_usd, line_items_count.
    """
    file_bytes, size_kb, filename = _read_file(str(xlsx_path))
    sheets, rows = _count_sheets_and_rows(file_bytes)
    _print_header(filename, size_kb, sheets, rows)

    if quiet:
        import logging

        logging.getLogger("src").setLevel(logging.WARNING)

    print("Running extraction pipeline...")
    start_time = time.time()

    try:
        result = asyncio.run(_run_extraction(file_bytes))
    except Exception as e:
        print(f"\nExtraction FAILED: {type(e).__name__}: {e}")
        return {
            "fixture_name": xlsx_path.stem,
            "duration": 0,
            "result": {},
            "triage_acc": None,
            "mapping_acc": None,
            "tokens_used": 0,
            "cost_usd": 0,
            "line_items_count": 0,
            "failed": True,
        }

    duration = time.time() - start_time
    print(f"Extraction completed in {duration:.1f}s\n")

    _print_result_summary(result, duration)
    _print_triage_details(result)

    with open(expected_path) as f:
        expected = json.load(f)

    triage_acc = _evaluate_triage_accuracy(result, expected)
    mapping_acc = _evaluate_mapping_accuracy(result, expected)
    _print_accuracy(triage_acc, mapping_acc)

    if save:
        _save_results(result, duration, filename, triage_acc, mapping_acc)

    return {
        "fixture_name": xlsx_path.stem,
        "duration": duration,
        "result": result,
        "triage_acc": triage_acc,
        "mapping_acc": mapping_acc,
        "tokens_used": result.get("tokens_used", 0),
        "cost_usd": result.get("cost_usd", 0),
        "line_items_count": len(result.get("line_items", [])),
        "failed": False,
    }


def _print_aggregate_summary(results: list):
    """Print aggregate summary across multiple fixture runs."""
    print()
    print("=" * 70)
    print("AGGREGATE BENCHMARK SUMMARY")
    print("=" * 70)

    successful = [r for r in results if not r.get("failed")]
    failed = [r for r in results if r.get("failed")]

    if failed:
        print(f"\nFailed fixtures ({len(failed)}):")
        for r in failed:
            print(f"  - {r['fixture_name']}")

    if not successful:
        print("\nNo successful fixture runs.")
        return

    # Per-fixture table
    print(
        f"\n{'Fixture':<25s} {'Triage':>8s} {'Mapping':>8s} {'Items':>6s} "
        f"{'Tokens':>8s} {'Cost':>8s} {'Time':>6s}"
    )
    print("-" * 70)

    total_mapping_correct = 0
    total_mapping_items = 0
    total_triage_correct = 0
    total_triage_items = 0
    total_tokens = 0
    total_cost = 0.0
    total_duration = 0.0

    for r in successful:
        t_acc = r.get("triage_acc") or {}
        m_acc = r.get("mapping_acc") or {}
        t_pct = f"{t_acc.get('accuracy', 0):.0%}" if t_acc else "N/A"
        m_pct = f"{m_acc.get('accuracy', 0):.0%}" if m_acc else "N/A"

        print(
            f"  {r['fixture_name']:<23s} {t_pct:>8s} {m_pct:>8s} "
            f"{r['line_items_count']:>6d} {r['tokens_used']:>8,} "
            f"${r['cost_usd']:>6.2f} {r['duration']:>5.0f}s"
        )

        total_mapping_correct += m_acc.get("correct", 0)
        total_mapping_items += m_acc.get("total", 0)
        total_triage_correct += t_acc.get("correct", 0)
        total_triage_items += t_acc.get("total", 0)
        total_tokens += r["tokens_used"]
        total_cost += r["cost_usd"]
        total_duration += r["duration"]

    print("-" * 70)

    overall_triage = total_triage_correct / max(total_triage_items, 1)
    overall_mapping = total_mapping_correct / max(total_mapping_items, 1)

    print(
        f"  {'TOTAL':<23s} {overall_triage:>7.0%} {overall_mapping:>7.0%} "
        f"{'':>6s} {total_tokens:>8,} "
        f"${total_cost:>6.2f} {total_duration:>5.0f}s"
    )
    print()
    print(
        f"Overall triage accuracy:  {total_triage_correct}/{total_triage_items} "
        f"({overall_triage:.1%})"
    )
    print(
        f"Overall mapping accuracy: {total_mapping_correct}/{total_mapping_items} "
        f"({overall_mapping:.1%})"
    )
    print(f"Total fixtures: {len(successful)} passed, {len(failed)} failed")
    print()


def main():
    parser = argparse.ArgumentParser(
        description="Benchmark DebtFund extraction pipeline with real Claude API calls."
    )
    parser.add_argument(
        "file",
        nargs="?",
        default=None,
        help="Path to Excel file to extract (e.g., tests/fixtures/realistic_model.xlsx)",
    )
    parser.add_argument(
        "--expected",
        help="Path to expected results JSON for accuracy measurement",
        default=None,
    )
    parser.add_argument(
        "--save",
        action="store_true",
        help="Save detailed results to data/benchmark_results/",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress per-stage log output",
    )
    parser.add_argument(
        "--compare",
        nargs="?",
        const="latest",
        default=None,
        help="Compare against a previous run. Use 'latest' (default) or a file path.",
    )
    parser.add_argument(
        "--fixture-dir",
        default=None,
        help="Run all fixtures in DIR that have matching *_expected.json files.",
    )
    args = parser.parse_args()

    # Validate that exactly one of file or --fixture-dir is provided
    if args.file and args.fixture_dir:
        parser.error("Cannot specify both a file and --fixture-dir.")
    if not args.file and not args.fixture_dir:
        parser.error("Must specify either a file or --fixture-dir.")

    # Validate API key
    _check_api_key()

    # Multi-fixture mode
    if args.fixture_dir:
        pairs = _discover_fixtures(args.fixture_dir)
        if not pairs:
            print(f"No fixture pairs found in {args.fixture_dir}")
            print("  (Looking for *.xlsx files with matching *_expected.json)")
            sys.exit(1)

        print(f"Found {len(pairs)} fixture(s) to benchmark:")
        for xlsx, exp in pairs:
            print(f"  {xlsx.name} + {exp.name}")
        print()

        results = []
        for xlsx_path, expected_path in pairs:
            print()
            print("*" * 70)
            print(f"FIXTURE: {xlsx_path.name}")
            print("*" * 70)
            r = _run_single_fixture(xlsx_path, expected_path, args.save, args.quiet)
            results.append(r)

        _print_aggregate_summary(results)

        failed = [r for r in results if r.get("failed")]
        if failed:
            sys.exit(1)
        return

    # Single file mode (original behavior)
    file_bytes, size_kb, filename = _read_file(args.file)
    sheets, rows = _count_sheets_and_rows(file_bytes)

    _print_header(filename, size_kb, sheets, rows)

    # Optionally suppress logging
    if args.quiet:
        import logging

        logging.getLogger("src").setLevel(logging.WARNING)

    # Run extraction
    print("Running extraction pipeline...")
    start_time = time.time()

    try:
        result = asyncio.run(_run_extraction(file_bytes))
    except Exception as e:
        print(f"\nExtraction FAILED: {type(e).__name__}: {e}")
        sys.exit(1)

    duration = time.time() - start_time
    print(f"Extraction completed in {duration:.1f}s\n")

    # Print results
    _print_result_summary(result, duration)
    _print_triage_details(result)

    # Accuracy evaluation
    triage_acc = None
    mapping_acc = None

    if args.expected:
        expected_path = Path(args.expected)
        if not expected_path.exists():
            print(f"WARNING: Expected results file not found: {expected_path}")
        else:
            with open(expected_path) as f:
                expected = json.load(f)

            triage_acc = _evaluate_triage_accuracy(result, expected)
            mapping_acc = _evaluate_mapping_accuracy(result, expected)
            _print_accuracy(triage_acc, mapping_acc)

    # Save results
    saved_path = None
    if args.save:
        saved_path = _save_results(result, duration, filename, triage_acc, mapping_acc)

    # Compare with previous run
    if args.compare:
        prev = _load_previous_result(args.compare)
        if prev and saved_path:
            # Load the just-saved result for consistent comparison
            with open(saved_path) as f:
                current_saved = json.load(f)
            diff = _compare_results(current_saved, prev)
            _print_comparison(diff)
        elif prev and mapping_acc:
            # Build a comparison dict from in-memory data
            current_dict = {
                "mapping_accuracy": mapping_acc,
                "triage_accuracy": triage_acc,
                "quality": result.get("validation", {}).get("quality", {}),
                "completeness": result.get("validation", {}).get("completeness", {}),
                "duration_seconds": duration,
                "tokens_used": result.get("tokens_used", 0),
            }
            diff = _compare_results(current_dict, prev)
            _print_comparison(diff)

    # Print canonical names found (useful for debugging)
    canonical_names = sorted(
        set(item.get("canonical_name", "unmapped") for item in result.get("line_items", []))
    )
    print(f"Canonical names extracted ({len(canonical_names)}):")
    for name in canonical_names:
        count = sum(
            1 for item in result.get("line_items", []) if item.get("canonical_name") == name
        )
        print(f"  {name}: {count} items")
    print()

    # Exit code based on basic sanity
    if not result.get("line_items"):
        print("FAIL: No line items extracted")
        sys.exit(1)
    if not result.get("triage"):
        print("FAIL: No triage results")
        sys.exit(1)

    print("Benchmark completed successfully.")


if __name__ == "__main__":
    main()
