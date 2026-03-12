"""Stage 1: Guided Parsing - Extract structured data from Excel files."""

import io
import re
import time
from typing import Any, Dict, List, Optional, Set, Tuple

import anthropic
import openpyxl
from openpyxl.utils import get_column_letter

from src.core.exceptions import (
    ClaudeAPIError,
    ExtractionError,
    InvalidFileError,
    RateLimitError,
)
from src.core.logging import extraction_logger as logger
from src.core.logging import log_exception, log_performance
from src.extraction.base import ExtractionStage, PipelineContext
from src.extraction.claude_client import get_claude_client
from src.extraction.period_parser import _DEFAULT_PARSER, check_period_consistency
from src.extraction.prompts import get_prompt
from src.extraction.utils import extract_json

# Pattern to detect subtotal/total rows from their label text
_SUBTOTAL_PATTERN = re.compile(r"\b(total|subtotal|net|gross)\b", re.IGNORECASE)

# Pattern to extract cell references from Excel formulas.
# Matches: A1, $A$1, Sheet2!B5, 'Sheet Name'!A1, A1:A10
_CELL_REF_PATTERN = re.compile(
    r"(?:"
    r"(?:'([^']+)'|([A-Za-z_]\w*))!"  # Optional sheet prefix
    r")?"
    r"\$?([A-Z]{1,3})\$?(\d+)"  # Cell reference (e.g. A1, $B$12)
    r"(?::(\$?[A-Z]{1,3})\$?(\d+))?"  # Optional range end (e.g. :A10)
)

# Pattern for detecting subtotal SUM formulas (same-column contiguous range)
_SUM_PATTERN = re.compile(r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$", re.IGNORECASE)

# ------------------------------------------------------------------
# Messy-sheet detection patterns (WS-3)
# ------------------------------------------------------------------

# Period-like header values: FY2022, CY2023, 2024E, Q1 2025, etc.
_YEAR_PATTERN = re.compile(r"^(?:FY|CY)?'?\s*(?:19|20)\d{2}\s*[EFPBA]?$", re.IGNORECASE)
_QUARTERLY_PATTERN = re.compile(r"^[QH][1-4]\s", re.IGNORECASE)

# Non-financial annotation rows
_NOTE_PATTERN = re.compile(
    r"^(source|note|disclaimer|prepared by|confidential|draft|revision"
    r"|management estimates|unaudited|as of\b|for the period|for the year"
    r"|preliminary|subject to)",
    re.IGNORECASE,
)
_SEPARATOR_PATTERN = re.compile(r"^[-=_]{3,}$")

# Unit / scale annotations
_UNIT_PATTERNS: List[Tuple[re.Pattern, str, float]] = [
    (re.compile(r"\(?\$?\s*(?:in\s+)?thousands\)?", re.I), "thousands", 1_000.0),
    (re.compile(r"\(?\$?\s*(?:in\s+)?millions?\)?", re.I), "millions", 1_000_000.0),
    (re.compile(r"\(?\$?\s*(?:in\s+)?billions?\)?", re.I), "billions", 1_000_000_000.0),
    (re.compile(r"000s", re.I), "thousands", 1_000.0),
    (re.compile(r"\(?\$?\s*(?:in\s+)?'?000\)?", re.I), "thousands", 1_000.0),
]


# ------------------------------------------------------------------
# Color & cell-type helpers
# ------------------------------------------------------------------

# Standard Office theme palette (12 colors, default theme)
_THEME_COLORS = [
    "000000",  # 0: dk1 (black)
    "FFFFFF",  # 1: lt1 (white)
    "44546A",  # 2: dk2 (dark gray-blue)
    "E7E6E6",  # 3: lt2 (light gray)
    "4472C4",  # 4: accent1 (blue)
    "ED7D31",  # 5: accent2 (orange)
    "A5A5A5",  # 6: accent3 (gray)
    "FFC000",  # 7: accent4 (gold)
    "5B9BD5",  # 8: accent5 (light blue)
    "70AD47",  # 9: accent6 (green)
    "0563C1",  # 10: hyperlink (blue)
    "954F72",  # 11: followed hyperlink (purple)
]

# Legacy indexed color palette (indices 0-7, most common)
_INDEXED_COLORS = [
    "000000",  # 0: black
    "FFFFFF",  # 1: white
    "FF0000",  # 2: red
    "00FF00",  # 3: green
    "0000FF",  # 4: blue
    "FFFF00",  # 5: yellow
    "FF00FF",  # 6: magenta
    "00FFFF",  # 7: cyan
]


def _apply_tint(hex_color: str, tint: float) -> str:
    """Apply Excel tint transformation to a 6-char hex color.

    tint < 0: darken — channel * (1 + tint)
    tint > 0: lighten — channel + (255 - channel) * tint
    """
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    def _tint_channel(c: int) -> int:
        if tint < 0:
            return max(0, int(c * (1.0 + tint)))
        return min(255, int(c + (255 - c) * tint))

    return f"{_tint_channel(r):02x}{_tint_channel(g):02x}{_tint_channel(b):02x}"


def _normalize_color(color_obj: Any) -> Optional[str]:
    """Convert an openpyxl Color object to a 6-char lowercase hex RGB string.

    Handles theme colors (with tint), indexed palette colors (0-7),
    and direct RGB/ARGB strings. Returns None for unresolvable colors.
    """
    if color_obj is None:
        return None

    color_type = getattr(color_obj, "type", None)

    # Theme colors: look up in palette and apply tint
    if color_type == "theme":
        theme_idx = getattr(color_obj, "theme", None)
        if theme_idx is not None and 0 <= theme_idx < len(_THEME_COLORS):
            base = _THEME_COLORS[theme_idx]
            tint = getattr(color_obj, "tint", 0.0) or 0.0
            if tint != 0.0:
                return _apply_tint(base, tint)
            return base.lower()
        return None

    # Indexed colors: look up common palette entries
    if color_type == "indexed":
        idx = getattr(color_obj, "indexed", None)
        if idx is not None and 0 <= idx < len(_INDEXED_COLORS):
            return _INDEXED_COLORS[idx].lower()
        return None

    # Direct RGB/ARGB strings
    rgb = getattr(color_obj, "rgb", None)
    if rgb is None:
        return None
    rgb = str(rgb)
    if rgb in ("00000000", "0", "auto", ""):
        return None
    if len(rgb) == 8:
        return rgb[2:].lower()
    if len(rgb) == 6:
        return rgb.lower()
    return None


def _extract_font_color(font: Any) -> Optional[str]:
    """Extract font color hex from an openpyxl Font object."""
    if font is None or font.color is None:
        return None
    return _normalize_color(font.color)


def _extract_fill_color(fill: Any) -> Optional[str]:
    """Extract fill/background color hex from an openpyxl PatternFill.

    Only reads fgColor from patternType fills (not 'none' or None).
    """
    if fill is None:
        return None
    pattern = getattr(fill, "patternType", None) or getattr(fill, "fill_type", None)
    if pattern is None or pattern == "none":
        return None
    fg = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
    return _normalize_color(fg)


def _is_blue_font(hex_color: Optional[str]) -> bool:
    """Check if a hex color qualifies as 'blue' for input-cell detection.

    Blue dominant: b >= 0x80, both r and g < 0x80, and blue exceeds both.
    Catches navy (000080), accent1 (4472C4), dark blue (003399), etc.
    """
    if not hex_color or len(hex_color) != 6:
        return False
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return b >= 0x80 and r < 0x80 and g < 0x80 and b > r and b > g
    except ValueError:
        return False


def _is_input_fill(hex_color: Optional[str]) -> bool:
    """Check if a fill color is light yellow or light green (input cell convention).

    Light yellow: high R, high G, low B.
    Light green: dominant G channel with R and B subdued.
    """
    if not hex_color or len(hex_color) != 6:
        return False
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        # Light yellow: high R, high G, B noticeably lower (e.g. FFFFCC)
        if r > 0xCC and g > 0xCC and b < 0xE0 and b < r and b < g:
            return True
        if g > 0xCC and r < 0xE0 and b < 0xE0 and g > r and g > b:
            return True
        return False
    except ValueError:
        return False


def _col_letter_to_index(col_letter: str) -> int:
    """Convert an Excel column letter (A, B, ..., Z, AA, ...) to 0-based index."""
    result = 0
    for ch in col_letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1  # 0-based


def _derive_cell_type(
    cell_dict: Dict[str, Any],
    col_index_0based: int,
    label_col_index: Optional[int] = None,
) -> str:
    """Derive semantic cell type from extracted cell metadata.

    Priority: formula > input > label > value.

    If *label_col_index* is given (from dynamic label-column detection), a
    string cell at or left of that column is "label".  Otherwise the legacy
    heuristic (columns A-C) is used.
    """
    if cell_dict.get("formula") is not None:
        return "formula"

    font_color = cell_dict.get("font_color")
    fill_color = cell_dict.get("fill_color")
    if _is_blue_font(font_color) or _is_input_fill(fill_color):
        return "input"

    value = cell_dict.get("value")
    if isinstance(value, str):
        if label_col_index is not None:
            if col_index_0based <= label_col_index:
                return "label"
        elif col_index_0based < 3:
            return "label"

    return "value"


class ParsingStage(ExtractionStage):
    """Stage 1: Parse Excel with Claude."""

    @property
    def name(self) -> str:
        return "parsing"

    @property
    def stage_number(self) -> int:
        return 1

    @property
    def timeout_seconds(self):
        return 120.0

    def get_timeout(self, context):
        """Scale timeout with file size: 60s base + 10s per MB."""
        file_mb = len(context.file_bytes) / (1024 * 1024)
        return max(60.0, 60.0 + file_mb * 10.0)

    @property
    def max_retries(self):
        return 3

    def validate_output(self, result):
        parsed = result.get("parsed", {})
        sheets = parsed.get("sheets", [])
        if not sheets:
            return "Parsing produced zero sheets"
        total_rows = sum(len(s.get("rows", [])) for s in sheets)
        if total_rows == 0:
            return f"Parsing found {len(sheets)} sheets but zero rows"
        return None

    async def execute(self, context: PipelineContext) -> Dict[str, Any]:
        """Parse Excel file bytes using Claude."""
        logger.info("Stage 1: Parsing started")
        start_time = time.time()

        # Pre-process Excel to structured representation
        structured = self._excel_to_structured_repr(context.file_bytes)

        # Run messy-sheet detection heuristics per sheet (pure Python, no Claude)
        for sheet in structured["sheets"]:
            sheet.update(ParsingStage._detect_sheet_metadata(sheet))
            # Reclassify cell types now that label_column is known
            ParsingStage._reclassify_cell_types(sheet)

        logger.debug(
            f"Structured extraction: {structured['sheet_count']} sheets, "
            f"{structured['total_rows']} rows"
        )

        # Detect periods deterministically from structured data
        period_parser = _DEFAULT_PARSER
        for sheet in structured["sheets"]:
            detection = period_parser.detect_periods_from_sheet(sheet)
            if detection.periods:
                sheet["detected_periods"] = detection.to_dict()

        # Check if file is too large for single-pass processing
        if self._should_chunk(structured):
            logger.info(
                f"Stage 1: Large file detected "
                f"({self._estimate_token_count(structured)} est. tokens). "
                f"Chunking into {len(structured['sheets'])} per-sheet calls."
            )
            return await self._execute_chunked(context, structured, start_time)

        # Convert to token-efficient markdown for Claude
        excel_text = self._structured_to_markdown(structured)
        logger.debug(f"Converted structured repr to markdown ({len(excel_text)} chars)")

        prompt_text = get_prompt("parsing").content
        full_prompt = (
            f"Below is the content of an Excel file, with each sheet shown separately.\n\n"
            f"{excel_text}\n\n"
            f"{prompt_text}"
        )

        content, input_tokens, output_tokens = self._call_claude(full_prompt)
        parsed = extract_json(content)

        # Enrich parsed rows with source cell provenance from structured repr
        if isinstance(parsed, dict):
            self._enrich_with_source_cells(parsed, structured)

        tokens = input_tokens + output_tokens
        duration = time.time() - start_time

        sheets_count = len(parsed.get("sheets", []) if isinstance(parsed, dict) else [])  # type: ignore[union-attr]

        log_performance(
            logger,
            "stage_1_parsing",
            duration,
            {"tokens": tokens, "sheets": sheets_count},
        )
        logger.info(f"Stage 1: Parsing completed - {sheets_count} sheets found")

        # Collect detected periods from all sheets
        all_detected_periods = {
            s["sheet_name"]: s["detected_periods"]
            for s in structured["sheets"]
            if "detected_periods" in s
        }

        try:
            period_warnings = check_period_consistency(all_detected_periods)
        except Exception:
            period_warnings = []

        return {
            "parsed": parsed,
            "tokens": tokens,
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "lineage_metadata": {
                "sheets_count": sheets_count,
                "file_size_bytes": len(context.file_bytes),
            },
            "structured": structured,
            "detected_periods": all_detected_periods,
            "period_warnings": period_warnings,
        }

    async def _execute_chunked(
        self,
        context: PipelineContext,
        structured: Dict[str, Any],
        start_time: float,
    ) -> Dict[str, Any]:
        """Execute parsing by calling Claude per-sheet and merging results."""
        sheet_results: List[Dict[str, Any]] = []
        total_input_tokens = 0
        total_output_tokens = 0

        prompt_text = get_prompt("parsing").content

        for sheet in structured["sheets"]:
            single = self._make_single_sheet_structured(structured, sheet)
            excel_text = self._structured_to_markdown(single)

            full_prompt = (
                f"Below is the content of an Excel file, with each sheet shown separately.\n\n"
                f"{excel_text}\n\n"
                f"{prompt_text}"
            )

            content, inp_tok, out_tok = self._call_claude(full_prompt)
            parsed_sheet = extract_json(content)
            sheet_results.append(parsed_sheet)  # type: ignore[arg-type]
            total_input_tokens += inp_tok
            total_output_tokens += out_tok

        parsed = self._merge_parsed_sheets(sheet_results)

        # Enrich parsed rows with source cell provenance from structured repr
        self._enrich_with_source_cells(parsed, structured)

        tokens = total_input_tokens + total_output_tokens
        duration = time.time() - start_time
        sheets_count = len(parsed.get("sheets", []))

        log_performance(
            logger,
            "stage_1_parsing_chunked",
            duration,
            {"tokens": tokens, "sheets": sheets_count, "chunks": len(sheet_results)},
        )
        logger.info(
            f"Stage 1: Chunked parsing completed - {sheets_count} sheets "
            f"from {len(sheet_results)} chunks"
        )

        # Collect detected periods from all sheets
        all_detected_periods = {
            s["sheet_name"]: s["detected_periods"]
            for s in structured["sheets"]
            if "detected_periods" in s
        }

        try:
            period_warnings = check_period_consistency(all_detected_periods)
        except Exception:
            period_warnings = []

        return {
            "parsed": parsed,
            "tokens": tokens,
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "lineage_metadata": {
                "sheets_count": sheets_count,
                "file_size_bytes": len(context.file_bytes),
                "chunked": True,
            },
            "structured": structured,
            "detected_periods": all_detected_periods,
            "period_warnings": period_warnings,
        }

    def _call_claude(self, prompt: str) -> Tuple[str, int, int]:
        """
        Call Claude API with error handling.

        Returns (response_text, input_tokens, output_tokens).
        """
        try:
            with get_claude_client().messages.stream(
                model="claude-sonnet-4-20250514",
                max_tokens=32768,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                response = stream.get_final_message()

            if response.stop_reason == "max_tokens":
                logger.warning(
                    f"Stage 1: Response truncated at max_tokens "
                    f"({response.usage.output_tokens} tokens)."
                )
                raise ExtractionError(
                    "Parsing response truncated: output exceeded token limit. "
                    "The financial model may be too large for single-pass parsing.",
                    stage="parsing",
                )

            content = response.content[0].text  # type: ignore[union-attr]
            return content, response.usage.input_tokens, response.usage.output_tokens

        except anthropic.RateLimitError as e:
            retry_after = getattr(e.response, "headers", {}).get("retry-after")
            logger.warning(f"Stage 1: Rate limit hit (retry-after={retry_after})")
            raise RateLimitError(
                "Rate limit exceeded",
                stage="parsing",
                retry_after=int(retry_after) if retry_after else None,
            )

        except anthropic.APIError as e:
            logger.error(f"Stage 1: Claude API error - {str(e)}")
            error = ClaudeAPIError(
                str(e),
                stage="parsing",
                status_code=getattr(e, "status_code", None),
            )
            log_exception(logger, error)
            raise error

        except ExtractionError:
            raise

        except Exception as e:
            logger.error(f"Stage 1: Unexpected error - {str(e)}")
            error = ExtractionError(f"Parsing failed: {str(e)}", stage="parsing")  # type: ignore[assignment]
            log_exception(logger, error)
            raise error

    # ------------------------------------------------------------------
    # Chunking support for large files
    # ------------------------------------------------------------------

    @staticmethod
    def _estimate_token_count(structured: Dict[str, Any]) -> int:
        """Estimate token count from structured repr using char/4 heuristic."""
        md = ParsingStage._structured_to_markdown(structured)
        return len(md) // 4

    @staticmethod
    def _should_chunk(
        structured: Dict[str, Any],
        threshold: int = 50_000,
    ) -> bool:
        """Return True if estimated token count exceeds threshold."""
        return ParsingStage._estimate_token_count(structured) > threshold

    @staticmethod
    def _make_single_sheet_structured(
        structured: Dict[str, Any],
        sheet: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Create a structured repr containing only one sheet."""
        return {
            "sheets": [sheet],
            "named_ranges": structured.get("named_ranges", {}),
            "sheet_count": 1,
            "total_rows": len(sheet.get("rows", [])),
        }

    @staticmethod
    def _merge_parsed_sheets(
        sheet_results: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """Merge per-sheet Claude parsing responses into a single result."""
        merged_sheets: List[Dict[str, Any]] = []
        for result in sheet_results:
            if isinstance(result, dict):
                merged_sheets.extend(result.get("sheets", []))
            elif isinstance(result, list):
                merged_sheets.extend(result)
        return {"sheets": merged_sheets}

    # ------------------------------------------------------------------
    # Provenance enrichment
    # ------------------------------------------------------------------

    @staticmethod
    def _enrich_with_source_cells(
        parsed: Dict[str, Any],
        structured: Dict[str, Any],
    ) -> None:
        """Enrich parsed rows with source cell provenance from structured repr.

        For each parsed row, cross-references the structured Excel extraction
        to find the original cells that contributed to the label and values.
        Adds ``source_cells`` and ``parsing_metadata`` to each row in-place.
        """
        # Build lookup: {sheet_name: {row_index: cells_list}}
        struct_lookup: Dict[str, Dict[int, List[Dict[str, Any]]]] = {}
        for sheet in structured.get("sheets", []):
            name = sheet.get("sheet_name", "")
            row_map: Dict[int, List[Dict[str, Any]]] = {}
            for row in sheet.get("rows", []):
                row_map[row["row_index"]] = row.get("cells", [])
            struct_lookup[name] = row_map

        for sheet in parsed.get("sheets", []):
            sheet_name = sheet.get("sheet_name", "")
            sheet_rows = struct_lookup.get(sheet_name, {})

            for row in sheet.get("rows", []):
                row_index = row.get("row_index")
                struct_cells = sheet_rows.get(row_index, [])

                source_cells: List[Dict[str, Any]] = []
                label_is_bold = False

                if struct_cells:
                    cell_ref = row.get("cell_ref")

                    # Label cell
                    for cell in struct_cells:
                        if cell.get("ref") == cell_ref:
                            source_cells.append(
                                {
                                    "sheet": sheet_name,
                                    "cell_ref": cell["ref"],
                                    "raw_value": cell.get("value"),
                                }
                            )
                            label_is_bold = cell.get("is_bold", False)
                            break

                    # Value cells — match parsed values to structured cells
                    values = row.get("values", {})
                    if values:
                        consumed: set = set()
                        # Get numeric cells excluding the label cell
                        numeric_cells = [
                            c
                            for c in struct_cells
                            if c.get("ref") != cell_ref
                            and c.get("value") is not None
                            and isinstance(c.get("value"), (int, float))
                        ]
                        for _period, val in values.items():
                            if val is None:
                                continue
                            for nc in numeric_cells:
                                if nc["ref"] in consumed:
                                    continue
                                nc_val = nc.get("value")
                                try:
                                    if abs(float(nc_val) - float(val)) < 0.01:  # type: ignore[arg-type]
                                        entry: Dict[str, Any] = {
                                            "sheet": sheet_name,
                                            "cell_ref": nc["ref"],
                                            "raw_value": nc_val,
                                        }
                                        if nc.get("formula"):
                                            entry["formula"] = nc["formula"]
                                        source_cells.append(entry)
                                        consumed.add(nc["ref"])
                                        break
                                except (TypeError, ValueError):
                                    continue

                row["source_cells"] = source_cells
                row["parsing_metadata"] = {
                    "hierarchy_level": row.get("hierarchy_level", 1),
                    "is_bold": label_is_bold,
                    "is_formula": row.get("is_formula", False),
                    "is_subtotal": row.get("is_subtotal", False),
                }

    # ------------------------------------------------------------------
    # Structured Excel extraction
    # ------------------------------------------------------------------

    @staticmethod
    def _excel_to_structured_repr(file_bytes: bytes) -> Dict[str, Any]:
        """
        Extract a rich structured representation from an Excel file.

        Opens the workbook **twice**:
          1. ``data_only=False`` (default) to capture formulas and formatting.
          2. ``data_only=True`` to capture computed/cached values.

        The two passes are merged so every cell carries both its computed
        *value* and its *formula* string (if any).

        Returns a dict with keys: sheets, named_ranges, sheet_count, total_rows.
        Raises InvalidFileError for password-protected or corrupted files.
        """
        try:
            # Pass 1 – formulas + formatting (read_only=False required for styles)
            wb_formulas = openpyxl.load_workbook(
                io.BytesIO(file_bytes), data_only=False, read_only=False
            )
        except Exception as e:
            raise InvalidFileError(
                f"Cannot open Excel file: {e}",
                file_type="xlsx",
            )

        try:
            # Pass 2 – computed/cached values
            wb_values = openpyxl.load_workbook(
                io.BytesIO(file_bytes), data_only=True, read_only=True
            )
        except Exception:
            # If the second pass fails, close the first workbook and raise
            wb_formulas.close()
            raise InvalidFileError(
                "Cannot open Excel file for value extraction",
                file_type="xlsx",
            )

        try:
            sheets: List[Dict[str, Any]] = []
            total_rows = 0

            for sheet_name in wb_formulas.sheetnames:
                ws_fmt = wb_formulas[sheet_name]
                ws_val = wb_values[sheet_name]

                # Build a lookup of computed values by (row, col) from the
                # values-only workbook
                value_lookup: Dict[tuple, Any] = {}
                for row in ws_val.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            value_lookup[(cell.row, cell.column)] = cell.value

                # Detect merged cell regions
                merged_regions = [str(m) for m in ws_fmt.merged_cells.ranges]

                sheet_visibility = ws_fmt.sheet_state  # 'visible', 'hidden', or 'veryHidden'
                is_hidden = sheet_visibility != "visible"

                row_dicts: List[Dict[str, Any]] = []
                for row_cells in ws_fmt.iter_rows():
                    cells_out: List[Dict[str, Any]] = []
                    row_empty = True

                    for cell in row_cells:
                        raw = cell.value
                        computed = value_lookup.get((cell.row, cell.column), raw)

                        # Determine formula
                        formula: Optional[str] = None
                        if isinstance(raw, str) and raw.startswith("="):
                            formula = raw
                            # If there is a cached computed value use it;
                            # otherwise fall back to None (formula not evaluated)
                            value = computed if computed != raw else None
                        else:
                            value = raw

                        # Skip truly empty cells
                        if value is None and formula is None:
                            continue

                        row_empty = False

                        # Formatting clues
                        font = cell.font
                        alignment = cell.alignment

                        is_bold = bool(font and font.bold)
                        is_underline = bool(font and font.underline)
                        indent_level = int(
                            alignment.indent if alignment and alignment.indent else 0
                        )
                        number_format = cell.number_format or "General"

                        # Rich formatting (WS-1)
                        font_color = _extract_font_color(font)
                        fill_color = _extract_fill_color(cell.fill)
                        border = cell.border
                        has_border_bottom = bool(
                            border and border.bottom and border.bottom.style is not None
                        )
                        has_border_right = bool(
                            border and border.right and border.right.style is not None
                        )
                        comment_text = (
                            str(cell.comment.text).strip()
                            if cell.comment and cell.comment.text
                            else None
                        )

                        col_letter = get_column_letter(cell.column)
                        ref = f"{col_letter}{cell.row}"

                        cell_dict: Dict[str, Any] = {
                            "ref": ref,
                            "value": value,
                            "formula": formula,
                            "is_bold": is_bold,
                            "is_underline": is_underline,
                            "indent_level": indent_level,
                            "number_format": number_format,
                            "font_color": font_color,
                            "fill_color": fill_color,
                            "has_border_bottom": has_border_bottom,
                            "has_border_right": has_border_right,
                        }

                        if comment_text:
                            cell_dict["comment"] = comment_text

                        # Label-based subtotal detection
                        if isinstance(value, str) and _SUBTOTAL_PATTERN.search(value):
                            cell_dict["is_subtotal"] = True

                        # Extract formula cell references for dependency tracking
                        if formula:
                            references = ParsingStage._extract_cell_references(formula)
                            if references:
                                cell_dict["references"] = references

                        # Derive semantic cell type
                        cell_dict["cell_type"] = _derive_cell_type(cell_dict, cell.column - 1)

                        cells_out.append(cell_dict)

                    if row_empty or not cells_out:
                        continue  # skip entirely empty rows

                    row_index = row_cells[0].row
                    row_dict: Dict[str, Any] = {
                        "row_index": row_index,
                        "cells": cells_out,
                    }

                    # Mark row-level subtotal if the first cell is a subtotal
                    if cells_out and cells_out[0].get("is_subtotal"):
                        row_dict["is_subtotal"] = True

                    row_dicts.append(row_dict)

                # Propagate merged cell values to all cells in merged regions
                row_dicts = ParsingStage._propagate_merged_cells(row_dicts, ws_fmt, value_lookup)

                # Detect section boundaries and build formula graph summary
                section_boundaries = ParsingStage._detect_section_boundaries(row_dicts)
                formula_graph_summary = ParsingStage._build_formula_graph_summary(
                    row_dicts, sheet_name
                )

                total_rows += len(row_dicts)

                sheets.append(
                    {
                        "sheet_name": sheet_name,
                        "is_hidden": is_hidden,
                        "visibility": sheet_visibility,
                        "merged_regions": merged_regions,
                        "section_boundaries": section_boundaries,
                        "formula_graph_summary": formula_graph_summary,
                        "rows": row_dicts,
                        "max_row": ws_fmt.max_row or 0,
                    }
                )

            # Collect named ranges (openpyxl 3.1+ uses a dict-like API)
            named_ranges: Dict[str, str] = {}
            dn = wb_formulas.defined_names
            # Support both openpyxl < 3.1 (.definedName) and >= 3.1 (.values())
            defn_iter = getattr(dn, "definedName", None) or dn.values()
            for defn in defn_iter:
                try:
                    destinations = list(defn.destinations)
                    if destinations:
                        ws_title, coord = destinations[0]
                        named_ranges[defn.name] = f"{ws_title}!{coord}"
                except Exception:
                    # Some defined names are malformed or refer to external
                    # sheets; skip gracefully.
                    pass

            return {
                "sheets": sheets,
                "named_ranges": named_ranges,
                "sheet_count": len(sheets),
                "total_rows": total_rows,
            }

        finally:
            wb_formulas.close()
            wb_values.close()

    # ------------------------------------------------------------------
    # Formula reference extraction
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_cell_references(formula: str) -> List[str]:
        """
        Extract cell references from an Excel formula.

        Handles simple refs (A1), absolute refs ($A$1), ranges (A1:A10),
        cross-sheet refs (Sheet2!B5, 'Sheet Name'!A1).

        Returns a deduplicated list of reference strings.
        """
        if not formula or not formula.startswith("="):
            return []

        refs: List[str] = []
        for match in _CELL_REF_PATTERN.finditer(formula):
            sheet_quoted = match.group(1)
            sheet_plain = match.group(2)
            col1 = match.group(3)
            row1 = match.group(4)
            col2 = match.group(5)
            row2 = match.group(6)

            sheet_prefix = ""
            if sheet_quoted:
                sheet_prefix = f"'{sheet_quoted}'!"
            elif sheet_plain:
                sheet_prefix = f"{sheet_plain}!"

            refs.append(f"{sheet_prefix}{col1}{row1}")

            if col2 and row2:
                col2_clean = col2.lstrip("$")
                refs.append(f"{sheet_prefix}{col2_clean}{row2}")

        # Deduplicate while preserving order
        seen: set = set()
        unique: List[str] = []
        for ref in refs:
            if ref not in seen:
                seen.add(ref)
                unique.append(ref)
        return unique

    # ------------------------------------------------------------------
    # Merged cell propagation
    # ------------------------------------------------------------------

    @staticmethod
    def _propagate_merged_cells(
        row_dicts: List[Dict[str, Any]],
        ws_fmt: Any,
        value_lookup: Dict[tuple, Any],
    ) -> List[Dict[str, Any]]:
        """
        Propagate merged cell values to all cells in merged regions.

        openpyxl makes non-top-left cells in a merged region return None,
        so they get skipped during the main extraction loop. This method
        copies the top-left cell's value and formatting to every cell in
        the region and adds ``is_merged=True`` to each.
        """
        if not ws_fmt.merged_cells.ranges:
            return row_dicts

        row_lookup: Dict[int, Dict[str, Any]] = {r["row_index"]: r for r in row_dicts}

        for merged_range in ws_fmt.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col

            # Get top-left cell data
            tl_cell = ws_fmt.cell(row=min_row, column=min_col)
            tl_raw = tl_cell.value
            tl_formula: Optional[str] = None
            if isinstance(tl_raw, str) and tl_raw.startswith("="):
                tl_formula = tl_raw
                tl_value = value_lookup.get((min_row, min_col))
            else:
                tl_value = tl_raw

            tl_font = tl_cell.font
            tl_alignment = tl_cell.alignment
            tl_is_bold = bool(tl_font and tl_font.bold)
            tl_is_underline = bool(tl_font and tl_font.underline)
            tl_indent = int(tl_alignment.indent if tl_alignment and tl_alignment.indent else 0)
            tl_number_format = tl_cell.number_format or "General"

            # Rich formatting from top-left cell
            tl_ref = f"{get_column_letter(min_col)}{min_row}"
            tl_font_color = _extract_font_color(tl_font)
            tl_fill_color = _extract_fill_color(tl_cell.fill)
            tl_border = tl_cell.border
            tl_has_border_bottom = bool(
                tl_border and tl_border.bottom and tl_border.bottom.style is not None
            )
            tl_has_border_right = bool(
                tl_border and tl_border.right and tl_border.right.style is not None
            )

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    col_letter = get_column_letter(c)
                    ref = f"{col_letter}{r}"

                    cell_dict: Dict[str, Any] = {
                        "ref": ref,
                        "value": tl_value,
                        "formula": tl_formula,
                        "is_bold": tl_is_bold,
                        "is_underline": tl_is_underline,
                        "indent_level": tl_indent,
                        "number_format": tl_number_format,
                        "is_merged": True,
                        "merge_origin": tl_ref,
                        "font_color": tl_font_color,
                        "fill_color": tl_fill_color,
                        "has_border_bottom": tl_has_border_bottom,
                        "has_border_right": tl_has_border_right,
                    }

                    # Derive cell type for propagated cell
                    cell_dict["cell_type"] = _derive_cell_type(cell_dict, c - 1)

                    if isinstance(tl_value, str) and _SUBTOTAL_PATTERN.search(tl_value):
                        cell_dict["is_subtotal"] = True

                    if r not in row_lookup:
                        new_row: Dict[str, Any] = {
                            "row_index": r,
                            "cells": [cell_dict],
                        }
                        if cell_dict.get("is_subtotal"):
                            new_row["is_subtotal"] = True
                        row_dicts.append(new_row)
                        row_lookup[r] = new_row
                    else:
                        existing = next(
                            (cell for cell in row_lookup[r]["cells"] if cell["ref"] == ref),
                            None,
                        )
                        if existing:
                            existing["is_merged"] = True
                            existing["merge_origin"] = tl_ref
                        else:
                            row_lookup[r]["cells"].append(cell_dict)

        # Re-sort rows and cells after possible insertions
        row_dicts.sort(key=lambda r: r["row_index"])
        for row_dict in row_dicts:
            row_dict["cells"].sort(
                key=lambda c: (
                    len(c["ref"].rstrip("0123456789")),
                    c["ref"].rstrip("0123456789"),
                )
            )

        return row_dicts

    # ------------------------------------------------------------------
    # Section boundary detection
    # ------------------------------------------------------------------

    _MAX_SECTION_BOUNDARIES = 20

    @staticmethod
    def _detect_section_boundaries(
        row_dicts: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Detect section header rows based on formatting cues.

        A row is a section boundary if its first cell is bold or underlined
        with a string value AND it is the first row, preceded by a gap in
        ``row_index``, or the first cell has ``has_border_bottom=True``.

        Returns at most ``_MAX_SECTION_BOUNDARIES`` entries. If truncated,
        the last entry gets ``"truncated": True``.
        """
        boundaries: List[Dict[str, Any]] = []
        prev_row_index = 0

        for row in row_dicts:
            cells = row.get("cells", [])
            if not cells:
                prev_row_index = row["row_index"]
                continue

            first_cell = cells[0]
            is_header_style = first_cell.get("is_bold") or first_cell.get("is_underline")
            if not is_header_style:
                prev_row_index = row["row_index"]
                continue

            value = first_cell.get("value")
            if not isinstance(value, str) or not value.strip():
                prev_row_index = row["row_index"]
                continue

            is_first_row = prev_row_index == 0
            preceded_by_gap = row["row_index"] - prev_row_index > 1
            has_border = first_cell.get("has_border_bottom", False)

            if is_first_row or preceded_by_gap or has_border:
                boundaries.append(
                    {
                        "row_index": row["row_index"],
                        "label": value.strip(),
                    }
                )

            prev_row_index = row["row_index"]

        # Cap at max to avoid cluttering markdown output
        if len(boundaries) > ParsingStage._MAX_SECTION_BOUNDARIES:
            boundaries = boundaries[: ParsingStage._MAX_SECTION_BOUNDARIES]
            if boundaries:
                boundaries[-1]["truncated"] = True

        return boundaries

    # ------------------------------------------------------------------
    # Cell type reclassification (post label_column detection)
    # ------------------------------------------------------------------

    @staticmethod
    def _reclassify_cell_types(sheet: Dict[str, Any]) -> None:
        """Re-derive cell_type using the detected label column.

        Called after ``_detect_sheet_metadata`` populates ``label_column``.
        If labels live beyond column C (e.g. column D or E), string cells
        at or left of that column were initially classified as "value" —
        this re-derives them using ``_derive_cell_type`` with the correct
        label column index.
        """
        label_col_letter = sheet.get("label_column")
        if not label_col_letter or label_col_letter in ("A", "B", "C"):
            return  # default heuristic already covers columns A-C

        label_col_index = _col_letter_to_index(label_col_letter)

        for row in sheet.get("rows", []):
            for cell in row.get("cells", []):
                ref = cell.get("ref", "")
                col_match = re.match(r"([A-Z]+)", ref)
                if not col_match:
                    continue
                col_idx = _col_letter_to_index(col_match.group(1))
                cell["cell_type"] = _derive_cell_type(cell, col_idx, label_col_index)

    # ------------------------------------------------------------------
    # Messy-sheet metadata detection heuristics (WS-3)
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_label_column(rows: List[Dict[str, Any]]) -> Optional[str]:
        """Detect which column contains row labels (most string cells).

        Returns a column letter (e.g. "A", "B") or None if uncertain.
        Prefers the leftmost column on ties.
        """
        if not rows:
            return None

        string_counts: Dict[str, int] = {}
        total_counts: Dict[str, int] = {}

        for row in rows:
            for cell in row.get("cells", []):
                ref = cell.get("ref", "")
                m = re.match(r"([A-Z]+)", ref)
                if not m:
                    continue
                col = m.group(1)
                val = cell.get("value")
                if val is None:
                    continue
                total_counts[col] = total_counts.get(col, 0) + 1
                if isinstance(val, str) and val.strip():
                    string_counts[col] = string_counts.get(col, 0) + 1

        if not string_counts:
            return None

        # Exclude columns where >50% of non-empty cells are numeric
        candidates: Dict[str, int] = {}
        for col, count in string_counts.items():
            total = total_counts.get(col, 0)
            if total > 0 and count / total >= 0.5:
                candidates[col] = count

        if not candidates:
            return None

        # Require at least 3 string cells
        candidates = {c: n for c, n in candidates.items() if n >= 3}
        if not candidates:
            return None

        # Return leftmost column with highest string count
        best = max(
            candidates.items(),
            key=lambda x: (x[1], -len(x[0]), [-ord(c) for c in x[0]]),
        )
        return best[0]

    @staticmethod
    def _detect_header_row(rows: List[Dict[str, Any]]) -> Optional[int]:
        """Detect the row containing period headers (FY2022, 2023, dates, etc.).

        Scans the first 20 rows. Returns the row_index with the most
        period-like values (minimum 2 matches), or None.
        """
        from datetime import datetime

        if not rows:
            return None

        best_row: Optional[int] = None
        best_count = 0

        for row in rows[:20]:
            period_count = 0
            for cell in row.get("cells", []):
                val = cell.get("value")
                if val is None:
                    continue
                if isinstance(val, datetime):
                    period_count += 1
                    continue
                s = str(val).strip()
                if _YEAR_PATTERN.match(s) or _QUARTERLY_PATTERN.match(s):
                    period_count += 1

            if period_count > best_count:
                best_count = period_count
                best_row = row["row_index"]

        return best_row if best_count >= 2 else None

    @staticmethod
    def _detect_table_regions(
        rows: List[Dict[str, Any]],
    ) -> List[Dict[str, int]]:
        """Detect contiguous data regions separated by 2+ blank rows.

        Returns a list of {start_row, end_row} dicts. If no gaps are
        found, returns a single region spanning all rows.
        """
        if not rows:
            return []

        indices = sorted(r["row_index"] for r in rows)
        if not indices:
            return []

        regions: List[Dict[str, int]] = []
        region_start = indices[0]

        for i in range(1, len(indices)):
            gap = indices[i] - indices[i - 1]
            if gap >= 3:  # 2+ blank rows between data rows
                regions.append(
                    {
                        "start_row": region_start,
                        "end_row": indices[i - 1],
                    }
                )
                region_start = indices[i]

        # Close the last region
        regions.append(
            {
                "start_row": region_start,
                "end_row": indices[-1],
            }
        )

        return regions

    # Additional patterns for transposition detection (broader than header patterns)
    _TRANSPOSED_PERIOD_RE = re.compile(
        r"^(?:"
        r"(?:FY|CY)'?\s*\d{2,4}(?:/\d{2,4})?"  # FY2024, FY26/27
        r"|(?:19|20)\d{2}\s*[EFPBA]?"  # 2024, 2024E
        r"|[QH][1-4]\s"  # Q1 2025, H1 2024
        r"|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*[\s/-]"  # Jan-24
        r"|Year\s+\d"  # Year 1, Year 2
        r")\s*",
        re.IGNORECASE,
    )

    @staticmethod
    def _detect_transposed(
        rows: List[Dict[str, Any]],
        label_column: Optional[str],
    ) -> bool:
        """Detect transposed layouts where periods run down a column.

        Returns True if >60% of column A (or label_column) values match
        period patterns (minimum 3 matches), OR if the first data row has
        period-like values going across AND column A has sequential years
        going down. Conservative default: False.
        """
        from datetime import datetime

        if not rows:
            return False

        target_col = label_column or "A"
        period_matches = 0
        total_values = 0

        for row in rows:
            for cell in row.get("cells", []):
                ref = cell.get("ref", "")
                m = re.match(r"([A-Z]+)", ref)
                if not m or m.group(1) != target_col:
                    continue
                val = cell.get("value")
                if val is None:
                    continue
                total_values += 1
                if isinstance(val, datetime):
                    period_matches += 1
                elif isinstance(val, (int, float)):
                    # Bare numeric years (2024, 2025.0)
                    n = int(val) if isinstance(val, float) and val == int(val) else val
                    if isinstance(n, int) and 1990 <= n <= 2100:
                        period_matches += 1
                else:
                    s = str(val).strip()
                    if ParsingStage._TRANSPOSED_PERIOD_RE.match(s):
                        period_matches += 1
                break  # only check one cell per row for the target column

        if total_values < 3 or period_matches < 3:
            return False

        return period_matches / total_values > 0.6

    @staticmethod
    def _detect_non_financial_rows(
        rows: List[Dict[str, Any]],
    ) -> Set[int]:
        """Detect rows containing notes, disclaimers, or separators.

        Returns a set of row_index values for non-financial rows.
        """
        non_financial: Set[int] = set()

        for row in rows:
            cells = row.get("cells", [])
            if not cells:
                continue

            # Find the first text cell
            first_text = None
            for cell in cells:
                val = cell.get("value")
                if isinstance(val, str) and val.strip():
                    first_text = val.strip()
                    break

            if not first_text:
                continue

            if _NOTE_PATTERN.match(first_text):
                non_financial.add(row["row_index"])
            elif _SEPARATOR_PATTERN.match(first_text):
                non_financial.add(row["row_index"])
            elif len(cells) == 1 and len(first_text) < 5 and not first_text.isalnum():
                non_financial.add(row["row_index"])

        return non_financial

    @staticmethod
    def _detect_unit_hint(
        rows: List[Dict[str, Any]],
        sheet_name: str,
    ) -> Tuple[Optional[str], Optional[float]]:
        """Detect unit/scale annotations like '(in thousands)'.

        Checks the first 10 rows and the sheet name for unit patterns.
        Returns (unit_hint, multiplier) or (None, None).
        """
        # Check sheet name first
        for pattern, hint, multiplier in _UNIT_PATTERNS:
            if pattern.search(sheet_name):
                return hint, multiplier

        # Check first 10 rows
        for row in rows[:10]:
            for cell in row.get("cells", []):
                val = cell.get("value")
                if not isinstance(val, str):
                    continue
                for pattern, hint, multiplier in _UNIT_PATTERNS:
                    if pattern.search(val):
                        return hint, multiplier

        return None, None

    @staticmethod
    def _detect_sheet_metadata(sheet: Dict[str, Any]) -> Dict[str, Any]:
        """Run all messy-sheet detection heuristics on a single sheet.

        Returns a dict of metadata keys to merge into the sheet dict.
        """
        rows = sheet.get("rows", [])
        sheet_name = sheet.get("sheet_name", "")

        label_column = ParsingStage._detect_label_column(rows)
        header_row_index = ParsingStage._detect_header_row(rows)
        table_regions = ParsingStage._detect_table_regions(rows)
        is_transposed = ParsingStage._detect_transposed(rows, label_column)
        non_financial_rows = ParsingStage._detect_non_financial_rows(rows)
        unit_hint, unit_multiplier = ParsingStage._detect_unit_hint(
            rows,
            sheet_name,
        )

        return {
            "label_column": label_column,
            "header_row_index": header_row_index,
            "table_regions": table_regions,
            "is_transposed": is_transposed,
            "non_financial_rows": non_financial_rows,
            "unit_hint": unit_hint,
            "unit_multiplier": unit_multiplier,
        }

    # ------------------------------------------------------------------
    # Formula dependency graph summary
    # ------------------------------------------------------------------

    @staticmethod
    def _build_formula_graph_summary(
        rows: List[Dict[str, Any]],
        sheet_name: str,
    ) -> Dict[str, Any]:
        """Build a lightweight formula dependency summary for a sheet.

        Returns counts and samples of subtotal SUM formulas and cross-sheet
        references.  Does NOT build a full dependency graph.
        """
        formula_count = 0
        subtotal_formulas: List[Dict[str, str]] = []
        cross_sheet_refs: List[Dict[str, str]] = []
        sample_formulas: List[str] = []

        for row in rows:
            for cell in row.get("cells", []):
                formula = cell.get("formula")
                if not formula:
                    continue

                formula_count += 1

                if len(sample_formulas) < 5:
                    sample_formulas.append(f"{cell['ref']}={formula}")

                # Detect subtotal SUM formulas (same column)
                ref_col_m = re.match(r"([A-Z]+)", cell.get("ref", ""))
                if ref_col_m:
                    m = _SUM_PATTERN.match(formula)
                    if m and m.group(1).upper() == m.group(3).upper() == ref_col_m.group(1):
                        subtotal_formulas.append(
                            {
                                "cell": cell["ref"],
                                "range": f"{m.group(1)}{m.group(2)}:{m.group(3)}{m.group(4)}",
                            }
                        )

                # Detect cross-sheet references
                for r in cell.get("references", []):
                    if "!" in r:
                        cross_sheet_refs.append(
                            {
                                "from_cell": cell["ref"],
                                "to_ref": r,
                            }
                        )

        return {
            "formula_count": formula_count,
            "subtotal_count": len(subtotal_formulas),
            "subtotal_samples": subtotal_formulas[:5],
            "cross_sheet_ref_count": len(cross_sheet_refs),
            "cross_sheet_samples": cross_sheet_refs[:5],
            "sample_formulas": sample_formulas,
        }

    # ------------------------------------------------------------------
    # Markdown serialisation for Claude
    # ------------------------------------------------------------------

    @staticmethod
    def _structured_to_markdown(structured: Dict[str, Any]) -> str:
        """
        Convert the structured dict to a column-aligned markdown table
        suitable for sending to Claude.

        Each sheet becomes a ``## Sheet: <name>`` section.  Merged regions
        are listed, then a pipe-delimited table whose columns use actual
        Excel column letters (A, B, C, …) so that every row lines up with
        the header row.  A trailing ``Bold`` column flags section headers.

        Hidden sheets are prefixed with ``[HIDDEN]``.  A format annotation
        row (``| Fmt | ... |``) is emitted when columns have a dominant
        number format (``%``, ``$``, or ``#``).

        Values are rendered with full precision to avoid rounding errors.
        """
        parts: List[str] = []

        for sheet in structured.get("sheets", []):
            # Sheet header with visibility annotation
            visibility = sheet.get("visibility", "visible")
            if visibility == "veryHidden":
                header = f"## Sheet: [VERY HIDDEN] {sheet['sheet_name']}"
            elif sheet.get("is_hidden"):
                header = f"## Sheet: [HIDDEN] {sheet['sheet_name']}"
            else:
                header = f"## Sheet: {sheet['sheet_name']}"
            parts.append(header)

            if sheet.get("merged_regions"):
                parts.append("Merged: " + ", ".join(sheet["merged_regions"]))

            # Emit messy-sheet metadata annotations (WS-3)
            label_col = sheet.get("label_column")
            if label_col and label_col != "A":
                parts.append(f"Labels: column {label_col}")
            header_row = sheet.get("header_row_index")
            if header_row is not None and header_row != 1:
                parts.append(f"Header row: {header_row}")
            if sheet.get("unit_hint"):
                parts.append(f"Units: {sheet['unit_hint']}")
            if sheet.get("is_transposed"):
                parts.append("Layout: TRANSPOSED (periods in rows, labels in columns)")
            table_regions = sheet.get("table_regions", [])
            if len(table_regions) > 1:
                regions_str = ", ".join(f"{r['start_row']}-{r['end_row']}" for r in table_regions)
                parts.append(f"Table regions: {regions_str}")

            # Period detection annotation
            detected = sheet.get("detected_periods")
            if detected and detected.get("periods"):
                period_parts = [
                    f"{p['column_letter']}->{p['normalized']}" for p in detected["periods"]
                ]
                parts.append(
                    f"Periods: {', '.join(period_parts)} "
                    f"({detected['dominant_type']}, "
                    f"confidence={detected['confidence']:.2f})"
                )

            rows = sheet.get("rows", [])
            if not rows:
                parts.append("_(empty sheet)_")
                parts.append("")
                continue

            # Collect every column letter that appears in any row so the
            # table is aligned by actual Excel column, not cell position.
            all_columns: set = set()
            for row in rows:
                for cell in row["cells"]:
                    ref = cell.get("ref", "")
                    m = re.match(r"([A-Z]+)", ref)
                    if m:
                        all_columns.add(m.group(1))

            # Sort in Excel order: A < B < … < Z < AA < AB …
            sorted_columns = sorted(all_columns, key=lambda c: (len(c), c))
            col_index = {c: i for i, c in enumerate(sorted_columns)}

            # Detect column-level number format patterns
            format_hints: Dict[str, str] = {}
            for col in sorted_columns:
                fmt_counts: Dict[str, int] = {
                    "$": 0,
                    "%": 0,
                    "x": 0,
                    "total": 0,
                }
                for row in rows:
                    for cell in row["cells"]:
                        ref = cell.get("ref", "")
                        cm = re.match(r"([A-Z]+)", ref)
                        if not cm or cm.group(1) != col:
                            continue
                        nf = cell.get("number_format", "General")
                        val = cell.get("value")
                        if val is not None and not isinstance(val, str):
                            fmt_counts["total"] += 1
                            if "%" in nf:
                                fmt_counts["%"] += 1
                            elif "$" in nf or "#,##0" in nf:
                                fmt_counts["$"] += 1
                            elif "0.0" in nf and "%" not in nf:
                                fmt_counts["x"] += 1
                if fmt_counts["total"] > 0:
                    for sym in ["%", "$", "x"]:
                        if fmt_counts[sym] / fmt_counts["total"] > 0.5:
                            format_hints[col] = sym
                            break

            # Header row (Type column between data columns and Bold)
            col_headers = ["Row"] + sorted_columns + ["Type", "Bold", "Indent", "Formula"]
            parts.append("| " + " | ".join(col_headers) + " |")
            parts.append("|" + "|".join(["---"] * len(col_headers)) + "|")

            # Format annotation row (only if any column has a dominant format)
            if format_hints:
                fmt_values = ["Fmt"]
                for col in sorted_columns:
                    fmt_values.append(format_hints.get(col, ""))
                fmt_values.extend(["", "", "", ""])  # Type, Bold, Indent, Formula empty
                parts.append("| " + " | ".join(fmt_values) + " |")

            # Build set of section boundary row indices for separator insertion
            section_rows = {b["row_index"] for b in sheet.get("section_boundaries", [])}

            for row in rows:
                cells = row["cells"]
                row_idx = str(row["row_index"])

                # Section separator before section-header rows
                if row["row_index"] in section_rows:
                    parts.append("|" + "|".join(["---"] * len(col_headers)) + "|")

                # Mark non-financial rows (WS-3)
                nf_rows = sheet.get("non_financial_rows", set())

                # Initialise every column to empty
                cell_values: List[str] = [""] * len(sorted_columns)
                bold_str = "NF" if row["row_index"] in nf_rows else "N"
                max_indent = 0
                first_formula = ""
                cell_types_seen: set = set()

                for cell in cells:
                    ref = cell.get("ref", "")
                    m = re.match(r"([A-Z]+)", ref)
                    if not m:
                        continue
                    idx = col_index.get(m.group(1))
                    if idx is None:
                        continue

                    val = cell.get("value")
                    if val is None:
                        cell_values[idx] = ""
                    elif isinstance(val, float):
                        import math

                        if not math.isfinite(val):
                            cell_values[idx] = str(val)
                        else:
                            nf = cell.get("number_format", "General")
                            if "%" in nf:
                                cell_values[idx] = f"{val:.4%}"
                            else:
                                # Full precision: integer-valued floats as int,
                                # otherwise Python's default repr.
                                if val == int(val):
                                    cell_values[idx] = str(int(val))
                                else:
                                    cell_values[idx] = str(val)
                    else:
                        cell_values[idx] = str(val)

                    if cell.get("is_bold") and bold_str != "NF":
                        bold_str = "Y"

                    indent = cell.get("indent_level", 0)
                    if indent > max_indent:
                        max_indent = indent

                    if not first_formula and cell.get("formula"):
                        first_formula = cell["formula"]

                    cell_types_seen.add(cell.get("cell_type", "value"))

                # Derive row-level type abbreviation (priority: I > F > L > V)
                if "input" in cell_types_seen:
                    type_str = "I"
                elif "formula" in cell_types_seen:
                    type_str = "F"
                elif "label" in cell_types_seen:
                    type_str = "L"
                else:
                    type_str = "V"

                row_parts = (
                    [row_idx] + cell_values + [type_str, bold_str, str(max_indent), first_formula]
                )
                parts.append("| " + " | ".join(row_parts) + " |")

            # Comment footnotes
            footnotes: List[str] = []
            for row in rows:
                for cell in row.get("cells", []):
                    comment = cell.get("comment")
                    if comment:
                        footnotes.append(
                            f"[{cell.get('ref', '?')}]: {comment.replace(chr(10), ' ')}"
                        )
            if footnotes:
                parts.append("")
                parts.append("**Notes:**")
                for fn in footnotes:
                    parts.append(f"- {fn}")

            parts.append("")  # blank line between sheets

        # Named ranges summary
        named = structured.get("named_ranges", {})
        if named:
            parts.append("## Named Ranges")
            for name, target in named.items():
                parts.append(f"- **{name}**: `{target}`")
            parts.append("")

        return "\n".join(parts)


# Self-register at import time
from src.extraction.registry import registry  # noqa: E402

registry.register(ParsingStage())
